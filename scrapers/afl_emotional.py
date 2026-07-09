"""
lib/scraper/afl_emotional.py

Generates emotional flags for the upcoming AFL round.

Data sources:
  - BettingEngine/outputs/afl_round_prep/rNN_YYYY/fixture_rNN_YYYY.csv  — fixture
  - data/afl/injuries/processed/latest-injuries.json                     — injuries
  - data/afl/team-news/latest.json                                        — team news
  - ANTHROPIC_API_KEY env var (from .env.local)

Deterministic:
  - rivalry_derby: fixture vs known AFL rivalry pair

Claude layer:
  - milestone, star_return, must_win, new_coach, farewell, personal_tragedy

Output:
  data/afl/emotional/raw/YYYY/round-N.json
  data/afl/emotional/processed/YYYY/round-N.json
  data/afl/emotional/processed/latest-emotional.json
  data/afl/emotional/logs/scrape.log

Usage:
  uv run --with anthropic python lib/scraper/afl_emotional.py --round 11
  uv run --with anthropic python lib/scraper/afl_emotional.py --round 0    # auto-detect
  uv run --with anthropic python lib/scraper/afl_emotional.py --round 11 --dry-run
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import sys
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

import anthropic

ROOT      = Path(__file__).resolve().parents[1]
BASE_DIR  = ROOT / "data" / "afl" / "emotional"

def _load_env() -> None:
    env_file = ROOT / '.env.local'
    if not env_file.exists():
        return
    for line in env_file.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            if k and k not in os.environ:
                os.environ[k] = v

_load_env()
RAW_DIR   = BASE_DIR / "raw"
PROC_DIR  = BASE_DIR / "processed"
LOG_DIR   = BASE_DIR / "logs"
LOG_PATH  = LOG_DIR / "scrape.log"

BETENGINE = ROOT / "BettingEngine"

VALID_FLAG_TYPES = {
    'milestone', 'new_coach', 'star_return',
    'farewell', 'personal_tragedy', 'rivalry_derby', 'must_win',
    'shame_blowout', 'losing_streak',
}
VALID_STRENGTHS = {'minor', 'normal', 'major'}

# Known AFL rivalry pairs — both orderings are checked
RIVALRIES: list[frozenset[str]] = [
    frozenset({'Collingwood Magpies', 'Carlton Blues'}),
    frozenset({'Collingwood Magpies', 'Richmond Tigers'}),
    frozenset({'Collingwood Magpies', 'Essendon Bombers'}),
    frozenset({'Carlton Blues', 'Essendon Bombers'}),
    frozenset({'Hawthorn Hawks', 'Essendon Bombers'}),           # Freeway Derby
    frozenset({'Richmond Tigers', 'Essendon Bombers'}),          # MCG rivals
    frozenset({'Hawthorn Hawks', 'Richmond Tigers'}),            # MCG rivals
    frozenset({'Melbourne Demons', 'Richmond Tigers'}),          # MCG rivals
    frozenset({'Geelong Cats', 'Collingwood Magpies'}),
    frozenset({'Geelong Cats', 'Hawthorn Hawks'}),
    frozenset({'Geelong Cats', 'Sydney Swans'}),                 # modern grand final rivalry
    frozenset({'Fremantle Dockers', 'West Coast Eagles'}),       # WA Derby
    frozenset({'Port Adelaide Power', 'Adelaide Crows'}),        # The Showdown
    frozenset({'Sydney Swans', 'Greater Western Sydney Giants'}),# Sydney Derby
    frozenset({'Brisbane Lions', 'Gold Coast Suns'}),            # Queensland Derby
    frozenset({'Melbourne Demons', 'Collingwood Magpies'}),
    frozenset({'North Melbourne Kangaroos', 'Collingwood Magpies'}),
    frozenset({'St Kilda Saints', 'Collingwood Magpies'}),
    frozenset({'Western Bulldogs', 'Collingwood Magpies'}),
]

log = logging.getLogger(__name__)

SYSTEM_PROMPT = """\
You are the AFL emotional context analyst for a sports pricing engine.

Your job is to identify MEANINGFUL emotional/motivational storylines that will materially
affect a team's performance in the upcoming round. A typical round has 2–5 flags.
Be thorough — missing a real flag is worse than including a borderline one.

You will receive:
1. The upcoming round fixture
2. Current injury/suspension/team news
3. Recent form (last 2 results per team including margin)
4. Auto-detected flags (rivalry_derby) already confirmed
5. Recent AFL news headlines

Your task: identify any ADDITIONAL flags from these types:
  milestone        — 100th/150th/200th/250th/300th AFL game, debut, first game as captain
  new_coach        — first game under a newly appointed head coach
  star_return      — elite/key player returning from 3+ weeks injured
  farewell         — confirmed retirement game, player's last season announcement
  personal_tragedy — team rallying around a recent death, serious illness, family tragedy
  must_win         — team needs a win to stay in finals contention, OR 3+ game losing
                     streak with finals implications
  shame_blowout    — team lost by 40+ points last week and must respond; flag the LOSING
                     team only (they feel the shame, not the winner)
  losing_streak    — team on 3+ consecutive losses heading into this game

Flag strength guide:
  major  — loss by 60+, 3+ game streak, captain tragedy, elimination game
  normal — loss by 40-59, 2-game streak, milestone 200th+, must-win but not eliminated
  minor  — loss by 30-39, borderline must-win, star return from shorter absence

For each flag return a JSON object:
  {
    "season": <int>,
    "round": <int>,
    "team": "<full team name>",
    "flag_type": "<one of the valid types>",
    "flag_strength": "<minor|normal|major>",
    "player_name": "<name or null>",
    "notes": "<1-2 sentence explanation>",
    "confidence": "<high|medium|low>"
  }

Return a JSON array. Return [] if genuinely nothing applies.
Do not repeat flags already in the auto-detected list.
Use full canonical AFL team names (e.g. "Collingwood Magpies" not "Magpies").
"""


def setup_logging() -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        handlers=[
            logging.FileHandler(LOG_PATH, encoding='utf-8'),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )


def load_json(path: Path) -> list | dict | None:
    if not path.exists():
        log.warning('File not found: %s', path)
        return None
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception as exc:
        log.warning('Failed to parse %s: %s', path, exc)
        return None


def auto_detect_round() -> int | None:
    """Find the highest round number in BettingEngine afl_round_prep outputs."""
    prep_base = BETENGINE / 'outputs' / 'afl_round_prep'
    if not prep_base.exists():
        return None
    dirs = [
        d for d in prep_base.iterdir()
        if d.is_dir() and re.match(r'^r(\d+)_\d{4}$', d.name)
    ]
    if not dirs:
        return None
    dirs.sort(key=lambda d: int(re.match(r'^r(\d+)_', d.name).group(1)))
    latest = dirs[-1]
    m = re.match(r'^r(\d+)_', latest.name)
    return int(m.group(1)) if m else None


def load_fixture(round_number: int, season: int) -> list[dict]:
    """Read fixture CSV from BettingEngine round prep output."""
    fixture_path = BETENGINE / 'outputs' / 'afl_round_prep' / f'r{round_number}_{season}' / f'fixture_r{round_number}_{season}.csv'
    if not fixture_path.exists():
        log.warning('AFL fixture not found: %s', fixture_path)
        return []
    games = []
    with open(fixture_path, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            games.append({
                'home_team':     row.get('home_team', ''),
                'away_team':     row.get('away_team', ''),
                'venue':         row.get('venue', ''),
                'kickoff_local': row.get('date', ''),
            })
    log.info('Loaded %d games from fixture CSV', len(games))
    return games


def fetch_afl_news(round_number: int, season: int, teams: set[str] | None = None,
                   max_age_days: int = 10) -> str:
    """
    Fetch recent AFL news headlines from Google News RSS.

    Headlines older than max_age_days are DROPPED and every kept headline is
    prefixed with its publish date. Google News resurfaces months-old stories
    for query matches — on 2026-07-09 an April "captain ruled out after
    brother's death" story came back for an R18 query and nearly put a bogus
    major personal_tragedy flag (+2.5) on Adelaide's line.

    teams: full team names of clubs playing this round. Each gets its own news
    query — the three generic queries alone missed the NRL R19 Jai Arrow MND
    tribute game (2026-07-09); the same blind spot existed here.
    """
    from email.utils import parsedate_to_datetime

    queries = [
        (f'AFL+Round+{round_number}+{season}', 10),
        (f'AFL+{season}+milestone+farewell+emotional', 10),
        (f'AFL+{season}+personal+tragedy+must+win', 10),
    ]
    for team in sorted(teams or []):
        nickname = team.split()[-1]
        queries.append((f'AFL+{nickname}+news', 4))
    headlines: list[str] = []
    seen: set[str] = set()
    skipped_old = 0
    for query, per_query_limit in queries:
        url = f'https://news.google.com/rss/search?q={query}&hl=en-AU&gl=AU&ceid=AU:en'
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                root = ET.fromstring(resp.read())
            for item in root.findall('.//item')[:per_query_limit]:
                title = item.findtext('title', '').split(' - ')[0].strip()
                pub_dt = None
                try:
                    pub_raw = item.findtext('pubDate', '')
                    pub_dt = parsedate_to_datetime(pub_raw) if pub_raw else None
                except (TypeError, ValueError):
                    pass
                if pub_dt is not None:
                    if (datetime.now(timezone.utc) - pub_dt).days > max_age_days:
                        skipped_old += 1
                        continue
                    date_label = pub_dt.strftime('%Y-%m-%d')
                else:
                    date_label = 'undated'
                if title and title not in seen:
                    seen.add(title)
                    headlines.append(f'- [{date_label}] {title}')
        except Exception as exc:
            log.warning('News fetch failed for query "%s": %s', query, exc)
    log.info('Fetched %d AFL news headlines (%d dropped as older than %d days)',
             len(headlines), skipped_old, max_age_days)
    return '\n'.join(headlines) if headlines else '(no news fetched)'


def detect_rivalry_derby(fixture: list[dict]) -> list[dict]:
    flags = []
    for game in fixture:
        home = game.get('home_team', '')
        away = game.get('away_team', '')
        pair = frozenset({home, away})
        if pair in RIVALRIES:
            flags.append({
                'flag_type':     'rivalry_derby',
                'flag_strength': 'normal',
                'team':          home,
                'player_name':   None,
                'notes':         f'{home} vs {away} is a recognised AFL rivalry fixture.',
                'source':        'fixture_rivalry',
            })
            log.info('rivalry_derby detected: %s vs %s', home, away)
    return flags


def load_recent_form(fixture: list[dict], round_number: int) -> str:
    """Pull last 2 results per team from AFL history xlsx."""
    try:
        from openpyxl import load_workbook
        from datetime import datetime as _dt
        xlsx = BETENGINE / 'outputs' / 'afl_weekly_review' / 'historical' / 'latest.xlsx'
        if not xlsx.exists():
            return '  (history not available)'
        wb = load_workbook(xlsx, read_only=True)
        ws = wb.active
        games = []
        for row in ws.iter_rows(values_only=True):
            if not isinstance(row[0], _dt):
                continue
            if row[2] is None or row[3] is None:
                continue
            try:
                hs = int(row[5]) if row[5] is not None else None
                aws = int(row[6]) if row[6] is not None else None
            except (TypeError, ValueError):
                hs, aws = None, None
            games.append({
                'date': row[0].date(),
                'home': str(row[2]).strip(),
                'away': str(row[3]).strip(),
                'home_score': hs,
                'away_score': aws,
            })
        games.sort(key=lambda g: g['date'])

        # Short name → full name mapping for lookup
        SHORT_TO_FULL = {
            'Adelaide':        'Adelaide Crows',
            'Brisbane':        'Brisbane Lions',
            'Carlton':         'Carlton Blues',
            'Collingwood':     'Collingwood Magpies',
            'Essendon':        'Essendon Bombers',
            'Fremantle':       'Fremantle Dockers',
            'Geelong':         'Geelong Cats',
            'Gold Coast':      'Gold Coast Suns',
            'GWS Giants':      'Greater Western Sydney Giants',
            'Hawthorn':        'Hawthorn Hawks',
            'Melbourne':       'Melbourne Demons',
            'North Melbourne': 'North Melbourne Kangaroos',
            'Port Adelaide':   'Port Adelaide Power',
            'Richmond':        'Richmond Tigers',
            'St Kilda':        'St Kilda Saints',
            'Sydney':          'Sydney Swans',
            'West Coast':      'West Coast Eagles',
            'Western Bulldogs':'Western Bulldogs',
        }

        all_teams = set()
        for g in fixture:
            all_teams.add(g.get('home_team', ''))
            all_teams.add(g.get('away_team', ''))

        lines = []
        for full_name in sorted(all_teams):
            if not full_name:
                continue
            # find short name
            short = next((s for s, f in SHORT_TO_FULL.items() if f == full_name), full_name)
            team_games = [
                g for g in games
                if g['home'] == short or g['away'] == short
            ][-2:]
            if not team_games:
                continue
            results = []
            for g in team_games:
                if g['home_score'] is None:
                    continue
                if g['home'] == short:
                    margin = g['home_score'] - g['away_score']
                    result = 'W' if margin > 0 else ('L' if margin < 0 else 'D')
                    results.append(f"{result} {'+' if margin >= 0 else ''}{margin} vs {SHORT_TO_FULL.get(g['away'], g['away'])}")
                else:
                    margin = g['away_score'] - g['home_score']
                    result = 'W' if margin > 0 else ('L' if margin < 0 else 'D')
                    results.append(f"{result} {'+' if margin >= 0 else ''}{margin} vs {SHORT_TO_FULL.get(g['home'], g['home'])}")
            if results:
                lines.append(f"  {full_name}: {' | '.join(results)}")
        return '\n'.join(lines) if lines else '  (none)'
    except Exception as exc:
        return f'  (form load failed: {exc})'


def build_claude_prompt(
    season: int,
    round_number: int,
    fixture: list[dict],
    injuries: list[dict],
    team_news: dict,
    auto_flags: list[dict],
    news_headlines: str = '',
) -> str:
    fixture_lines = '\n'.join(
        f"  {g.get('home_team')} vs {g.get('away_team')}  ({g.get('kickoff_local', '')[:10]})"
        for g in fixture
    ) or '  (no fixture loaded)'

    # Injuries grouped by team
    by_team: dict[str, list[str]] = {}
    for inj in injuries:
        team = inj.get('team', 'Unknown')
        line = f"  [{inj.get('status','?')}] {inj.get('player','?')} — {inj.get('notes','')}"
        by_team.setdefault(team, []).append(line)

    injury_section = ''
    for team, lines in sorted(by_team.items()):
        injury_section += f"\n{team}:\n" + '\n'.join(lines)

    # Team news
    news_section = ''
    if isinstance(team_news, dict):
        teams_data = team_news.get('teams', team_news)
        for team, data in sorted(teams_data.items()):
            items = data.get('items', []) if isinstance(data, dict) else []
            if items:
                news_section += f"\n{team}:\n"
                for item in items:
                    news_section += f"  [{item.get('type','?')}] {item.get('player','?')} — {item.get('detail','')}\n"

    auto_section = (
        json.dumps(auto_flags, indent=2, ensure_ascii=False)
        if auto_flags else '  (none detected)'
    )

    form_section = load_recent_form(fixture, round_number)

    return f"""UPCOMING ROUND: Season {season}, Round {round_number}

FIXTURE:
{fixture_lines}

RECENT FORM (last 2 results per team — margin from team's perspective, positive = win):
{form_section}

INJURY LIST:
{injury_section or '  (none loaded)'}

TEAM NEWS:
{news_section or '  (none loaded)'}

RECENT AFL NEWS HEADLINES (each prefixed with its publish date — DISREGARD any event \
that clearly happened weeks or months ago even if the headline resurfaced; only flag \
situations that affect THIS round's games):
{news_headlines or '  (none fetched)'}

AUTO-DETECTED EMOTIONAL FLAGS (already confirmed — do NOT repeat these):
{auto_section}

Based on the above, identify ADDITIONAL emotional flags for Round {round_number}.
Pay particular attention to: teams off 40+ point losses (shame_blowout), teams on 3+ losing streaks, \
must-win finals situations, milestones in the news headlines.

Return a JSON array only. No prose. No markdown. No explanation outside the JSON.
"""


def call_claude(prompt: str, dry_run: bool) -> list[dict]:
    if dry_run:
        log.info('DRY RUN — skipping Claude API call')
        return []

    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        log.warning('ANTHROPIC_API_KEY not set — skipping Claude analysis')
        return []

    client = anthropic.Anthropic(api_key=api_key)
    try:
        msg = client.messages.create(
            model='claude-opus-4-7',
            max_tokens=1024,
            system=SYSTEM_PROMPT,
            messages=[{'role': 'user', 'content': prompt}],
        )
        raw = msg.content[0].text.strip()
        log.info('Claude response: %s', raw[:200])

        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)

        result = json.loads(raw)
        if isinstance(result, list):
            return result
        log.warning('Claude returned non-list: %s', type(result))
        return []
    except json.JSONDecodeError as exc:
        log.warning('Claude response not valid JSON: %s', exc)
        return []
    except Exception as exc:
        log.warning('Claude API call failed: %s', exc)
        return []


def validate_flag(flag: dict, season: int, round_number: int, playing_teams: set[str] | None = None) -> dict | None:
    ftype    = str(flag.get('flag_type', '')).strip().lower()
    strength = str(flag.get('flag_strength', 'normal')).strip().lower()
    team     = str(flag.get('team', '')).strip()

    if not team:
        return None
    if ftype not in VALID_FLAG_TYPES:
        log.warning('Invalid flag_type "%s" — skipping', ftype)
        return None
    if strength not in VALID_STRENGTHS:
        strength = 'normal'
    if playing_teams and team not in playing_teams:
        log.warning('Flag for "%s" rejected — team not in fixture this round (bye?)', team)
        return None

    return {
        'season':        season,
        'round':         round_number,
        'team':          team,
        'flag_type':     ftype,
        'flag_strength': strength,
        'player_name':   flag.get('player_name') or None,
        'notes':         flag.get('notes') or None,
    }


def write_outputs(flags: list[dict], raw_payload: dict, season: int, round_number: int) -> None:
    scraped_at = datetime.now(timezone.utc).isoformat()

    raw_year_dir = RAW_DIR / str(season)
    raw_year_dir.mkdir(parents=True, exist_ok=True)
    (raw_year_dir / f'round-{round_number}.json').write_text(
        json.dumps(raw_payload, indent=2, ensure_ascii=False), encoding='utf-8',
    )

    proc_year_dir = PROC_DIR / str(season)
    proc_year_dir.mkdir(parents=True, exist_ok=True)
    (proc_year_dir / f'round-{round_number}.json').write_text(
        json.dumps(flags, indent=2, ensure_ascii=False), encoding='utf-8',
    )

    if flags:
        PROC_DIR.mkdir(parents=True, exist_ok=True)
        (PROC_DIR / 'latest-emotional.json').write_text(
            json.dumps({
                'sport':      'AFL',
                'season':     season,
                'round':      round_number,
                'scraped_at': scraped_at,
                'flags':      flags,
            }, indent=2, ensure_ascii=False),
            encoding='utf-8',
        )
        log.info('Wrote latest-emotional.json — %d flags, R%d', len(flags), round_number)
    else:
        log.info('No flags generated — latest-emotional.json NOT overwritten')


def run(season: int, round_number: int, dry_run: bool) -> int:
    fixture   = load_fixture(round_number, season)
    injuries  = load_json(ROOT / 'data' / 'afl' / 'injuries' / 'processed' / 'latest-injuries.json') or []
    team_news = load_json(ROOT / 'data' / 'afl' / 'team-news' / 'latest.json') or {}
    playing_teams = {
        t for g in fixture for t in (g.get('home_team'), g.get('away_team')) if t
    }
    news_headlines = fetch_afl_news(round_number, season, playing_teams)

    if not fixture:
        log.warning('No fixture data — emotional flags may be incomplete')

    auto_flags = detect_rivalry_derby(fixture)
    log.info('Auto-detected: %d rivalry_derby', len(auto_flags))

    prompt = build_claude_prompt(season, round_number, fixture, injuries, team_news, auto_flags, news_headlines)
    claude_flags = call_claude(prompt, dry_run)
    log.info('Claude returned %d additional flags', len(claude_flags))

    playing_teams: set[str] = set()
    for g in fixture:
        playing_teams.add(g.get('home_team', ''))
        playing_teams.add(g.get('away_team', ''))
    playing_teams.discard('')

    all_raw = auto_flags + claude_flags
    validated: list[dict] = []
    for flag in all_raw:
        v = validate_flag(flag, season, round_number, playing_teams or None)
        if v:
            validated.append(v)

    seen: set[tuple] = set()
    deduped: list[dict] = []
    for v in validated:
        key = (v['team'], v['flag_type'], v.get('player_name'))
        if key not in seen:
            seen.add(key)
            deduped.append(v)

    log.info('Final: %d validated flags after dedup', len(deduped))

    if not dry_run:
        write_outputs(deduped, {
            'season': season, 'round': round_number,
            'auto_flags': auto_flags, 'claude_flags': claude_flags,
            'prompt': prompt,
        }, season, round_number)
    else:
        log.info('DRY RUN — output:')
        print(json.dumps(deduped, indent=2, ensure_ascii=False))

    if deduped:
        print(f'\n  AFL Emotional flags for R{round_number}:')
        for f in deduped:
            player = f' — {f["player_name"]}' if f.get('player_name') else ''
            print(f'  [{f["flag_type"]:<18}] {f["flag_strength"]:<6}  {f["team"]}{player}')
            if f.get('notes'):
                print(f'    {f["notes"]}')
    else:
        print(f'\n  No AFL emotional flags for R{round_number} this week.')

    return len(deduped)


def main() -> None:
    setup_logging()
    p = argparse.ArgumentParser(description='Generate emotional flags for AFL round')
    p.add_argument('--season',  type=int, default=2026)
    p.add_argument('--round',   type=int, required=True, dest='round_number',
                   help='Round to generate flags for (0 = auto-detect)')
    p.add_argument('--dry-run', action='store_true',
                   help='Print output without writing files or calling Claude')
    args = p.parse_args()

    round_number = args.round_number
    if round_number == 0:
        round_number = auto_detect_round()
        if not round_number:
            log.error('Could not auto-detect AFL round — pass --round explicitly')
            sys.exit(1)
        log.info('Auto-detected round: %d', round_number)

    log.info('Generating AFL emotional flags: season=%d round=%d', args.season, round_number)
    count = run(args.season, round_number, args.dry_run)
    sys.exit(0 if count >= 0 else 1)


if __name__ == '__main__':
    main()
