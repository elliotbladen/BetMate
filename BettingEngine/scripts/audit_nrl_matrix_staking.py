#!/usr/bin/env python3
"""Audit 2026 logged NRL H2H/handicap bets against the existing matrices.

The matrices are treated as a separate confidence layer. They are never used to
recreate a price or to select a bet after seeing its result.

The script reads the BetMate actual-bet ledger, ignores totals and non-NRL
markets, maps eligible bets to the 2026 fixture in model.db, then evaluates
the *frozen 2022-25* H2H and handicap matrices that were used during 2026.

Usage:
    python scripts/audit_nrl_matrix_staking.py

Outputs:
    outputs/results/nrl_2026_actual_bets_matrix_staking_audit.md
    outputs/results/nrl_2026_actual_bets_matrix_staking_audit.csv
"""

from __future__ import annotations

import csv
import re
import sqlite3
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BETMATE = ROOT.parent
LEDGER = BETMATE / "lib" / "researchData.ts"
DB = ROOT / "data" / "model.db"
H2H_MATRIX = ROOT / "outputs" / "nrl_h2h_matrix.xlsx"
HANDICAP_MATRIX = ROOT / "outputs" / "nrl_handicap_matrix.csv"
OUT_DIR = ROOT / "outputs" / "results"

TEAM_ALIASES = {
    "Brisbane Broncos": ["brisbane broncos", "broncos", "brisbane"],
    "Canberra Raiders": ["canberra raiders", "canberra", "raiders"],
    "Canterbury Bulldogs": ["canterbury bulldogs", "canterbury", "bulldogs"],
    "Cronulla Sharks": ["cronulla sharks", "cronulla", "sharks"],
    "Dolphins": ["dolphins"],
    "Gold Coast Titans": ["gold coast titans", "titans", "gold coast"],
    "Manly Sea Eagles": ["manly sea eagles", "sea eagles", "manly"],
    "Melbourne Storm": ["melbourne storm", "storm"],
    "New Zealand Warriors": ["new zealand warriors", "nz warriors", "warriors"],
    "Newcastle Knights": ["newcastle knights", "knights", "newcastle"],
    "North QLD Cowboys": ["north queensland cowboys", "north qld cowboys", "cowboys", "north queensland"],
    "Parramatta Eels": ["parramatta eels", "parramatta", "eels"],
    "Penrith Panthers": ["penrith panthers", "panthers", "penrith"],
    "South Sydney Rabbitohs": ["south sydney rabbitohs", "rabbitohs", "souths"],
    "St George Dragons": ["st george illawarra dragons", "st george dragons", "st george", "dragons"],
    "Sydney Roosters": ["sydney roosters", "roosters"],
    "Wests Tigers": ["wests tigers", "west tigers", "tigers"],
}

DB_TO_MATRIX = {
    "Brisbane Broncos": "Brisbane Broncos",
    "Canberra Raiders": "Canberra Raiders",
    "Canterbury-Bankstown Bulldogs": "Canterbury Bulldogs",
    "Cronulla-Sutherland Sharks": "Cronulla Sharks",
    "Dolphins": "Dolphins",
    "Gold Coast Titans": "Gold Coast Titans",
    "Manly-Warringah Sea Eagles": "Manly Sea Eagles",
    "Melbourne Storm": "Melbourne Storm",
    "New Zealand Warriors": "New Zealand Warriors",
    "Newcastle Knights": "Newcastle Knights",
    "North Queensland Cowboys": "North QLD Cowboys",
    "Parramatta Eels": "Parramatta Eels",
    "Penrith Panthers": "Penrith Panthers",
    "South Sydney Rabbitohs": "South Sydney Rabbitohs",
    "St George Illawarra Dragons": "St George Dragons",
    "St. George Illawarra Dragons": "St George Dragons",
    "Sydney Roosters": "Sydney Roosters",
    "Wests Tigers": "Wests Tigers",
}

MONTHS = {3: "March", 4: "April", 5: "May", 6: "June", 7: "July", 8: "August", 9: "September", 10: "October"}


@dataclass
class Bet:
    ident: int
    date: str
    match: str
    market: str
    odds: float
    result: str
    stake: float


def load_matrix_xlsx(path: Path) -> dict[str, dict[str, tuple[float, str]]]:
    """Read visible matrix rows: team -> category -> (relative edge %, direction)."""
    out: dict[str, dict[str, tuple[float, str]]] = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        rows: dict[str, tuple[float, str]] = {}
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            if not row or not row[0] or len(row) < 5 or not row[4]:
                continue
            match = re.match(r"^([0-9.]+)%\s+(.+)$", str(row[4]).strip())
            if match:
                rows[str(row[0]).strip()] = (float(match.group(1)), match.group(2).strip().lower())
        out[sheet] = rows
    wb.close()
    return out


def load_handicap_matrix(path: Path) -> dict[str, dict[str, tuple[float, str]]]:
    out: dict[str, dict[str, tuple[float, str]]] = {}
    with path.open(encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            try:
                edge = float(row["edge_pct"])
            except (KeyError, TypeError, ValueError):
                continue
            out.setdefault(row["team"].strip(), {})[row["category"].strip()] = (edge, row["direction"].strip().lower())
    return out


def parse_ledger(path: Path) -> list[Bet]:
    """Parse only object literals in the checked-in actual-bet ledger."""
    text = path.read_text(encoding="utf-8-sig")
    bets: list[Bet] = []
    for item in re.finditer(r"\{\s*id:(\d+),(.*?)\}\s*,?", text, re.S):
        ident, body = int(item.group(1)), item.group(2)
        fields = dict(re.findall(r"\b(date|match|market|result|sport|notes):'([^']*)'", body))
        if fields.get("sport") != "NRL" or not fields.get("date", "").startswith("2026-"):
            continue
        odds_match = re.search(r"\bodds:([0-9.]+|null)", body)
        if not odds_match or odds_match.group(1) == "null":
            continue
        stake_match = re.search(r"Stake \$([0-9]+(?:\.[0-9]{1,2})?)", fields.get("notes", ""), re.I)
        # The ledger documents bets before id 326 as one unit. We use $50 for
        # comparison only; it does not alter the incremental double-stake P/L.
        stake = float(stake_match.group(1)) if stake_match else 50.0
        bets.append(Bet(ident, fields["date"], fields.get("match", ""), fields.get("market", ""), float(odds_match.group(1)), fields.get("result", ""), stake))
    return sorted(bets, key=lambda b: (b.date, b.ident))


def mentioned_teams(text: str) -> list[str]:
    low = text.lower()
    found = []
    for team, aliases in TEAM_ALIASES.items():
        if any(re.search(rf"(?<![a-z]){re.escape(alias)}(?![a-z])", low) for alias in aliases):
            found.append(team)
    return found


def selected_team(bet: Bet, fixture_teams: tuple[str, str]) -> str | None:
    mentioned = [t for t in mentioned_teams(bet.market) if t in fixture_teams]
    return mentioned[0] if len(mentioned) == 1 else None


def market_type(market: str) -> str | None:
    low = market.lower()
    if any(word in low for word in ("under", "over", "total", "try", "goal", "multi", "same game")):
        return None
    if re.search(r"[+-]\d+(?:\.\d+)?", market):
        return "handicap"
    if "win" in low or "game" in low or re.search(r"\b\d+(?:\.\d+)?\b", low):
        return "h2h"
    return None


def load_fixtures() -> list[dict]:
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT m.match_date, m.kickoff_datetime, v.venue_name,
               th.team_name AS home_db, ta.team_name AS away_db,
               rh.home_score, rh.away_score
        FROM matches m
        JOIN teams th ON th.team_id = m.home_team_id
        JOIN teams ta ON ta.team_id = m.away_team_id
        JOIN venues v ON v.venue_id = m.venue_id
        LEFT JOIN results rh ON rh.match_id = m.match_id
        WHERE m.sport = 'NRL' AND m.season = 2026
        ORDER BY m.kickoff_datetime
    """).fetchall()
    conn.close()
    return [{
        "date": r["match_date"], "kickoff": r["kickoff_datetime"], "venue": r["venue_name"],
        "home": DB_TO_MATRIX[r["home_db"]], "away": DB_TO_MATRIX[r["away_db"]],
        "hs": r["home_score"], "as": r["away_score"],
    } for r in rows]


def fixture_for_bet(bet: Bet, fixtures: list[dict]) -> dict | None:
    teams = set(mentioned_teams(bet.match))
    # The ledger date is sometimes the placement date (the evening before a
    # Friday/Saturday game), so accept a one-day gap only when the two teams
    # identify one fixture uniquely.
    bet_day = datetime.fromisoformat(bet.date).date()
    candidates = [
        f for f in fixtures
        if abs((datetime.fromisoformat(f["date"]).date() - bet_day).days) <= 1
        and {f["home"], f["away"]}.issubset(teams)
    ]
    return candidates[0] if len(candidates) == 1 else None


def form_context(fixtures: list[dict], game: dict, team: str) -> tuple[int | None, str | None]:
    earlier = [f for f in fixtures if f["kickoff"] < game["kickoff"] and team in (f["home"], f["away"]) and f["hs"] is not None]
    if not earlier:
        return None, None
    prev = earlier[-1]
    rest = (datetime.fromisoformat(game["date"]).date() - datetime.fromisoformat(prev["date"]).date()).days
    team_score = prev["hs"] if prev["home"] == team else prev["as"]
    opp_score = prev["as"] if prev["home"] == team else prev["hs"]
    return rest, "win" if team_score > opp_score else "loss" if team_score < opp_score else "draw"


def applicable_rows(game: dict, team: str, market: str, line: float | None, fixtures: list[dict]) -> list[str]:
    dt = datetime.fromisoformat(game["kickoff"].replace("Z", "+00:00"))
    # Source timestamps are UTC; NRL local match classification only needs the
    # broad evening/day label. Convert to Sydney time without a third-party lib.
    local_hour = (dt.hour + 10 + (1 if 10 <= dt.month or dt.month <= 4 else 0)) % 24
    rest, form = form_context(fixtures, game, team)
    rows = ["Win % — Home" if team == game["home"] and market == "h2h" else
            "Win % — Away" if market == "h2h" else
            "Cover Rate — Home" if team == game["home"] else "Cover Rate — Away"]
    rows.append("Night Games (kick-off ≥ 18:00)" if local_hour >= 18 else "Day Games (kick-off < 18:00)")
    rows.append("Thursday / Friday Games" if dt.weekday() in (3, 4) else "Saturday Games" if dt.weekday() == 5 else "Sunday Games" if dt.weekday() == 6 else "")
    if rest is not None:
        if rest <= 6: rows.append("Short Rest (≤ 6 days)")
        elif rest >= 10: rows.append("Long Rest (≥ 10 days)")
        elif market == "handicap": rows.append("Normal Rest (7-9 days)")
    if form == "win": rows.append("After a Win")
    elif form == "loss": rows.append("After a Loss")
    rows += [MONTHS.get(dt.month, ""), f"vs {game['away'] if team == game['home'] else game['home']}", game["venue"]]
    if market == "handicap" and line is not None:
        rows.append("As Favourite (line < 0)" if line < 0 else "As Underdog  (line > 0)")
        if line <= -9.5: rows.append("Heavy Fav (line ≤ -9.5)")
        elif -9 <= line <= -1: rows.append("Slight Fav (line -1 to -9)")
        elif 1 <= line <= 9: rows.append("Slight Dog (line +1 to +9)")
        elif line >= 9.5: rows.append("Big Dog    (line ≥ +9.5)")
    return [r for r in rows if r]


def direction_supports(direction: str, market: str) -> bool:
    return direction == ("backing" if market == "h2h" else "covers")


def line_from_market(market: str, team: str) -> float | None:
    match = re.search(r"([+-]\d+(?:\.\d+)?)", market)
    return float(match.group(1)) if match else None


def pnl(bet: Bet) -> float:
    if bet.result == "win": return bet.stake * (bet.odds - 1)
    if bet.result == "loss": return -bet.stake
    return 0.0


def main() -> None:
    h2h, handicap = load_matrix_xlsx(H2H_MATRIX), load_handicap_matrix(HANDICAP_MATRIX)
    fixtures, bets = load_fixtures(), parse_ledger(LEDGER)
    records, exclusions = [], []
    for bet in bets:
        kind = market_type(bet.market)
        if not kind:
            exclusions.append((bet, "totals/prop/multi or unsupported market")); continue
        game = fixture_for_bet(bet, fixtures)
        if not game:
            exclusions.append((bet, "fixture not uniquely matched")); continue
        team = selected_team(bet, (game["home"], game["away"]))
        if not team:
            exclusions.append((bet, "selected team not explicit in ledger")); continue
        line = line_from_market(bet.market, team) if kind == "handicap" else None
        matrix = h2h if kind == "h2h" else handicap
        supports, conflicts = [], []
        for category in applicable_rows(game, team, kind, line, fixtures):
            value = matrix.get(team, {}).get(category)
            if not value: continue
            edge, direction = value
            (supports if direction_supports(direction, kind) else conflicts).append((edge, category, direction))
        records.append({"bet": bet, "game": game, "team": team, "kind": kind, "support": supports, "conflict": conflicts})

    # Predefined rule from existing T9 practice: 3+ aligned 20% signals in the
    # bet's relevant market, with no opposing signal of 20%+. This is a
    # retrospective audit only, not proof that the rule should be deployed.
    for r in records:
        r["qualifies_double"] = sum(e >= 20 for e, *_ in r["support"]) >= 3 and not any(e >= 20 for e, *_ in r["conflict"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path = OUT_DIR / "nrl_2026_actual_bets_matrix_staking_audit.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["id", "date", "match", "market", "type", "team", "odds", "stake", "result", "support_10", "support_20", "support_30", "support_40", "conflict_20", "max_support", "qualifies_double", "support_rows", "conflict_rows"])
        for r in records:
            b = r["bet"]; s, c = r["support"], r["conflict"]
            writer.writerow([b.ident, b.date, b.match, b.market, r["kind"], r["team"], b.odds, b.stake, b.result,
                             *[sum(e >= t for e, *_ in s) for t in (10, 20, 30, 40)], sum(e >= 20 for e, *_ in c),
                             max((e for e, *_ in s), default=0), r["qualifies_double"],
                             "; ".join(f"{e:.1f}% {row}" for e, row, _ in s), "; ".join(f"{e:.1f}% {row}" for e, row, _ in c)])

    actual_pnl = sum(pnl(r["bet"]) for r in records)
    doubled = [r for r in records if r["qualifies_double"]]
    double_pnl = actual_pnl + sum(pnl(r["bet"]) for r in doubled)
    threshold = {t: [r for r in records if sum(e >= t for e, *_ in r["support"]) >= 3 and not any(e >= t for e, *_ in r["conflict"])] for t in (10, 20, 30, 40)}
    md_path = OUT_DIR / "nrl_2026_actual_bets_matrix_staking_audit.md"
    with md_path.open("w", encoding="utf-8") as fh:
        fh.write("# 2026 NRL actual bets — H2H/handicap matrix staking audit\n\n")
        fh.write("## Scope\n\n")
        fh.write("- Source: checked-in BetMate actual-bet ledger.\n- Included: 2026 NRL bets with an explicit team and a fixture match.\n- Excluded: totals, props, multis, State of Origin and ledger records whose selected side is unclear.\n- Matrix source: frozen 2022–2025 NRL H2H and handicap matrices; no 2026 results were added to the matrix.\n\n")
        fh.write("## Eligible-bet signal counts\n\n")
        fh.write("| Matrix support rule | Bets | Wins–losses | P&L at actual stake | Increment if doubled |\n|---|---:|---:|---:|---:|\n")
        for t, rows in threshold.items():
            wl = Counter(r["bet"].result for r in rows)
            increment = sum(pnl(r["bet"]) for r in rows)
            fh.write(f"| 3+ aligned {t}% signals; no opposing {t}% signal | {len(rows)} | {wl['win']}–{wl['loss']} | ${increment:+.2f} | ${increment:+.2f} |\n")
        fh.write("\n## Predefined double-stake test\n\n")
        fh.write("Rule: double only where the relevant matrix had **3+ aligned signals at 20%+** and **zero opposing 20%+ signal**. This mirrors the existing T9 confluence convention; it was fixed before calculating the staking result.\n\n")
        fh.write(f"- Eligible bets: **{len(records)}**; excluded/unmatched: **{len(exclusions)}**.\n")
        fh.write(f"- Actual eligible-bet P&L: **${actual_pnl:+.2f}**.\n")
        fh.write(f"- Qualifying double-stake bets: **{len(doubled)}**.\n")
        fh.write(f"- Counterfactual P&L: **${double_pnl:+.2f}**.\n")
        fh.write(f"- Change from doubling: **${double_pnl - actual_pnl:+.2f}**.\n\n")
        fh.write("## Qualifying bets\n\n| Date | Bet | Result | Stake | Support / conflict |\n|---|---|---:|---:|---|\n")
        for r in doubled:
            b = r["bet"]
            signal = ", ".join(f"{e:.0f}% {row}" for e, row, _ in sorted(r["support"], reverse=True) if e >= 20)
            fh.write(f"| {b.date} | {b.market} — {b.match} | {b.result} | ${b.stake:.2f} | {signal} |\n")
        fh.write("\n## Exclusions\n\n")
        reason_counts = Counter(reason for _, reason in exclusions)
        fh.write("Totals are deliberately out of scope. The remaining exclusions lack an unambiguous selected side in the current ledger, so guessing would make the experiment unreliable.\n\n")
        for reason, count in sorted(reason_counts.items()):
            fh.write(f"- {reason}: **{count}**\n")
        fh.write("\n## Important limitation\n\n")
        fh.write("This is a retrospective, small-sample staking experiment—not evidence that the matrix creates a predictive edge. The matrix uses overlapping historical splits, so three signals are not three independent pieces of evidence. Do not deploy doubled stakes from this result alone; test the rule prospectively next season with the rule frozen.\n")
    print(f"Eligible: {len(records)} | excluded: {len(exclusions)} | double qualifiers: {len(doubled)}")
    print(f"Actual P&L: ${actual_pnl:+.2f} | counterfactual: ${double_pnl:+.2f} | change: ${double_pnl - actual_pnl:+.2f}")
    print(f"Wrote {md_path}\nWrote {csv_path}")


if __name__ == "__main__":
    main()
