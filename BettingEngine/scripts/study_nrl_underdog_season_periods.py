#!/usr/bin/env python3
"""NRL regular-season closing-price underdog ROI by season period, 2022-2025."""
from __future__ import annotations

import argparse
import random
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "outputs/nrl_weekly_review/historical/latest.xlsx"
OUTPUT = ROOT / "outputs/studies/nrl_underdog_roi_by_season_period_4_6_8_years.md"
SEASONS = tuple(range(2018, 2026))
ROUND_ANCHORS = {
    2018: date(2018, 3, 8),
    2019: date(2019, 3, 14),
    2021: date(2021, 3, 11),
    2022: date(2022, 3, 10),
    2023: date(2023, 3, 2),
    2024: date(2024, 3, 7),
    2025: date(2025, 3, 6),
}
PERIODS = ("Early (R1-10)", "Origin (R11-20)", "Late (R21-end)")


def round_number(season: int, played: date) -> int:
    if season == 2020:
        # COVID interruption: R1-2 in March, R3 restart on 28 May.
        if played < date(2020, 5, 1):
            return (played - date(2020, 3, 12)).days // 7 + 1
        return 3 + (played - date(2020, 5, 28)).days // 7
    anchor = ROUND_ANCHORS[season]
    round_no = 1 if played < anchor else (played - anchor).days // 7 + 1
    # The 2018 ANZAC Wednesday opened R8 before the normal Thu-Sun boundary.
    if season == 2018 and played == date(2018, 4, 25):
        round_no += 1
    # Standalone representative weekends created a blank premiership week.
    rep_restarts = {
        2018: date(2018, 6, 28),
        2019: date(2019, 6, 27),
        2021: date(2021, 7, 1),
        2022: date(2022, 6, 30),
    }
    if season in rep_restarts and played >= rep_restarts[season]:
        round_no -= 1
    return round_no


def period_for(round_no: int) -> str:
    if round_no <= 10:
        return PERIODS[0]
    if round_no <= 20:
        return PERIODS[1]
    return PERIODS[2]


def load_bets(path: Path) -> tuple[list[dict], int]:
    sheet = load_workbook(path, read_only=True, data_only=True).active
    bets, equal_prices = [], 0
    for raw in sheet.iter_rows(min_row=3, values_only=True):
        played_raw = raw[0]
        if not played_raw or played_raw.year not in SEASONS:
            continue
        if raw[5] is None or raw[6] is None or bool(raw[7]):
            continue
        used_opening_fallback = raw[16] is None or raw[20] is None
        home_odds = raw[16] if raw[16] is not None else raw[13]
        away_odds = raw[20] if raw[20] is not None else raw[17]
        if home_odds is None or away_odds is None:
            continue
        home_odds, away_odds = float(home_odds), float(away_odds)
        if home_odds == away_odds:
            equal_prices += 1
            continue
        underdog_home = home_odds > away_odds
        odds = home_odds if underdog_home else away_odds
        home_score, away_score = int(raw[5]), int(raw[6])
        won = home_score > away_score if underdog_home else away_score > home_score
        played = played_raw.date()
        round_no = round_number(played.year, played)
        bets.append({
            "season": played.year,
            "round": round_no,
            "period": period_for(round_no),
            "odds": odds,
            "won": won,
            "pnl": odds - 1.0 if won else -1.0,
            "used_opening_fallback": used_opening_fallback,
        })
    return bets, equal_prices


def summary(rows: list[dict]) -> dict:
    n = len(rows)
    wins = sum(row["won"] for row in rows)
    profit = sum(row["pnl"] for row in rows)
    return {
        "bets": n,
        "wins": wins,
        "win_rate": wins / n if n else 0,
        "average_odds": sum(row["odds"] for row in rows) / n if n else 0,
        "profit": profit,
        "roi": profit / n if n else 0,
    }


def bootstrap_roi(rows: list[dict], runs: int = 20_000) -> tuple[float, float]:
    rng = random.Random(42)
    pnls = [row["pnl"] for row in rows]
    n = len(pnls)
    samples = sorted(sum(rng.choice(pnls) for _ in range(n)) / n for _ in range(runs))
    return samples[int(runs * 0.025)], samples[int(runs * 0.975)]


def pct(value: float) -> str:
    return f"{value * 100:+.1f}%"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=SOURCE)
    parser.add_argument("--output", type=Path, default=OUTPUT)
    args = parser.parse_args()
    bets, equal_prices = load_bets(args.source)

    lines = [
        "# NRL closing-price underdog ROI by season period — 4, 6 and 8 years", "",
        "Regular season only. One flat $1 bet on the longer-priced H2H team in every game. "
        "Closing price is used where available; the historical workbook's opening-price fallback "
        "is used only where its close is absent. Equal-priced teams are excluded.", "",
        "Periods are fixed before calculation: early R1–10, Origin R11–20, late R21 through the "
        "end of the regular season. Finals are excluded.", "",
        "## Pooled comparison", "",
        "| Window | Period | Bets | Wins | Win rate | Avg dog odds | Profit | ROI | Bootstrap 95% ROI interval |",
        "|---|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    windows = (("4 seasons (2022–25)", tuple(range(2022, 2026))),
               ("6 seasons (2020–25)", tuple(range(2020, 2026))),
               ("8 seasons (2018–25)", tuple(range(2018, 2026))))
    for window_name, window_seasons in windows:
        for period in PERIODS:
            rows = [row for row in bets if row["season"] in window_seasons and row["period"] == period]
            result = summary(rows)
            low, high = bootstrap_roi(rows)
            lines.append(
                f"| {window_name} | {period} | {result['bets']} | {result['wins']} | "
                f"{result['win_rate']:.1%} | {result['average_odds']:.2f} | "
                f"${result['profit']:+.2f} | {pct(result['roi'])} | {pct(low)} to {pct(high)} |"
            )

    lines += [
        "", "## True-closing-price sensitivity", "",
        "This removes every match where either side required the opening-price fallback.", "",
        "| Window | Period | Bets | Profit | ROI |", "|---|---|---:|---:|---:|",
    ]
    for window_name, window_seasons in windows:
        for period in PERIODS:
            rows = [row for row in bets if row["season"] in window_seasons and row["period"] == period and not row["used_opening_fallback"]]
            result = summary(rows)
            lines.append(f"| {window_name} | {period} | {result['bets']} | ${result['profit']:+.2f} | {pct(result['roi'])} |")

    lines += ["", "## ROI by season", "", "| Season | Early | Origin | Late |", "|---:|---:|---:|---:|"]
    for season in SEASONS:
        values = []
        for period in PERIODS:
            rows = [row for row in bets if row["season"] == season and row["period"] == period]
            values.append(pct(summary(rows)["roi"]) if rows else "—")
        lines.append(f"| {season} | {values[0]} | {values[1]} | {values[2]} |")

    lines += [
        "", "## Interpretation", "",
        "The early season is the only profitable pooled period in all three windows, but the apparent "
        "advantage shrinks sharply as older seasons are added and was negative in 2019–2021. Origin "
        "and late-season underdogs lost money in every pooled window.", "",
        "The 2020 season was COVID-disrupted and Origin was played after the NRL season. It remains in "
        "the requested six/eight-year samples, but R11–20 is a middle-season band rather than a true "
        "in-season Origin window for that year.", "",
        "The bootstrap interval is wide because long-priced winners create volatile returns. The "
        "result is descriptive and is not, by itself, proof of a future betting edge. A prospective "
        "rule should be frozen before testing 2026.", "",
        f"Data checks: {len(bets)} qualifying regular-season games; {equal_prices} equal-price games excluded.",
    ]
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("\n".join(lines))
    print(f"\nSaved: {args.output}")


if __name__ == "__main__":
    main()
