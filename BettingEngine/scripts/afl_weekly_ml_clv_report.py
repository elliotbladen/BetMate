#!/usr/bin/env python3
"""Build weekly AFL CLV comparison for normal, ML, and market signals."""

from __future__ import annotations

import argparse
import csv
import re
import sqlite3
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "model.db"
DEFAULT_WORKBOOK = ROOT / "outputs" / "afl_weekly_review" / "historical" / "latest.xlsx"
DEFAULT_REPORT_DIR = ROOT / "outputs" / "afl_weekly_review" / "reports"

AFL_ALIASES = {
    "adelaide": "Adelaide Crows",
    "adelaide crows": "Adelaide Crows",
    "brisbane": "Brisbane Lions",
    "brisbane lions": "Brisbane Lions",
    "carlton": "Carlton Blues",
    "carlton blues": "Carlton Blues",
    "collingwood": "Collingwood Magpies",
    "collingwood magpies": "Collingwood Magpies",
    "essendon": "Essendon Bombers",
    "essendon bombers": "Essendon Bombers",
    "fremantle": "Fremantle Dockers",
    "fremantle dockers": "Fremantle Dockers",
    "geelong": "Geelong Cats",
    "geelong cats": "Geelong Cats",
    "gold coast": "Gold Coast Suns",
    "gold coast suns": "Gold Coast Suns",
    "gws giants": "Greater Western Sydney Giants",
    "greater western sydney giants": "Greater Western Sydney Giants",
    "hawthorn": "Hawthorn Hawks",
    "hawthorn hawks": "Hawthorn Hawks",
    "melbourne": "Melbourne Demons",
    "melbourne demons": "Melbourne Demons",
    "north melbourne": "North Melbourne Kangaroos",
    "north melbourne kangaroos": "North Melbourne Kangaroos",
    "port adelaide": "Port Adelaide Power",
    "port adelaide power": "Port Adelaide Power",
    "richmond": "Richmond Tigers",
    "richmond tigers": "Richmond Tigers",
    "st kilda": "St Kilda Saints",
    "st kilda saints": "St Kilda Saints",
    "sydney": "Sydney Swans",
    "sydney swans": "Sydney Swans",
    "west coast": "West Coast Eagles",
    "west coast eagles": "West Coast Eagles",
    "western bulldogs": "Western Bulldogs",
}


def canon_team(raw: str) -> str:
    key = re.sub(r"\s+", " ", (raw or "").strip().lower())
    return AFL_ALIASES.get(key, raw.strip())


def as_float(value: Any) -> float | None:
    if value in ("", None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    return None


def load_workbook_rows(path: Path, season: int) -> dict[tuple[date, str, str], dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"AFL workbook not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
    headers = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    rows: dict[tuple[date, str, str], dict[str, Any]] = {}
    for values in ws.iter_rows(min_row=3, values_only=True):
        row = dict(zip(headers, values))
        match_date = as_date(row.get("Date"))
        if not match_date or match_date.year != season:
            continue
        home = canon_team(str(row.get("Home Team") or ""))
        away = canon_team(str(row.get("Away Team") or ""))
        rows[(match_date, home, away)] = row
    wb.close()
    return rows


def find_workbook_row(
    workbook_rows: dict[tuple[date, str, str], dict[str, Any]],
    match_date: date | None,
    home: str,
    away: str,
) -> dict[str, Any] | None:
    if match_date is None:
        return None
    exact = workbook_rows.get((match_date, home, away))
    if exact:
        return exact
    # AusSportsBetting AFL dates can be one calendar day earlier than model
    # dates for night games depending on export timezone. Keep team matching exact.
    for delta in (-1, 1):
        row = workbook_rows.get((match_date + timedelta(days=delta), home, away))
        if row:
            return row
    return None


def load_predictions(season: int, rnd: int | None = None) -> list[dict[str, Any]]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    params: list[Any] = [season]
    where = "WHERE season = ?"
    if rnd:
        where += " AND round_number = ?"
        params.append(rnd)
    rows = [dict(r) for r in conn.execute(f"""
        SELECT *
        FROM afl_shadow_predictions
        {where}
        ORDER BY round_number, game_date, id
    """, params)]
    conn.close()
    return rows


def auto_round(season: int, workbook_rows: dict[tuple[date, str, str], dict[str, Any]]) -> int:
    rounds: dict[int, list[dict[str, Any]]] = {}
    for row in load_predictions(season):
        rounds.setdefault(int(row["round_number"]), []).append(row)
    completed: list[int] = []
    for rnd, rows in rounds.items():
        if not rows:
            continue
        ok = True
        for pred in rows:
            market = find_workbook_row(
                workbook_rows,
                as_date(pred["game_date"]),
                canon_team(pred["home_team"]),
                canon_team(pred["away_team"]),
            )
            if not market or market.get("Home Score") is None or market.get("Away Score") is None:
                ok = False
                break
        if ok:
            completed.append(rnd)
    if not completed:
        raise SystemExit(f"No completed AFL round found for {season} with pricing and workbook results")
    return max(completed)


def h2h_selection(home: str, away: str, home_prob: float | None) -> str:
    if home_prob is None:
        return ""
    return home if home_prob >= 0.5 else away


def h2h_market_move(home: str, away: str, row: dict[str, Any]) -> str:
    h_open = as_float(row.get("Home Odds Open"))
    h_close = as_float(row.get("Home Odds Close"))
    a_open = as_float(row.get("Away Odds Open"))
    a_close = as_float(row.get("Away Odds Close"))
    if None in (h_open, h_close, a_open, a_close):
        return ""
    home_move = h_open - h_close
    away_move = a_open - a_close
    if abs(home_move - away_move) < 1e-9:
        return "no_move"
    return home if home_move > away_move else away


def odds_for_selection(row: dict[str, Any], selection: str, home: str, away: str, prefix: str) -> float | None:
    if selection == home:
        return as_float(row.get(f"Home Odds {prefix}"))
    if selection == away:
        return as_float(row.get(f"Away Odds {prefix}"))
    return None


def decimal_clv(open_odds: float | None, close_odds: float | None) -> float | None:
    if not open_odds or not close_odds:
        return None
    return round((open_odds / close_odds) - 1.0, 4)


def h2h_result(selection: str, home: str, away: str, home_score: float, away_score: float) -> str:
    if selection in ("", "no_move"):
        return ""
    if home_score == away_score:
        return "push"
    winner = home if home_score > away_score else away
    return "win" if selection == winner else "loss"


def handicap_pick(model_margin: float | None, home_line: float | None, home: str, away: str) -> str:
    if model_margin is None or home_line is None:
        return ""
    edge = model_margin + home_line
    if abs(edge) < 1e-9:
        return "no_edge"
    return home if edge > 0 else away


def handicap_market_move(home: str, away: str, row: dict[str, Any]) -> str:
    open_line = as_float(row.get("Home Line Open"))
    close_line = as_float(row.get("Home Line Close"))
    if open_line is None or close_line is None:
        return ""
    move = close_line - open_line
    if abs(move) < 1e-9:
        return "no_move"
    return home if move < 0 else away


def line_for_selection(row: dict[str, Any], selection: str, home: str, away: str, prefix: str) -> float | None:
    if selection == home:
        return as_float(row.get(f"Home Line {prefix}"))
    if selection == away:
        return as_float(row.get(f"Away Line {prefix}"))
    return None


def line_odds_for_selection(row: dict[str, Any], selection: str, home: str, away: str, prefix: str) -> float | None:
    if selection == home:
        return as_float(row.get(f"Home Line Odds {prefix}"))
    if selection == away:
        return as_float(row.get(f"Away Line Odds {prefix}"))
    return None


def handicap_clv(selection: str, open_line: float | None, close_line: float | None) -> float | None:
    if selection in ("", "no_edge", "no_move") or open_line is None or close_line is None:
        return None
    return round(open_line - close_line, 2)


def spread_result(selection: str, home: str, away: str, home_score: float, away_score: float, line: float | None) -> str:
    if line is None:
        return ""
    if selection == home:
        adjusted = home_score + line - away_score
    elif selection == away:
        adjusted = away_score + line - home_score
    else:
        return ""
    if abs(adjusted) < 1e-9:
        return "push"
    return "win" if adjusted > 0 else "loss"


def total_pick(model_total: float | None, total_open: float | None) -> str:
    if model_total is None or total_open is None:
        return ""
    edge = model_total - total_open
    if abs(edge) < 1e-9:
        return "no_edge"
    return "over" if edge > 0 else "under"


def total_market_move(row: dict[str, Any]) -> str:
    open_total = as_float(row.get("Total Score Open"))
    close_total = as_float(row.get("Total Score Close"))
    if open_total is None or close_total is None:
        return ""
    move = close_total - open_total
    if abs(move) < 1e-9:
        return "no_move"
    return "over" if move > 0 else "under"


def total_line_clv(selection: str, open_total: float | None, close_total: float | None) -> float | None:
    if selection in ("", "no_edge", "no_move") or open_total is None or close_total is None:
        return None
    if selection == "over":
        return round(close_total - open_total, 2)
    if selection == "under":
        return round(open_total - close_total, 2)
    return None


def total_odds(row: dict[str, Any], selection: str, prefix: str) -> float | None:
    if selection == "over":
        return as_float(row.get(f"Total Score Over {prefix}"))
    if selection == "under":
        return as_float(row.get(f"Total Score Under {prefix}"))
    return None


def total_result(selection: str, actual_total: float, line: float | None) -> str:
    if line is None or selection not in ("over", "under"):
        return ""
    diff = actual_total - line
    if abs(diff) < 1e-9:
        return "push"
    if selection == "over":
        return "win" if diff > 0 else "loss"
    return "win" if diff < 0 else "loss"


def fair_odds(home_prob: float | None) -> tuple[float | None, float | None]:
    if home_prob is None or home_prob <= 0 or home_prob >= 1:
        return None, None
    return round(1 / home_prob, 3), round(1 / (1 - home_prob), 3)


def build_rows(season: int, rnd: int, workbook_rows: dict[tuple[date, str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    predictions = load_predictions(season, rnd)
    if not predictions:
        raise SystemExit(f"No AFL predictions found for season={season} round={rnd}")

    out: list[dict[str, Any]] = []
    for pred in predictions:
        match_date = as_date(pred["game_date"])
        home = canon_team(pred["home_team"])
        away = canon_team(pred["away_team"])
        market = find_workbook_row(workbook_rows, match_date, home, away)
        if not market:
            raise SystemExit(f"No workbook row found for {match_date} {home} vs {away}")
        home_score = as_float(market.get("Home Score"))
        away_score = as_float(market.get("Away Score"))
        if home_score is None or away_score is None:
            raise SystemExit(f"Workbook result missing for {match_date} {home} vs {away}")

        actual_margin = home_score - away_score
        actual_total = home_score + away_score
        common = {
            "season": season,
            "round": rnd,
            "date": match_date.isoformat() if match_date else "",
            "home_team": home,
            "away_team": away,
            "home_score": int(home_score),
            "away_score": int(away_score),
            "actual_margin_home": actual_margin,
            "actual_total": actual_total,
            "bookmakers_surveyed": market.get("Bookmakers Surveyed"),
        }

        signals = {
            "normal": {
                "margin": as_float(pred.get("rules_margin")),
                "total": as_float(pred.get("rules_total")),
                "home_prob": as_float(pred.get("rules_home_prob")),
            },
            "ml": {
                "margin": as_float(pred.get("ml_margin")),
                "total": as_float(pred.get("ml_total")),
                "home_prob": as_float(pred.get("ml_h2h")),
            },
            "market": {},
        }

        for signal, values in signals.items():
            if signal == "market":
                selection = h2h_market_move(home, away, market)
                home_fair, away_fair = None, None
            else:
                home_fair, away_fair = fair_odds(values["home_prob"])
                selection = h2h_selection(home, away, values["home_prob"])
            open_odds = odds_for_selection(market, selection, home, away, "Open")
            close_odds = odds_for_selection(market, selection, home, away, "Close")
            out.append({
                **common,
                "market": "h2h",
                "signal": signal,
                "selection": selection,
                "model_number": "",
                "model_home_fair_odds": home_fair,
                "model_away_fair_odds": away_fair,
                "open_number": "",
                "close_number": "",
                "open_odds": open_odds,
                "close_odds": close_odds,
                "clv": decimal_clv(open_odds, close_odds),
                "result": h2h_result(selection, home, away, home_score, away_score),
            })

        home_line_open = as_float(market.get("Home Line Open"))
        for signal, values in signals.items():
            if signal == "market":
                selection = handicap_market_move(home, away, market)
                model_number = ""
            else:
                selection = handicap_pick(values["margin"], home_line_open, home, away)
                model_number = values["margin"]
            open_line = line_for_selection(market, selection, home, away, "Open")
            close_line = line_for_selection(market, selection, home, away, "Close")
            out.append({
                **common,
                "market": "handicap",
                "signal": signal,
                "selection": selection,
                "model_number": model_number,
                "model_home_fair_odds": "",
                "model_away_fair_odds": "",
                "open_number": open_line,
                "close_number": close_line,
                "open_odds": line_odds_for_selection(market, selection, home, away, "Open"),
                "close_odds": line_odds_for_selection(market, selection, home, away, "Close"),
                "clv": handicap_clv(selection, open_line, close_line),
                "result": spread_result(selection, home, away, home_score, away_score, open_line),
            })

        total_open = as_float(market.get("Total Score Open"))
        total_close = as_float(market.get("Total Score Close"))
        for signal, values in signals.items():
            if signal == "market":
                selection = total_market_move(market)
                model_number = ""
            else:
                selection = total_pick(values["total"], total_open)
                model_number = values["total"]
            out.append({
                **common,
                "market": "total",
                "signal": signal,
                "selection": selection,
                "model_number": model_number,
                "model_home_fair_odds": "",
                "model_away_fair_odds": "",
                "open_number": total_open,
                "close_number": total_close,
                "open_odds": total_odds(market, selection, "Open"),
                "close_odds": total_odds(market, selection, "Close"),
                "clv": total_line_clv(selection, total_open, total_close),
                "result": total_result(selection, actual_total, total_open),
            })

    add_winners(out)
    return out


def add_winners(rows: list[dict[str, Any]]) -> None:
    score = {"win": 1, "push": 0, "loss": -1, "": 0}
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    for row in rows:
        key = (row["date"], row["home_team"], row["away_team"], row["market"])
        grouped.setdefault(key, []).append(row)
    for group in grouped.values():
        ordered = sorted(
            group,
            key=lambda r: (
                score.get(r.get("result", ""), 0),
                float(r["clv"]) if r.get("clv") not in ("", None) else -999.0,
            ),
            reverse=True,
        )
        best_score = score.get(ordered[0].get("result", ""), 0)
        best_clv = float(ordered[0]["clv"]) if ordered[0].get("clv") not in ("", None) else -999.0
        winners = [
            r["signal"] for r in ordered
            if score.get(r.get("result", ""), 0) == best_score
            and (float(r["clv"]) if r.get("clv") not in ("", None) else -999.0) == best_clv
        ]
        value = "tie" if len(winners) > 1 else winners[0]
        for row in group:
            row["winner"] = value


def write_report(rows: list[dict[str, Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "season", "round", "date", "home_team", "away_team",
        "home_score", "away_score", "actual_margin_home", "actual_total",
        "market", "signal", "selection",
        "model_number", "model_home_fair_odds", "model_away_fair_odds",
        "open_number", "close_number", "open_odds", "close_odds",
        "clv", "result", "winner", "bookmakers_surveyed",
    ]
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def print_summary(rows: list[dict[str, Any]], out_path: Path) -> None:
    print(f"Written {len(rows)} rows -> {out_path}")
    for signal in ("normal", "ml", "market"):
        print(f"{signal:>6}: {dict(Counter(r['result'] for r in rows if r['signal'] == signal))}")
    print(f"winner: {dict(Counter(r['winner'] for r in rows if r['signal'] == 'normal'))}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build AFL weekly normal-vs-ML-vs-market CLV report.")
    parser.add_argument("--season", type=int, default=2026)
    parser.add_argument("--round", type=int, default=0, dest="round_number", help="0 = latest completed AFL round")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out", type=Path, default=None)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_rows = load_workbook_rows(args.workbook, args.season)
    rnd = args.round_number or auto_round(args.season, workbook_rows)
    out_path = args.out or DEFAULT_REPORT_DIR / f"r{rnd}_afl_ml_clv_comparison_{args.season}.csv"
    rows = build_rows(args.season, rnd, workbook_rows)
    write_report(rows, out_path)
    print_summary(rows, out_path)


if __name__ == "__main__":
    main()
