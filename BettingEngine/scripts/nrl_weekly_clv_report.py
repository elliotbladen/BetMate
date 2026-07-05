#!/usr/bin/env python3
"""
Build a weekly NRL model-vs-market CLV report.

Inputs:
  - outputs/nrl_weekly_review/historical/latest.xlsx
  - results/r{round}_pricing_{season}.csv
  - data/import/r{round}_results_{season}.csv

Output:
  - outputs/nrl_weekly_review/reports/r{round}_nrl_clv_report_{season}.csv

The report compares:
  - model signal: your pricing against the opening market
  - market signal: the open-to-close move

Both are settled against actual scores and measured for CLV against the close.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = ROOT / "outputs" / "nrl_weekly_review" / "historical" / "latest.xlsx"
DEFAULT_REPORT_DIR = ROOT / "outputs" / "nrl_weekly_review" / "reports"
RESULTS_DIR = ROOT / "results"
IMPORT_DIR = ROOT / "data" / "import"


TEAM_ALIASES = {
    "brisbane broncos": "Brisbane Broncos",
    "canterbury bulldogs": "Canterbury-Bankstown Bulldogs",
    "canterbury-bankstown bulldogs": "Canterbury-Bankstown Bulldogs",
    "canberra raiders": "Canberra Raiders",
    "cronulla sharks": "Cronulla-Sutherland Sharks",
    "cronulla-sutherland sharks": "Cronulla-Sutherland Sharks",
    "dolphins": "Dolphins",
    "gold coast titans": "Gold Coast Titans",
    "manly sea eagles": "Manly-Warringah Sea Eagles",
    "manly-warringah sea eagles": "Manly-Warringah Sea Eagles",
    "melbourne storm": "Melbourne Storm",
    "new zealand warriors": "New Zealand Warriors",
    "newcastle knights": "Newcastle Knights",
    "north qld cowboys": "North Queensland Cowboys",
    "north queensland cowboys": "North Queensland Cowboys",
    "parramatta eels": "Parramatta Eels",
    "penrith panthers": "Penrith Panthers",
    "south sydney rabbitohs": "South Sydney Rabbitohs",
    "st george dragons": "St. George Illawarra Dragons",
    "st. george illawarra dragons": "St. George Illawarra Dragons",
    "sydney roosters": "Sydney Roosters",
    "wests tigers": "Wests Tigers",
}


@dataclass(frozen=True)
class SignalResult:
    selection: str
    clv: float | None
    won: str
    result_margin: float | None


def canon_team(raw: str) -> str:
    key = re.sub(r"\s+", " ", (raw or "").strip().lower())
    return TEAM_ALIASES.get(key, raw.strip())


def as_float(value: Any) -> float | None:
    if value in ("", None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    return None


def auto_round(season: int) -> int:
    rounds: list[int] = []
    for path in IMPORT_DIR.glob(f"r*_results_{season}.csv"):
        match = re.fullmatch(r"r(\d+)_results_\d{4}\.csv", path.name)
        if not match:
            continue
        rnd = int(match.group(1))
        if pricing_path(season, rnd).exists():
            rounds.append(rnd)
    if not rounds:
        raise SystemExit(f"No completed round found with both results and pricing for {season}")
    return max(rounds)


def pricing_path(season: int, rnd: int) -> Path:
    preferred = RESULTS_DIR / f"r{rnd}_pricing_{season}.csv"
    if preferred.exists():
        return preferred
    candidates = sorted(RESULTS_DIR.glob(f"r{rnd}*pricing*{season}.csv"))
    if candidates:
        return candidates[0]
    return preferred


def load_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise SystemExit(f"Required file not found: {path}")
    with path.open(newline="", encoding="latin-1") as fh:
        return list(csv.DictReader(fh))


def load_workbook_rows(path: Path, season: int) -> dict[tuple[date, str, str], dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
    headers = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    rows: dict[tuple[date, str, str], dict[str, Any]] = {}

    for values in ws.iter_rows(min_row=3, values_only=True):
        row = dict(zip(headers, values))
        match_date = as_date(row.get("Date"))
        if not match_date or match_date.year != season:
            continue
        home = canon_team(str(row.get("Home Team") or ""))
        away = canon_team(str(row.get("Away Team") or ""))
        rows[(match_date, home, away)] = row

    wb.close()
    return rows


def h2h_pick_from_odds(home: str, away: str, home_odds: float | None, away_odds: float | None) -> str:
    if home_odds is None or away_odds is None:
        return ""
    return home if home_odds <= away_odds else away


def h2h_market_move(home: str, away: str, row: dict[str, Any]) -> str:
    h_open = as_float(row.get("Home Odds Open"))
    h_close = as_float(row.get("Home Odds Close"))
    a_open = as_float(row.get("Away Odds Open"))
    a_close = as_float(row.get("Away Odds Close"))
    if None in (h_open, h_close, a_open, a_close):
        return ""
    home_move = h_open - h_close
    away_move = a_open - a_close
    if abs(home_move - away_move) < 1e-9:
        return "no_move"
    return home if home_move > away_move else away


def odds_for_selection(row: dict[str, Any], selection: str, home: str, away: str, prefix: str) -> float | None:
    if selection == home:
        return as_float(row.get(f"Home Odds {prefix}"))
    if selection == away:
        return as_float(row.get(f"Away Odds {prefix}"))
    return None


def h2h_result(selection: str, home: str, away: str, home_score: float, away_score: float) -> str:
    if home_score == away_score:
        return "push"
    winner = home if home_score > away_score else away
    return "win" if selection == winner else "loss"


def decimal_clv(open_odds: float | None, close_odds: float | None) -> float | None:
    if not open_odds or not close_odds:
        return None
    return round((open_odds / close_odds) - 1.0, 4)


def handicap_pick(model_margin: float | None, home_line: float | None, home: str, away: str) -> str:
    if model_margin is None or home_line is None:
        return ""
    edge = model_margin + home_line
    if abs(edge) < 1e-9:
        return "no_edge"
    return home if edge > 0 else away


def handicap_market_move(home: str, away: str, row: dict[str, Any]) -> str:
    open_line = as_float(row.get("Home Line Open"))
    close_line = as_float(row.get("Home Line Close"))
    if open_line is None or close_line is None:
        return ""
    move = close_line - open_line
    if abs(move) < 1e-9:
        return "no_move"
    return home if move < 0 else away


def line_for_selection(row: dict[str, Any], selection: str, home: str, away: str, prefix: str) -> float | None:
    if selection == home:
        return as_float(row.get(f"Home Line {prefix}"))
    if selection == away:
        return as_float(row.get(f"Away Line {prefix}"))
    return None


def handicap_clv(selection: str, open_line: float | None, close_line: float | None) -> float | None:
    if open_line is None or close_line is None or selection in ("", "no_edge", "no_move"):
        return None
    return round(open_line - close_line, 2)


def spread_result(selection: str, home: str, away: str, home_score: float, away_score: float, line: float | None) -> str:
    if line is None:
        return ""
    if selection == home:
        adjusted = home_score + line - away_score
    elif selection == away:
        adjusted = away_score + line - home_score
    else:
        return ""
    if abs(adjusted) < 1e-9:
        return "push"
    return "win" if adjusted > 0 else "loss"


def total_pick(model_total: float | None, total_open: float | None) -> str:
    if model_total is None or total_open is None:
        return ""
    edge = model_total - total_open
    if abs(edge) < 1e-9:
        return "no_edge"
    return "over" if edge > 0 else "under"


def total_market_move(row: dict[str, Any]) -> str:
    open_total = as_float(row.get("Total Score Open"))
    close_total = as_float(row.get("Total Score Close"))
    if open_total is None or close_total is None:
        return ""
    move = close_total - open_total
    if abs(move) < 1e-9:
        return "no_move"
    return "over" if move > 0 else "under"


def total_line_clv(selection: str, open_total: float | None, close_total: float | None) -> float | None:
    if open_total is None or close_total is None or selection in ("", "no_edge", "no_move"):
        return None
    if selection == "over":
        return round(close_total - open_total, 2)
    if selection == "under":
        return round(open_total - close_total, 2)
    return None


def total_result(selection: str, actual_total: float, line: float | None) -> str:
    if line is None or selection not in ("over", "under"):
        return ""
    diff = actual_total - line
    if abs(diff) < 1e-9:
        return "push"
    if selection == "over":
        return "win" if diff > 0 else "loss"
    return "win" if diff < 0 else "loss"


def winner(model_won: str, market_won: str, model_clv: float | None, market_clv: float | None) -> str:
    score = {"win": 1, "push": 0, "loss": -1, "": 0}
    if score.get(model_won, 0) > score.get(market_won, 0):
        return "model"
    if score.get(model_won, 0) < score.get(market_won, 0):
        return "market"
    mc = model_clv if model_clv is not None else -999.0
    kc = market_clv if market_clv is not None else -999.0
    if mc > kc:
        return "model_clv"
    if kc > mc:
        return "market_clv"
    return "tie"


def build_market_rows(
    season: int,
    rnd: int,
    pricing_rows: list[dict[str, str]],
    actual_rows: list[dict[str, str]],
    workbook_rows: dict[tuple[date, str, str], dict[str, Any]],
) -> list[dict[str, Any]]:
    actual_by_match = {
        (as_date(r.get("match_date")), canon_team(r["home_team"]), canon_team(r["away_team"])): r
        for r in actual_rows
    }

    output: list[dict[str, Any]] = []
    for p in pricing_rows:
        match_date = as_date(p.get("date"))
        home = canon_team(p.get("home_team", ""))
        away = canon_team(p.get("away_team", ""))
        key = (match_date, home, away)
        if key not in actual_by_match:
            raise SystemExit(f"No actual result found for {match_date} {home} vs {away}")
        if key not in workbook_rows:
            raise SystemExit(f"No workbook row found for {match_date} {home} vs {away}")

        actual = actual_by_match[key]
        market = workbook_rows[key]
        home_score = float(actual["home_score"])
        away_score = float(actual["away_score"])
        actual_margin = home_score - away_score
        actual_total = home_score + away_score

        common = {
            "season": season,
            "round": rnd,
            "date": match_date.isoformat() if match_date else "",
            "home_team": home,
            "away_team": away,
            "home_score": int(home_score),
            "away_score": int(away_score),
            "actual_margin_home": actual_margin,
            "actual_total": actual_total,
            "bookmakers_surveyed": market.get("Bookmakers Surveyed"),
        }

        model_margin = as_float(p.get("final_margin") or p.get("fair_hcap_line"))
        model_total = as_float(p.get("final_total") or p.get("fair_total_line"))
        model_home_odds = as_float(p.get("fair_home_odds"))
        model_away_odds = as_float(p.get("fair_away_odds"))

        # H2H
        model_sel = h2h_pick_from_odds(home, away, model_home_odds, model_away_odds)
        market_sel = h2h_market_move(home, away, market)
        for signal_name, selection in (("model", model_sel), ("market", market_sel)):
            open_odds = odds_for_selection(market, selection, home, away, "Open")
            close_odds = odds_for_selection(market, selection, home, away, "Close")
            clv = decimal_clv(open_odds, close_odds)
            won = h2h_result(selection, home, away, home_score, away_score)
            output.append({
                **common,
                "market": "h2h",
                "signal": signal_name,
                "selection": selection,
                "open_number": "",
                "close_number": "",
                "open_odds": open_odds,
                "close_odds": close_odds,
                "model_number": "",
                "model_home_fair_odds": model_home_odds,
                "model_away_fair_odds": model_away_odds,
                "clv": clv,
                "result": won,
            })

        # Handicap
        home_line_open = as_float(market.get("Home Line Open"))
        model_sel = handicap_pick(model_margin, home_line_open, home, away)
        market_sel = handicap_market_move(home, away, market)
        for signal_name, selection in (("model", model_sel), ("market", market_sel)):
            open_line = line_for_selection(market, selection, home, away, "Open")
            close_line = line_for_selection(market, selection, home, away, "Close")
            clv = handicap_clv(selection, open_line, close_line)
            won = spread_result(selection, home, away, home_score, away_score, open_line)
            output.append({
                **common,
                "market": "handicap",
                "signal": signal_name,
                "selection": selection,
                "open_number": open_line,
                "close_number": close_line,
                "open_odds": line_for_selection_odds(market, selection, home, away, "Open"),
                "close_odds": line_for_selection_odds(market, selection, home, away, "Close"),
                "model_number": model_margin,
                "model_home_fair_odds": "",
                "model_away_fair_odds": "",
                "clv": clv,
                "result": won,
            })

        # Totals
        total_open = as_float(market.get("Total Score Open"))
        total_close = as_float(market.get("Total Score Close"))
        model_sel = total_pick(model_total, total_open)
        market_sel = total_market_move(market)
        for signal_name, selection in (("model", model_sel), ("market", market_sel)):
            clv = total_line_clv(selection, total_open, total_close)
            won = total_result(selection, actual_total, total_open)
            output.append({
                **common,
                "market": "total",
                "signal": signal_name,
                "selection": selection,
                "open_number": total_open,
                "close_number": total_close,
                "open_odds": total_odds(market, selection, "Open"),
                "close_odds": total_odds(market, selection, "Close"),
                "model_number": model_total,
                "model_home_fair_odds": "",
                "model_away_fair_odds": "",
                "clv": clv,
                "result": won,
            })

    pair_winners(output)
    return output


def line_for_selection_odds(row: dict[str, Any], selection: str, home: str, away: str, prefix: str) -> float | None:
    if selection == home:
        return as_float(row.get(f"Home Line Odds {prefix}"))
    if selection == away:
        return as_float(row.get(f"Away Line Odds {prefix}"))
    return None


def total_odds(row: dict[str, Any], selection: str, prefix: str) -> float | None:
    if selection == "over":
        return as_float(row.get(f"Total Score Over {prefix}"))
    if selection == "under":
        return as_float(row.get(f"Total Score Under {prefix}"))
    return None


def pair_winners(rows: list[dict[str, Any]]) -> None:
    grouped: dict[tuple[Any, ...], dict[str, dict[str, Any]]] = {}
    for row in rows:
        key = (row["date"], row["home_team"], row["away_team"], row["market"])
        grouped.setdefault(key, {})[row["signal"]] = row

    for pair in grouped.values():
        model = pair.get("model")
        market = pair.get("market")
        if not model or not market:
            continue
        value = winner(model["result"], market["result"], model["clv"], market["clv"])
        model["winner"] = value
        market["winner"] = value


def write_report(rows: list[dict[str, Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "season", "round", "date", "home_team", "away_team",
        "home_score", "away_score", "actual_margin_home", "actual_total",
        "market", "signal", "selection",
        "model_number", "model_home_fair_odds", "model_away_fair_odds",
        "open_number", "close_number", "open_odds", "close_odds",
        "clv", "result", "winner", "bookmakers_surveyed",
    ]
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def print_summary(rows: list[dict[str, Any]], out_path: Path) -> None:
    by_signal = {"model": {"win": 0, "loss": 0, "push": 0}, "market": {"win": 0, "loss": 0, "push": 0}}
    winners: dict[str, int] = {}
    for row in rows:
        if row["signal"] in by_signal and row["result"] in by_signal[row["signal"]]:
            by_signal[row["signal"]][row["result"]] += 1
        if row["signal"] == "model":
            winners[row.get("winner", "")] = winners.get(row.get("winner", ""), 0) + 1
    print(f"Written {len(rows)} rows -> {out_path}")
    print(f"Model results:  {by_signal['model']}")
    print(f"Market results: {by_signal['market']}")
    print(f"Winner counts:  {winners}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build NRL weekly CLV model-vs-market report.")
    parser.add_argument("--season", type=int, default=2026)
    parser.add_argument("--round", type=int, default=0, dest="round_number", help="0 = latest completed round")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out", type=Path, default=None)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rnd = args.round_number or auto_round(args.season)
    price_path = pricing_path(args.season, rnd)
    actual_path = IMPORT_DIR / f"r{rnd}_results_{args.season}.csv"
    out_path = args.out or DEFAULT_REPORT_DIR / f"r{rnd}_nrl_clv_report_{args.season}.csv"

    pricing_rows = load_csv(price_path)
    actual_rows = load_csv(actual_path)
    workbook_rows = load_workbook_rows(args.workbook, args.season)
    rows = build_market_rows(args.season, rnd, pricing_rows, actual_rows, workbook_rows)
    write_report(rows, out_path)
    print_summary(rows, out_path)


if __name__ == "__main__":
    main()
