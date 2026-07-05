#!/usr/bin/env python3
"""
scripts/matrix_confluence.py — T9 Matrix Confluence Analyser

Scans the upcoming round's fixture and flags games where 3+ applicable
matrix edges of 20%+ all point in the same direction for any market.

This is a RESEARCH/FLAGGING tool only. Do not incorporate into pricing
until end-of-season CLV review confirms edge value.

Usage:
    python scripts/matrix_confluence.py --season 2026 --round 13
    python scripts/matrix_confluence.py           # auto-detect round
    python scripts/matrix_confluence.py --min-edges 2   # lower threshold
    python scripts/matrix_confluence.py --push    # push results to Supabase
"""

import argparse
import csv
import json
import os
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.supabase_push import push as _sb_push, load_env as _load_env

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

SCRIPTS_DIR   = Path(__file__).resolve().parent
ENGINE_ROOT   = SCRIPTS_DIR.parent
OUTPUTS_DIR   = ENGINE_ROOT / 'outputs'
DB_PATH       = ENGINE_ROOT / 'data' / 'model.db'
BETMATE_ROOT  = Path(os.environ.get('BETMATE_ROOT', ENGINE_ROOT.parent))
FIXTURE_PATH  = BETMATE_ROOT / 'data' / 'nrl' / 'fixture' / 'processed' / 'latest-fixture.json'

MIN_EDGE   = 20.0   # % threshold for an edge to count
MIN_COUNT  = 3      # how many edges in the same direction to flag

# Per-market overrides for NRL totals (edges can be meaningful at 10%+)
MIN_EDGE_BY_MARKET  = {'totals': 10.0}
MIN_COUNT_BY_MARKET = {'totals': 3}


# ---------------------------------------------------------------------------
# Matrix loaders
# ---------------------------------------------------------------------------

def _parse_edge(val):
    """'6.2% opposing' → (6.2, 'opposing') | None for '—' or empty."""
    if not val:
        return None
    s = str(val).strip()
    if s in ('—', '-', ''):
        return None
    import re
    m = re.match(r'^([\d.]+)%\s+(.+)$', s)
    if not m:
        return None
    return float(m.group(1)), m.group(2).strip()


def load_xlsx_matrix(path: Path) -> dict:
    """Sheet names = team keys. Returns {team: {category: (edgePct, direction)}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    result = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        data = {}
        for row in rows[2:]:            # skip title row + header row
            cat = row[0]
            raw = row[4] if len(row) > 4 else None
            if not cat:
                continue
            parsed = _parse_edge(raw)
            data[str(cat)] = parsed     # None if no edge
        result[sheet] = data
    return result


def load_handicap_csv(path: Path) -> dict:
    """Returns {team: {category: (edgePct, direction)}}."""
    result = defaultdict(dict)
    with open(path, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            team = row['team'].strip()
            cat  = row['category'].strip()
            try:
                edge = float(row['edge_pct'])
            except (ValueError, TypeError):
                continue
            direction = row['direction'].strip()
            if team and cat and direction:
                result[team][cat] = (edge, direction)
    return dict(result)


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def get_team_context(conn, team_name: str, game_date: date) -> dict:
    """
    Returns {'rest_days': int|None, 'last_result': 'win'|'loss'|'draw'|None}
    by looking at the team's most recent completed game before game_date.
    """
    row = conn.execute("""
        SELECT m.match_id, m.match_date, m.home_team_id, m.away_team_id,
               r.home_score, r.away_score
        FROM matches m
        JOIN results r ON m.match_id = r.match_id
        JOIN teams th ON m.home_team_id = th.team_id
        JOIN teams ta ON m.away_team_id = ta.team_id
        WHERE (th.team_name = ? OR ta.team_name = ?)
          AND m.match_date < ?
        ORDER BY m.match_date DESC
        LIMIT 1
    """, (team_name, team_name, game_date.isoformat())).fetchone()

    if not row:
        return {'rest_days': None, 'last_result': None}

    last_date   = datetime.strptime(row['match_date'], '%Y-%m-%d').date()
    rest_days   = (game_date - last_date).days

    home_id = row['home_team_id']
    team_id = conn.execute(
        "SELECT team_id FROM teams WHERE team_name = ?", (team_name,)
    ).fetchone()

    if team_id:
        tid = team_id['team_id']
        hs, aws = row['home_score'], row['away_score']
        if hs is None or aws is None:
            last_result = None
        elif tid == home_id:
            last_result = 'win' if hs > aws else ('loss' if hs < aws else 'draw')
        else:
            last_result = 'win' if aws > hs else ('loss' if aws < hs else 'draw')
    else:
        last_result = None

    return {'rest_days': rest_days, 'last_result': last_result}


# ---------------------------------------------------------------------------
# Moon phase helper
# ---------------------------------------------------------------------------

_MONTH_LABELS = {
    3: 'March', 4: 'April', 5: 'May', 6: 'June',
    7: 'July',  8: 'August', 9: 'September', 10: 'October',
}


def get_moon_phase_row(game_date: date) -> str | None:
    """Return 'New Moon (±1 day)' / 'Full Moon (±1 day)' or None."""
    try:
        import ephem
        d = ephem.Date(game_date.strftime('%Y/%m/%d 12:00:00'))
        prev_new  = ephem.previous_new_moon(d)
        next_new  = ephem.next_new_moon(d)
        prev_full = ephem.previous_full_moon(d)
        next_full = ephem.next_full_moon(d)
        closest_new  = min(abs(float(d) - float(prev_new)),  abs(float(next_new)  - float(d)))
        closest_full = min(abs(float(d) - float(prev_full)), abs(float(next_full) - float(d)))
        if closest_new  <= 1.0:
            return 'New Moon (±1 day)'
        if closest_full <= 1.0:
            return 'Full Moon (±1 day)'
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Row selection
# ---------------------------------------------------------------------------

def applicable_row_names(
    kickoff_hour: int,
    kickoff_weekday: int,   # 0=Mon ... 6=Sun
    game_date: date,
    team_role: str,         # 'home' or 'away'
    home_matrix_key: str,
    away_matrix_key: str,
    venue: str,
    rest_days: int | None,
    last_result: str | None,
    generic_row: str,       # e.g. 'Win % — Home'
) -> list[str]:
    """Return all matrix row names that apply for this game context."""
    rows = [generic_row]

    # Time of day
    if kickoff_hour >= 18:
        rows.append('Night Games (kick-off ≥ 18:00)')
    else:
        rows.append('Day Games (kick-off < 18:00)')

    # Day of week
    if kickoff_weekday in (3, 4):
        rows.append('Thursday / Friday Games')
    elif kickoff_weekday == 5:
        rows.append('Saturday Games')
    elif kickoff_weekday == 6:
        rows.append('Sunday Games')

    # Rest
    if rest_days is not None:
        if rest_days <= 6:
            rows.append('Short Rest (≤ 6 days)')
        elif rest_days >= 10:
            rows.append('Long Rest (≥ 10 days)')

    # Form
    if last_result == 'win':
        rows.append('After a Win')
    elif last_result == 'loss':
        rows.append('After a Loss')

    # Month
    month_label = _MONTH_LABELS.get(game_date.month)
    if month_label:
        rows.append(month_label)

    # Moon phase
    moon_row = get_moon_phase_row(game_date)
    if moon_row:
        rows.append(moon_row)

    # Head-to-head vs specific opponent
    opponent_key = away_matrix_key if team_role == 'home' else home_matrix_key
    rows.append(f'vs {opponent_key}')

    # Venue
    rows.append(venue)

    return rows


# ---------------------------------------------------------------------------
# Direction normalisation
# ---------------------------------------------------------------------------

def normalise_direction(raw_direction: str, team_role: str, market: str) -> str | None:
    """
    Convert a raw matrix direction to a game-level direction string.
    Returns None if direction is unrecognised.
    """
    d = raw_direction.lower().strip()

    if market == 'h2h':
        # 'backing' = the TEAM wins, 'opposing' = the TEAM loses
        is_team_wins = (d == 'backing')
        is_home      = (team_role == 'home')
        return 'HOME_WIN' if (is_team_wins == is_home) else 'AWAY_WIN'

    elif market == 'handicap':
        is_covers = (d == 'covers')
        is_home   = (team_role == 'home')
        if is_home:
            return 'HOME_COVERS' if is_covers else 'AWAY_COVERS'
        else:
            return 'AWAY_COVERS' if is_covers else 'HOME_COVERS'

    elif market == 'totals':
        if 'overs' in d:
            return 'OVERS'
        elif 'unders' in d:
            return 'UNDERS'

    return None


# ---------------------------------------------------------------------------
# Core analysis
# ---------------------------------------------------------------------------

GENERIC_ROWS = {
    'h2h':      ('Win % — Home',        'Win % — Away'),
    'handicap': ('Cover Rate — Home',   'Cover Rate — Away'),
    'totals':   ('Total Points — Home', 'Total Points — Away'),
}


def analyse_game(
    home_team: str, away_team: str,
    home_matrix_key: str, away_matrix_key: str,
    venue: str, kickoff_dt: datetime,
    home_ctx: dict, away_ctx: dict,
    h2h_matrix: dict, totals_matrix: dict, handicap_matrix: dict,
    min_edge_by_market: dict | None = None,
    min_edge: float = MIN_EDGE,
) -> dict:
    """
    Returns dict of per-market confluence results, e.g.:
    {
      'h2h':      {'HOME_WIN': [(edge, row_label, team_label), ...], 'AWAY_WIN': [...]},
      'handicap': {'HOME_COVERS': [...], 'AWAY_COVERS': [...]},
      'totals':   {'OVERS': [...], 'UNDERS': [...]},
    }
    """
    hour    = kickoff_dt.hour
    weekday = kickoff_dt.weekday()   # 0=Mon
    gdate   = kickoff_dt.date()

    results = {}

    for market, matrix in [('h2h', h2h_matrix), ('totals', totals_matrix),
                            ('handicap', handicap_matrix)]:
        threshold = (min_edge_by_market or {}).get(market, min_edge)
        buckets = defaultdict(list)
        home_generic, away_generic = GENERIC_ROWS[market]

        for team_key, team_label, team_role, generic_row, ctx in [
            (home_matrix_key, home_team, 'home', home_generic, home_ctx),
            (away_matrix_key, away_team, 'away', away_generic, away_ctx),
        ]:
            team_data = matrix.get(team_key, {})
            row_names = applicable_row_names(
                hour, weekday, gdate, team_role,
                home_matrix_key, away_matrix_key, venue,
                ctx['rest_days'], ctx['last_result'], generic_row,
            )

            for row_name in row_names:
                val = team_data.get(row_name)
                if not val:
                    continue
                edge_pct, raw_dir = val
                if edge_pct < threshold:
                    continue
                norm_dir = normalise_direction(raw_dir, team_role, market)
                if not norm_dir:
                    continue
                buckets[norm_dir].append((edge_pct, row_name, team_label))

        results[market] = dict(buckets)

    return results


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

DIR_LABELS = {
    'HOME_WIN':    'back HOME',
    'AWAY_WIN':    'back AWAY',
    'HOME_COVERS': 'HOME covers',
    'AWAY_COVERS': 'AWAY covers',
    'OVERS':       'OVERS',
    'UNDERS':      'UNDERS',
}


def print_report(
    games: list,
    results_by_game: list,
    min_count: int = MIN_COUNT,
    min_edge: float = MIN_EDGE,
    min_count_by_market: dict | None = None,
    market_filter: str = 'all',
):
    w = 88
    bar = '═' * w
    thin = '─' * w

    market_label = f' [{market_filter.upper()} only]' if market_filter != 'all' else ''
    print(f'\n{bar}')
    print(f'  T9 MATRIX CONFLUENCE — NRL  (threshold: {int(min_edge)}%+ edge, {min_count}+ signals same direction){market_label}')
    print(bar)

    confluence_found = 0

    for game, analysis in zip(games, results_by_game):
        home  = game['home_team']
        away  = game['away_team']
        venue = game['venue']
        ko    = game['kickoff_local']

        # Check if any market has confluence
        flags = []
        for market, buckets in analysis.items():
            if market_filter != 'all' and market != market_filter:
                continue
            mcount = (min_count_by_market or {}).get(market, min_count)
            for direction, edges in buckets.items():
                if len(edges) >= mcount:
                    flags.append((market, direction, edges))

        if not flags:
            continue

        confluence_found += 1
        print(f'\n{thin}')
        print(f'  {home}  vs  {away}')
        print(f'  {ko[:16]}  |  {venue}')
        print(f'  Context:')
        print(f'    HOME rest={game["home_ctx"]["rest_days"]}d  form={game["home_ctx"]["last_result"] or "?"}')
        print(f'    AWAY rest={game["away_ctx"]["rest_days"]}d  form={game["away_ctx"]["last_result"] or "?"}')
        print()

        for market, direction, edges in sorted(flags, key=lambda x: -len(x[2])):
            label = DIR_LABELS.get(direction, direction)
            star  = '⚡' if len(edges) >= min_count else '  '
            print(f'  {star} {market.upper():<10}  {len(edges)}-way confluence  →  {label.upper()}')
            for edge_pct, row_name, team_label in sorted(edges, key=lambda x: -x[0]):
                short_team = team_label.split()[-1]    # last word (nickname)
                print(f'      {edge_pct:5.1f}%  {short_team:<10}  {row_name}')
            print()

    if confluence_found == 0:
        print(f'\n  No games with {min_count}+ edges ≥{int(min_edge)}% in the same direction this round.\n')
    else:
        print(f'{bar}')
        print(f'  {confluence_found} game(s) with {min_count}+ confluence signals found.')
    print(bar + '\n')


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description='T9 matrix confluence analyser')
    ap.add_argument('--season', type=int, default=2026)
    ap.add_argument('--round',  type=int, default=None)
    ap.add_argument('--min-edges', type=int, default=MIN_COUNT,
                    help='Minimum edges in same direction to flag (default 3)')
    ap.add_argument('--min-edge-pct', type=float, default=MIN_EDGE,
                    help='Minimum edge %% to count (default 20)')
    ap.add_argument('--market', choices=['h2h', 'handicap', 'totals', 'all'],
                    default='all', help='Show only signals for this market')
    ap.add_argument('--push', action='store_true',
                    help='Push confluence results to Supabase')
    args = ap.parse_args()

    # ── Load fixture ──────────────────────────────────────────────────────
    if not FIXTURE_PATH.exists():
        print(f'ERROR: Fixture not found at {FIXTURE_PATH}')
        sys.exit(1)

    fixture = json.loads(FIXTURE_PATH.read_text(encoding='utf-8'))
    round_num = args.round or fixture.get('round')
    games_raw = fixture.get('games', [])
    if not games_raw:
        print('ERROR: No games in fixture.')
        sys.exit(1)

    print(f'\nLoading R{round_num} fixture  ({len(games_raw)} games) ...')

    # ── Load matrices ─────────────────────────────────────────────────────
    print('Loading matrices ...')
    h2h_matrix      = load_xlsx_matrix(OUTPUTS_DIR / 'nrl_h2h_matrix.xlsx')
    totals_matrix   = load_xlsx_matrix(OUTPUTS_DIR / 'nrl_team_totals_matrix.xlsx')
    handicap_matrix = load_handicap_csv(OUTPUTS_DIR / 'nrl_handicap_matrix.csv')

    # ── Map team names to matrix keys ─────────────────────────────────────
    from utils.teams import to_matrix_key

    # ── Open DB ───────────────────────────────────────────────────────────
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # ── Analyse each game ─────────────────────────────────────────────────
    games_enriched   = []
    results_by_game  = []

    for g in games_raw:
        home_team = g['home_team']
        away_team = g['away_team']
        venue     = g['venue']

        ko_str = g.get('kickoff_local') or g.get('kickoff_utc', '')
        try:
            ko_dt = datetime.fromisoformat(ko_str.replace('Z', '+00:00'))
        except Exception:
            ko_dt = datetime.now()

        game_date = ko_dt.date()

        home_matrix_key = to_matrix_key(home_team)
        away_matrix_key = to_matrix_key(away_team)

        home_ctx = get_team_context(conn, home_team, game_date)
        away_ctx = get_team_context(conn, away_team, game_date)

        analysis = analyse_game(
            home_team, away_team,
            home_matrix_key, away_matrix_key,
            venue, ko_dt,
            home_ctx, away_ctx,
            h2h_matrix, totals_matrix, handicap_matrix,
            min_edge_by_market=MIN_EDGE_BY_MARKET,
            min_edge=args.min_edge_pct,
        )

        enriched = {**g,
                    'home_matrix_key': home_matrix_key,
                    'away_matrix_key': away_matrix_key,
                    'home_ctx': home_ctx,
                    'away_ctx': away_ctx}

        games_enriched.append(enriched)
        results_by_game.append(analysis)

    conn.close()

    # ── Print report ──────────────────────────────────────────────────────
    print_report(games_enriched, results_by_game, min_count=args.min_edges,
                 min_edge=args.min_edge_pct, min_count_by_market=MIN_COUNT_BY_MARKET,
                 market_filter=args.market)

    def _meets_count(market, n):
        return n >= MIN_COUNT_BY_MARKET.get(market, args.min_edges)

    # ── Write local JSON for Baz ──────────────────────────────────────────
    local_games = []
    for game, analysis in zip(games_enriched, results_by_game):
        flags = {}
        for market, buckets in analysis.items():
            for direction, edges in buckets.items():
                if _meets_count(market, len(edges)):
                    flags[f'{market}_{direction}'] = {
                        'count': len(edges),
                        'edges': [{'edge_pct': e, 'row': r, 'team': t} for e, r, t in edges],
                    }
        if flags:
            local_games.append({
                'home': game['home_team'],
                'away': game['away_team'],
                'confluence': flags,
            })
    confluence_out = OUTPUTS_DIR / 'nrl_t9_confluence_latest.json'
    confluence_out.write_text(json.dumps({
        'season': args.season,
        'round': round_num,
        'generated_at': datetime.now().isoformat(),
        'games': local_games,
    }, indent=2), encoding='utf-8')
    print(f'  Confluence JSON written -> {confluence_out}')

    # ── Optionally push to Supabase ───────────────────────────────────────
    if args.push:
        _load_env()
        payload = []
        for game, analysis in zip(games_enriched, results_by_game):
            flags = {}
            for market, buckets in analysis.items():
                for direction, edges in buckets.items():
                    if _meets_count(market, len(edges)):
                        flags[f'{market}_{direction}'] = [
                            {'edge_pct': e, 'row': r, 'team': t}
                            for e, r, t in edges
                        ]
            if flags:
                payload.append({
                    'home': game['home_team'],
                    'away': game['away_team'],
                    'venue': game['venue'],
                    'kickoff': game.get('kickoff_local', ''),
                    'confluence': flags,
                })
        key = f'nrl_t9_confluence_r{round_num}_{args.season}'
        if _sb_push(key, payload):
            print(f'  Pushed to Supabase: {key}')
        else:
            print('  Supabase push skipped (env vars not set)')


if __name__ == '__main__':
    main()
