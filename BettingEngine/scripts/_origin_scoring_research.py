"""
Deep research: Does NRL scoring drop in rounds played immediately after State of Origin games?

Method:
1. Pull all NRL match results from model.db + aussportsbetting historical xlsx
2. Map each match to whether it fell in a "post-Origin" round
3. Compare scoring distributions: post-Origin vs normal rounds
4. Run t-test for statistical significance
5. Break down by season, Origin game number, and days-since-Origin
"""
import sqlite3, csv, statistics
from pathlib import Path
from collections import defaultdict
from datetime import date, timedelta

# State of Origin game dates (historical + 2026)
# Format: (season, game_number, origin_date)
ORIGIN_DATES = [
    # 2026
    (2026, 1, date(2026, 5, 27)),
    (2026, 2, date(2026, 6, 17)),
    (2026, 3, date(2026, 7,  8)),
    # 2025
    (2025, 1, date(2025, 5, 28)),
    (2025, 2, date(2025, 6, 18)),
    (2025, 3, date(2025, 7,  9)),
    # 2024
    (2024, 1, date(2024, 5, 29)),
    (2024, 2, date(2024, 6, 19)),
    (2024, 3, date(2024, 7, 10)),
    # 2023
    (2023, 1, date(2023, 5, 31)),
    (2023, 2, date(2023, 6, 21)),
    (2023, 3, date(2023, 7, 12)),
    # 2022
    (2022, 1, date(2022, 6,  1)),
    (2022, 2, date(2022, 6, 22)),
    (2022, 3, date(2022, 7, 13)),
    # 2021
    (2021, 1, date(2021, 6, 9)),
    (2021, 2, date(2021, 6, 27)),
    (2021, 3, date(2021, 7, 14)),
    # 2019
    (2019, 1, date(2019, 5, 22)),
    (2019, 2, date(2019, 6, 19)),
    (2019, 3, date(2019, 7, 10)),
    # 2018
    (2018, 1, date(2018, 5, 30)),
    (2018, 2, date(2018, 6, 20)),
    (2018, 3, date(2018, 7, 11)),
    # 2017
    (2017, 1, date(2017, 5, 31)),
    (2017, 2, date(2017, 6, 21)),
    (2017, 3, date(2017, 7, 12)),
]

def days_since_nearest_origin(match_date, season):
    """Return (days_since, game_number) for nearest prior Origin game in same season."""
    candidates = [(d, g) for (s, g, d) in ORIGIN_DATES if s == season and d < match_date]
    if not candidates:
        return None, None
    nearest = max(candidates, key=lambda x: x[0])
    return (match_date - nearest[0]).days, nearest[1]

POST_ORIGIN_WINDOW = 10  # days after Origin game = "post-Origin round"

BE = Path("C:/Users/ElliotBladen/Apps/BettingEngine")

# ── Load from model.db ────────────────────────────────────────────────────
matches = []  # list of dicts: season, round, match_date, total, is_post_origin, days_since, game_num

db_path = BE / "data/model.db"
if db_path.exists():
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    # Check what seasons are available
    cur.execute("SELECT DISTINCT season FROM matches ORDER BY season")
    seasons_in_db = [r["season"] for r in cur.fetchall()]
    print(f"Seasons in model.db: {seasons_in_db}")

    cur.execute("""
        SELECT m.season, m.round_number, m.match_date,
               r.total_score
        FROM results r
        JOIN matches m ON m.match_id = r.match_id
        WHERE r.total_score > 0
        ORDER BY m.season, m.round_number, m.match_date
    """)
    for row in cur.fetchall():
        match_date = date.fromisoformat(row["match_date"])
        days, gnum = days_since_nearest_origin(match_date, row["season"])
        matches.append({
            "season": row["season"],
            "round": row["round_number"],
            "date": match_date,
            "total": int(row["total_score"]),
            "days_since": days,
            "game_num": gnum,
            "is_post_origin": days is not None and days <= POST_ORIGIN_WINDOW,
            "source": "db",
        })
    conn.close()
    print(f"Loaded {len(matches)} matches from model.db")

# ── Also load from aussportsbetting historical xlsx ───────────────────────
try:
    import openpyxl
    hist_path = BE / "outputs/afl_weekly_review/historical"
    # NRL xlsx
    nrl_xlsx = list(Path("C:/Users/ElliotBladen/Apps/data/nrl/historical").glob("*.xlsx"))
    if not nrl_xlsx:
        nrl_xlsx = list(Path("C:/Users/ElliotBladen/Apps/data").glob("nrl_*.xlsx"))
    print(f"NRL historical xlsx files: {[f.name for f in nrl_xlsx]}")

    for xlsx_path in nrl_xlsx:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"  {xlsx_path.name} headers: {headers[:15]}")

        date_col = next((i for i, h in enumerate(headers) if "date" in h), None)
        home_score_col = next((i for i, h in enumerate(headers) if "home" in h and "score" in h), None)
        away_score_col = next((i for i, h in enumerate(headers) if "away" in h and "score" in h), None)
        season_col = next((i for i, h in enumerate(headers) if "season" in h), None)

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if date_col is None or row[date_col] is None:
                    continue
                match_date = row[date_col]
                if hasattr(match_date, "date"):
                    match_date = match_date.date()
                elif isinstance(match_date, str):
                    match_date = date.fromisoformat(match_date[:10])
                else:
                    continue

                hs = int(row[home_score_col]) if home_score_col is not None and row[home_score_col] else None
                aws = int(row[away_score_col]) if away_score_col is not None and row[away_score_col] else None
                if hs is None or aws is None:
                    continue
                total = hs + aws
                season = int(row[season_col]) if season_col is not None and row[season_col] else match_date.year

                days, gnum = days_since_nearest_origin(match_date, season)
                # Check not already in matches
                matches.append({
                    "season": season,
                    "round": None,
                    "date": match_date,
                    "total": total,
                    "days_since": days,
                    "game_num": gnum,
                    "is_post_origin": days is not None and days <= POST_ORIGIN_WINDOW,
                    "source": "xlsx",
                })
            except Exception:
                continue
        wb.close()

except ImportError:
    print("openpyxl not available — skipping xlsx")
except Exception as e:
    print(f"xlsx error: {e}")

# ── Analysis ──────────────────────────────────────────────────────────────
print(f"\nTotal matches loaded: {len(matches)}")

# Filter to NRL seasons only (2017+, skip test data)
matches = [m for m in matches if m["season"] >= 2017]

# Deduplicate by (season, date, total) — xlsx and db may overlap
seen = set()
deduped = []
for m in matches:
    key = (m["season"], str(m["date"]), m["total"])
    if key not in seen:
        seen.add(key)
        deduped.append(m)
matches = deduped
print(f"After dedup (2017+): {len(matches)} matches")

post_origin = [m for m in matches if m["is_post_origin"]]
normal      = [m for m in matches if not m["is_post_origin"]]

print(f"\n  Post-Origin matches (within {POST_ORIGIN_WINDOW}d of an Origin game): {len(post_origin)}")
print(f"  Normal matches: {len(normal)}")

def stats(vals):
    n = len(vals)
    if n == 0:
        return {"n": 0, "mean": None, "median": None, "sd": None, "min": None, "max": None}
    mean = sum(vals) / n
    med = statistics.median(vals)
    sd = statistics.stdev(vals) if n > 1 else 0
    return {"n": n, "mean": mean, "median": med, "sd": sd, "min": min(vals), "max": max(vals)}

post_totals   = [m["total"] for m in post_origin]
normal_totals = [m["total"] for m in normal]

ps = stats(post_totals)
ns = stats(normal_totals)

print("\n" + "="*70)
print("  NRL SCORING: POST-ORIGIN vs NORMAL ROUNDS")
print("="*70)
print(f"  {'Metric':<15} {'Post-Origin':>14} {'Normal':>14} {'Diff':>10}")
print(f"  {'-'*53}")
print(f"  {'N (games)':<15} {ps['n']:>14} {ns['n']:>14}")
print(f"  {'Mean total':<15} {ps['mean']:>14.2f} {ns['mean']:>14.2f} {ps['mean']-ns['mean']:>+10.2f}")
print(f"  {'Median total':<15} {ps['median']:>14.1f} {ns['median']:>14.1f} {ps['median']-ns['median']:>+10.1f}")
print(f"  {'Std dev':<15} {ps['sd']:>14.2f} {ns['sd']:>14.2f}")
print(f"  {'Min':<15} {ps['min']:>14} {ns['min']:>14}")
print(f"  {'Max':<15} {ps['max']:>14} {ns['max']:>14}")

# T-test (Welch's)
try:
    import math
    n1, m1, s1 = ps['n'], ps['mean'], ps['sd']
    n2, m2, s2 = ns['n'], ns['mean'], ns['sd']
    se = math.sqrt(s1**2/n1 + s2**2/n2)
    t = (m1 - m2) / se
    # Welch df
    df = (s1**2/n1 + s2**2/n2)**2 / ((s1**2/n1)**2/(n1-1) + (s2**2/n2)**2/(n2-1))
    print(f"\n  Welch t-test: t={t:.3f}  df={df:.0f}")
    # Rough p-value guide
    abs_t = abs(t)
    if abs_t > 3.29:   p_est = "< 0.001 (***)"
    elif abs_t > 2.58: p_est = "< 0.01  (**)"
    elif abs_t > 1.96: p_est = "< 0.05  (*)"
    elif abs_t > 1.64: p_est = "< 0.10  (.)"
    else:              p_est = "> 0.10  (NS)"
    print(f"  p-value est: {p_est}")
    print(f"  Effect: {'LOWER' if t < 0 else 'HIGHER'} scoring post-Origin")
except Exception as e:
    print(f"  t-test error: {e}")

# ── By Origin game number ─────────────────────────────────────────────────
print(f"\n  By Origin Game Number:")
print(f"  {'Game':<8} {'N':>5} {'Mean':>8} {'vs Normal':>10}")
for gnum in [1, 2, 3]:
    g_matches = [m for m in post_origin if m["game_num"] == gnum]
    vals = [m["total"] for m in g_matches]
    if vals:
        mean_g = sum(vals)/len(vals)
        diff = mean_g - ns['mean']
        print(f"  Post-G{gnum}  {len(vals):>5} {mean_g:>8.1f} {diff:>+10.1f}")

# ── By season ────────────────────────────────────────────────────────────
print(f"\n  By Season:")
print(f"  {'Season':<8} {'Post-Orig N':>12} {'Post Mean':>10} {'Normal N':>10} {'Normal Mean':>12} {'Diff':>8}")
for season in sorted(set(m["season"] for m in matches)):
    s_post   = [m["total"] for m in post_origin if m["season"] == season]
    s_normal = [m["total"] for m in normal if m["season"] == season]
    if s_post and s_normal:
        pm = sum(s_post)/len(s_post)
        nm = sum(s_normal)/len(s_normal)
        print(f"  {season:<8} {len(s_post):>12} {pm:>10.1f} {len(s_normal):>10} {nm:>12.1f} {pm-nm:>+8.1f}")

# ── Days-since breakdown ─────────────────────────────────────────────────
print(f"\n  By Days Since Origin (post-Origin only):")
print(f"  {'Days':>6} {'N':>5} {'Mean':>8} {'vs Normal':>10}")
for days_bucket in [(1,4),(4,7),(7,11)]:
    lo, hi = days_bucket
    bucket = [m["total"] for m in post_origin if m["days_since"] is not None and lo <= m["days_since"] < hi]
    if bucket:
        bm = sum(bucket)/len(bucket)
        print(f"  {lo}-{hi-1}d    {len(bucket):>5} {bm:>8.1f} {bm-ns['mean']:>+10.1f}")

# ── Under rate ───────────────────────────────────────────────────────────
# How often does actual total fall under market typical line (48.0)?
MARKET_LINE = 48.0
for label, pool in [("Post-Origin", post_totals), ("Normal", normal_totals)]:
    if pool:
        under_rate = sum(1 for t in pool if t < MARKET_LINE) / len(pool)
        print(f"\n  Under {MARKET_LINE} rate — {label}: {under_rate:.1%}  (n={len(pool)})")

print("\n  Under 44.0 rate (below typical post-Origin actuals we saw 2026):")
for label, pool in [("Post-Origin", post_totals), ("Normal", normal_totals)]:
    if pool:
        under_rate = sum(1 for t in pool if t < 44.0) / len(pool)
        print(f"  Under 44.0 — {label}: {under_rate:.1%}  (n={len(pool)})")
