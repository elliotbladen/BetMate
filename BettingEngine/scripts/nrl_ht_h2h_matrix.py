"""
NRL Halftime H2H Matrix Builder
Generates one sheet per NRL team showing actual win % vs halftime market implied win %.
Source: data/inplay/nrl/halftime/processed/halftime_dataset.csv (seasons 2022–2026)
Output: outputs/nrl_ht_h2h_matrix.xlsx

Same structure as nrl_h2h_matrix.py, plus a HALFTIME SCORE POSITION section:
  - Home & Leading at HT
  - Home & Trailing at HT
  - Away & Leading at HT
  - Away & Trailing at HT
"""

import os
from datetime import timedelta, date
from collections import defaultdict
from pathlib import Path

import ephem
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

_ROOT = Path(__file__).resolve().parents[1]

SOURCE_PATH = str(_ROOT / "data" / "inplay" / "nrl" / "halftime" / "processed" / "halftime_dataset.csv")
OUTPUT_PATH = str(_ROOT / "outputs" / "nrl_ht_h2h_matrix.xlsx")

SEASONS        = (2022, 2023, 2024, 2025, 2026)
MIN_SAMPLE     = 3
EDGE_FLAG_PCT  = 15.0
MOON_WINDOW_DAYS = 1


# ─────────────────────────────────────────────
#  Team name normalisation
# ─────────────────────────────────────────────

TEAM_NAME_MAP = {
    "Brisbane Broncos":          "Brisbane Broncos",
    "Canberra Raiders":          "Canberra Raiders",
    "Canterbury":                "Canterbury Bulldogs",
    "Canterbury Bulldogs":       "Canterbury Bulldogs",
    "Cronulla Sharks":           "Cronulla Sharks",
    "Dolphins":                  "Dolphins",
    "Gold Coast":                "Gold Coast Titans",
    "Gold Coast Titans":         "Gold Coast Titans",
    "Illa Dragons":              "St George Illawarra Dragons",
    "Manly Sea Eagles":          "Manly Sea Eagles",
    "Melbourne Storm":           "Melbourne Storm",
    "New Zealand Warriors":      "New Zealand Warriors",
    "Newcastle Knights":         "Newcastle Knights",
    "North Qld Cowboys":         "North QLD Cowboys",
    "North Queensland Cowboy":   "North QLD Cowboys",
    "North Queensland Cowboys":  "North QLD Cowboys",
    "Parramatta Eels":           "Parramatta Eels",
    "Penrith Panthers":          "Penrith Panthers",
    "South Sydney Rabbitohs":    "South Sydney Rabbitohs",
    "St George Illawarra Dra":   "St George Illawarra Dragons",
    "St George/Illa Dragons":    "St George Illawarra Dragons",
    "Sydney":                    "Sydney Roosters",
    "Sydney Roosters":           "Sydney Roosters",
    "Wests Tigers":              "Wests Tigers",
}

# Known non-NRL entries to drop
_DROP_TEAMS = {"Canterbury Bulldogs Wom", "New Zealand Warriors Wo"}


# ─────────────────────────────────────────────
#  Moon phase helpers
# ─────────────────────────────────────────────

def build_moon_sets(start_date: date, end_date: date):
    new_moons, full_moons = set(), set()
    d = ephem.Date(start_date - timedelta(days=30))
    end_ephem = ephem.Date(end_date + timedelta(days=30))
    while d < end_ephem:
        nm = ephem.next_new_moon(d)
        fm = ephem.next_full_moon(d)
        nm_date = ephem.Date(nm).datetime().date()
        fm_date = ephem.Date(fm).datetime().date()
        for delta in range(-MOON_WINDOW_DAYS, MOON_WINDOW_DAYS + 1):
            new_moons.add(nm_date + timedelta(days=delta))
            full_moons.add(fm_date + timedelta(days=delta))
        d = nm + 1
    return new_moons, full_moons


# ─────────────────────────────────────────────
#  Data loading
# ─────────────────────────────────────────────

def load_data():
    df = pd.read_csv(SOURCE_PATH)

    # Drop non-NRL rows
    df = df[~df["home_team"].isin(_DROP_TEAMS) & ~df["away_team"].isin(_DROP_TEAMS)]

    # Normalise team names
    df["home_team"] = df["home_team"].map(lambda x: TEAM_NAME_MAP.get(x, x))
    df["away_team"] = df["away_team"].map(lambda x: TEAM_NAME_MAP.get(x, x))

    # Filter to known seasons
    df = df[df["year"].isin(SEASONS)]

    # Parse dates
    df["game_date"] = pd.to_datetime(df["date"]).dt.date

    # Require halftime prices and final result
    df = df.dropna(subset=["ht_home_impl_prob", "ht_away_impl_prob", "home_won"])

    rows = []
    for _, r in df.iterrows():
        rows.append({
            "game_date":     r["game_date"],
            "season":        int(r["year"]),
            "home_team":     r["home_team"],
            "away_team":     r["away_team"],
            "home_won":      int(r["home_won"]),
            "implied_home":  float(r["ht_home_impl_prob"]),
            "implied_away":  float(r["ht_away_impl_prob"]),
            "ht_leader":     str(r["ht_leader"]) if pd.notna(r["ht_leader"]) else "level",
        })

    rows.sort(key=lambda r: r["game_date"])
    return rows


def enrich_rows(rows):
    dates = [r["game_date"] for r in rows]
    new_moon_dates, full_moon_dates = build_moon_sets(min(dates), max(dates))

    for r in rows:
        d = r["game_date"]
        wd = d.weekday()
        r["is_thu_fri"]  = wd in (3, 4)
        r["is_saturday"] = wd == 5
        r["is_sunday"]   = wd == 6
        r["is_new_moon"]  = d in new_moon_dates
        r["is_full_moon"] = d in full_moon_dates

    team_games_map: dict[str, list] = defaultdict(list)
    for r in rows:
        team_games_map[r["home_team"]].append(r)
        team_games_map[r["away_team"]].append(r)

    for team, games in team_games_map.items():
        games.sort(key=lambda x: x["game_date"])
        for i, g in enumerate(games):
            key_rest = f"rest__{team}"
            key_prev = f"prev_win__{team}"
            if i == 0:
                g[key_rest] = None
                g[key_prev] = None
            else:
                prev = games[i - 1]
                g[key_rest] = (g["game_date"] - prev["game_date"]).days
                if prev["home_team"] == team:
                    g[key_prev] = prev["home_won"] == 1
                else:
                    g[key_prev] = prev["home_won"] == 0

    return rows


# ─────────────────────────────────────────────
#  Stats computation
# ─────────────────────────────────────────────

def team_win(game, team):
    if game["home_team"] == team:
        return game["home_won"] == 1
    else:
        return game["home_won"] == 0


def implied_prob(game, team):
    if game["home_team"] == team:
        return game["implied_home"]
    else:
        return game["implied_away"]


def compute_stats(games, team):
    n = len(games)
    if n < MIN_SAMPLE:
        return None
    wins        = sum(1 for g in games if team_win(g, team))
    actual_pct  = round((wins / n) * 100, 1)
    implied_pct = round((sum(implied_prob(g, team) for g in games) / n) * 100, 1)
    return actual_pct, implied_pct, n


def edge_label(actual_pct, implied_pct):
    diff = actual_pct - implied_pct
    if implied_pct == 0:
        return round(diff, 1), 0.0, "", False
    edge_pct  = abs(diff / implied_pct) * 100
    direction = "backing" if diff > 0 else "opposing"
    flag      = edge_pct >= EDGE_FLAG_PCT
    return round(diff, 1), round(edge_pct, 1), direction, flag


def get_team_games(all_rows, team):
    return [r for r in all_rows if r["home_team"] == team or r["away_team"] == team]


# ─────────────────────────────────────────────
#  Excel formatting
# ─────────────────────────────────────────────

HEADER_FILL      = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT      = Font(color="FFFFFF", bold=True, size=10)
SECTION_FILL     = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT     = Font(color="FFFFFF", bold=True, size=9)
LABEL_FILL       = PatternFill("solid", fgColor="D6E4F0")
ALT_ROW_FILL     = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL       = PatternFill("solid", fgColor="FFFFFF")
BLANK_FILL       = PatternFill("solid", fgColor="F2F2F2")
FLAG_FILL        = PatternFill("solid", fgColor="6CE58D")
STRONG_FLAG_FILL = PatternFill("solid", fgColor="00FF00")
HT_SECTION_FILL  = PatternFill("solid", fgColor="7030A0")
THIN             = Side(style="thin", color="CCCCCC")
THIN_BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_HEADERS = [
    "Category",
    "Actual Win %",
    "HT Market Implied Win %",
    "Difference (pp)",
    "Edge % & Direction",
    "N (Games)",
]


def style_header_row(ws, row):
    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_section_row(ws, row, label, fill=None):
    ws.cell(row=row, column=1).value = label
    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill or SECTION_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER


def write_data_row(ws, row, label, stats, alt_row=False):
    bg = ALT_ROW_FILL if alt_row else WHITE_FILL

    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.fill = LABEL_FILL
    label_cell.font = Font(size=9)
    label_cell.alignment = Alignment(vertical="center", indent=1)
    label_cell.border = THIN_BORDER

    if stats is None:
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c, value="—")
            cell.fill = BLANK_FILL
            cell.font = Font(color="AAAAAA", size=9)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        return

    actual_pct, implied_pct, n = stats
    diff, edge_pct, direction, flag = edge_label(actual_pct, implied_pct)
    edge_text = f"{edge_pct}% {direction}" if edge_pct >= 1.0 else "—"

    for c_idx, val in enumerate([actual_pct, implied_pct, diff, edge_text, n], start=2):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = Font(bold=(c_idx == 5 and flag), size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if c_idx in (2, 3, 4):
            cell.fill = bg
            cell.number_format = "0.0"
        elif c_idx == 5:
            cell.fill = STRONG_FLAG_FILL if (flag and edge_pct >= 30) else FLAG_FILL if flag else bg
        else:
            cell.fill = bg


# ─────────────────────────────────────────────
#  Sheet builder
# ─────────────────────────────────────────────

def build_team_sheet(wb, team, all_rows, all_teams):
    ws = wb.create_sheet(title=team[:31])

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12

    ws.merge_cells("A1:F1")
    tc = ws.cell(row=1, column=1, value=f"{team} — NRL Halftime H2H Matrix (2022–2026)")
    tc.fill = PatternFill("solid", fgColor="0D2137")
    tc.font = Font(color="FFFFFF", bold=True, size=12)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(COL_HEADERS, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2)
    ws.row_dimensions[2].height = 30

    cur_row = 3
    g = get_team_games(all_rows, team)
    home_games = [x for x in g if x["home_team"] == team]
    away_games = [x for x in g if x["away_team"] == team]

    def section(label, fill=None):
        nonlocal cur_row
        style_section_row(ws, cur_row, label, fill=fill)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

    def row(label, games, alt=False):
        nonlocal cur_row
        write_data_row(ws, cur_row, label, compute_stats(games, team), alt_row=alt)
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

    # ── OVERALL ──────────────────────────────────────────────────
    section("OVERALL")
    row("Win % — All Games", g)
    row("Win % — Home",      home_games, alt=True)
    row("Win % — Away",      away_games)

    # ── HALFTIME SCORE POSITION ───────────────────────────────────
    section("HALFTIME SCORE POSITION", fill=HT_SECTION_FILL)
    row("Home & Leading at HT",  [x for x in home_games if x["ht_leader"] == "home"])
    row("Home & Trailing at HT", [x for x in home_games if x["ht_leader"] == "away"], alt=True)
    row("Away & Leading at HT",  [x for x in away_games if x["ht_leader"] == "away"])
    row("Away & Trailing at HT", [x for x in away_games if x["ht_leader"] == "home"], alt=True)
    row("Level at HT (Home)",    [x for x in home_games if x["ht_leader"] == "level"])
    row("Level at HT (Away)",    [x for x in away_games if x["ht_leader"] == "level"], alt=True)

    # ── DAY OF WEEK ──────────────────────────────────────────────
    section("DAY OF WEEK")
    row("Thursday / Friday Games", [x for x in g if x["is_thu_fri"]])
    row("Saturday Games",          [x for x in g if x["is_saturday"]], alt=True)
    row("Sunday Games",            [x for x in g if x["is_sunday"]])

    # ── FORM ─────────────────────────────────────────────────────
    section("FORM")
    key_prev = f"prev_win__{team}"
    row("After a Win",  [x for x in g if x.get(key_prev) is True])
    row("After a Loss", [x for x in g if x.get(key_prev) is False], alt=True)

    # ── REST ─────────────────────────────────────────────────────
    section("REST")
    key_rest = f"rest__{team}"
    row("Short Rest (≤ 6 days)", [x for x in g if x.get(key_rest) is not None and x[key_rest] <= 6])
    row("Long Rest (≥ 10 days)", [x for x in g if x.get(key_rest) is not None and x[key_rest] >= 10], alt=True)

    # ── MOON PHASE ───────────────────────────────────────────────
    section("MOON PHASE")
    row("New Moon (±1 day)",  [x for x in g if x["is_new_moon"]])
    row("Full Moon (±1 day)", [x for x in g if x["is_full_moon"]], alt=True)

    # ── BY MONTH ─────────────────────────────────────────────────
    section("BY MONTH")
    nrl_months = [
        (3, "March"), (4, "April"), (5, "May"),    (6, "June"),
        (7, "July"),  (8, "August"),(9, "September"),(10, "October"),
    ]
    for i, (m, mname) in enumerate(nrl_months):
        row(mname, [x for x in g if x["game_date"].month == m], alt=(i % 2 == 1))

    # ── HEAD TO HEAD vs OPPONENT ─────────────────────────────────
    section("HEAD TO HEAD vs OPPONENT")
    opponents = sorted(t for t in all_teams if t != team)
    for i, opp in enumerate(opponents):
        opp_games = [x for x in g if x["home_team"] == opp or x["away_team"] == opp]
        row(f"vs {opp}", opp_games, alt=(i % 2 == 1))

    ws.freeze_panes = "B3"


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────

def main():
    print("Loading halftime dataset...")
    all_rows = load_data()
    print(f"  Loaded {len(all_rows)} games (seasons {SEASONS})")

    season_counts = defaultdict(int)
    for r in all_rows:
        season_counts[r["season"]] += 1
    for s in sorted(season_counts):
        print(f"    {s}: {season_counts[s]} games")

    print("Enriching rows (moon, rest, form)...")
    all_rows = enrich_rows(all_rows)

    all_teams = sorted(set(r["home_team"] for r in all_rows) | set(r["away_team"] for r in all_rows))
    print(f"  {len(all_teams)} teams")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Building sheets...")
    for team in all_teams:
        print(f"  {team}...")
        build_team_sheet(wb, team, all_rows, all_teams)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
