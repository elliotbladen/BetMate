#!/usr/bin/env python3
"""Build historical EFL Championship 1X2 and O/U 2.5 confluence workbooks.

The matrices compare actual outcomes with de-vigged closing market probabilities.
Default window: five recent development seasons. The sealed 2025/26 holdout is
deliberately excluded.
"""
from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from statistics import mean

try:
    import ephem
except ImportError:
    ephem = None

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "ml/football/data/championship/matches/championship_matches.csv"
DEFAULT_OUTPUT_DIR = ROOT / "outputs/football/championship"
DEFAULT_SEASONS = ("2020/21", "2021/22", "2022/23", "2023/24", "2024/25")
CURRENT_2026_27_TEAMS = (
    "Birmingham", "Blackburn", "Bolton", "Bristol City", "Burnley", "Cardiff",
    "Charlton", "Derby", "Lincoln", "Middlesbrough", "Millwall", "Norwich",
    "Portsmouth", "Preston", "QPR", "Sheffield United", "Southampton", "Stoke",
    "Swansea", "Watford", "West Brom", "West Ham", "Wolves", "Wrexham",
)
MIN_SAMPLE = 5
FLAG_EDGE_PP = 7.5

NAVY = PatternFill("solid", fgColor="17365D")
BLUE = PatternFill("solid", fgColor="2E75B6")
LIGHT = PatternFill("solid", fgColor="DCE6F1")
ALT = PatternFill("solid", fgColor="F3F6FA")
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
GREY = PatternFill("solid", fgColor="E7E6E6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9E1F2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def number(value: str | None) -> float | None:
    try:
        parsed = float(value or "")
        return parsed if parsed > 1.0 else None
    except (TypeError, ValueError):
        return None


def score(value: str | None) -> int | None:
    try:
        return int(float(value or ""))
    except (TypeError, ValueError):
        return None


def devig(odds: list[float]) -> list[float]:
    raw = [1.0 / value for value in odds]
    total = sum(raw)
    return [value / total for value in raw]


def parse_date(value: str) -> date | None:
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    return None


def closing_1x2(raw: dict) -> tuple[list[float], str] | None:
    for columns, source in (
        (("PSCH", "PSCD", "PSCA"), "Pinnacle close"),
        (("MaxH", "MaxD", "MaxA"), "market maximum"),
    ):
        odds = [number(raw.get(column)) for column in columns]
        if all(value is not None for value in odds):
            return devig(odds), source
    return None


def closing_totals(raw: dict) -> tuple[list[float], str] | None:
    for columns, source in (
        (("PC>2.5", "PC<2.5"), "Pinnacle close"),
        (("Max>2.5", "Max<2.5"), "market maximum"),
    ):
        odds = [number(raw.get(column)) for column in columns]
        if all(value is not None for value in odds):
            return devig(odds), source
    return None


def _get_moon_phase(game_date: date) -> str | None:
    """Return 'new' / 'full' if within +/-1 day, else None.  Same logic as NRL/AFL matrix."""
    if ephem is None:
        return None
    try:
        d = ephem.Date(game_date.strftime("%Y/%m/%d 12:00:00"))
        closest_new = min(
            abs(float(d) - float(ephem.previous_new_moon(d))),
            abs(float(ephem.next_new_moon(d)) - float(d)),
        )
        closest_full = min(
            abs(float(d) - float(ephem.previous_full_moon(d))),
            abs(float(ephem.next_full_moon(d)) - float(d)),
        )
        if closest_new <= 1.0:
            return "new"
        if closest_full <= 1.0:
            return "full"
    except Exception:
        pass
    return None


def load_rows(path: Path, seasons: tuple[str, ...]) -> list[dict]:
    rows: list[dict] = []
    with path.open(encoding="utf-8-sig") as handle:
        for raw in csv.DictReader(handle):
            if raw.get("Season") not in seasons:
                continue
            played = parse_date((raw.get("Date") or "").strip())
            home, away = (raw.get("HomeTeam") or "").strip(), (raw.get("AwayTeam") or "").strip()
            home_goals, away_goals = score(raw.get("FTHG")), score(raw.get("FTAG"))
            result = (raw.get("FTR") or "").strip()
            market_1x2, market_totals = closing_1x2(raw), closing_totals(raw)
            if not played or not home or not away or result not in {"H", "D", "A"}:
                continue
            if home_goals is None or away_goals is None:
                continue
            moon = _get_moon_phase(played)
            row = {
                "season": raw["Season"], "date": played, "home": home, "away": away,
                "home_goals": home_goals, "away_goals": away_goals, "result": result,
                "referee": (raw.get("Referee") or "").strip(),
                "month": played.month, "weekday": played.weekday(),
                "market_1x2": market_1x2, "market_totals": market_totals,
                "moon_phase": moon,
            }
            rows.append(row)
    rows.sort(key=lambda row: row["date"])
    enrich_team_history(rows)
    return rows


def enrich_team_history(rows: list[dict]) -> None:
    team_games: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        team_games[row["home"]].append(row)
        team_games[row["away"]].append(row)
    for team, games in team_games.items():
        games.sort(key=lambda row: row["date"])
        previous = None
        for game in games:
            game[f"rest__{team}"] = (game["date"] - previous["date"]).days if previous else None
            if previous is None:
                game[f"previous__{team}"] = None
            elif previous["result"] == "D":
                game[f"previous__{team}"] = "D"
            else:
                won = (previous["home"] == team and previous["result"] == "H") or (
                    previous["away"] == team and previous["result"] == "A"
                )
                game[f"previous__{team}"] = "W" if won else "L"
            previous = game


def team_rows(rows: list[dict], team: str) -> list[dict]:
    return [row for row in rows if team in (row["home"], row["away"])]


def odds_band(row: dict, team: str) -> str | None:
    market = row["market_1x2"]
    if not market:
        return None
    probs = market[0]
    probability = probs[0] if row["home"] == team else probs[2]
    fair = 1.0 / probability
    if fair < 1.80:
        return "Short favourite (<1.80 fair)"
    if fair < 2.50:
        return "Favourite (1.80–2.49 fair)"
    if fair < 3.50:
        return "Competitive (2.50–3.49 fair)"
    return "Underdog (3.50+ fair)"


def reliability(n: int) -> str:
    if n >= 30:
        return "Strong sample"
    if n >= 15:
        return "Useful sample"
    if n >= MIN_SAMPLE:
        return "Small sample"
    return "Insufficient"


def stats_1x2(games: list[dict], team: str) -> tuple | None:
    usable = [game for game in games if game["market_1x2"]]
    if len(usable) < MIN_SAMPLE:
        return None
    wins, draws, p_win, p_draw = 0, 0, [], []
    sources = Counter()
    for game in usable:
        probs, source = game["market_1x2"]
        team_won = (game["home"] == team and game["result"] == "H") or (
            game["away"] == team and game["result"] == "A"
        )
        wins += int(team_won)
        draws += int(game["result"] == "D")
        p_win.append(probs[0] if game["home"] == team else probs[2])
        p_draw.append(probs[1])
        sources[source] += 1
    n = len(usable)
    actual_win, market_win = wins / n * 100, mean(p_win) * 100
    actual_draw, market_draw = draws / n * 100, mean(p_draw) * 100
    return (actual_win, market_win, actual_win - market_win,
            actual_draw, market_draw, actual_draw - market_draw,
            n, reliability(n), sources.most_common(1)[0][0])


def stats_goals(games: list[dict], league_btts: float) -> tuple | None:
    usable = [game for game in games if game["market_totals"]]
    if len(usable) < MIN_SAMPLE:
        return None
    actual = [int(game["home_goals"] + game["away_goals"] > 2) for game in usable]
    implied = [game["market_totals"][0][0] for game in usable]
    sources = Counter(game["market_totals"][1] for game in usable)
    actual_over, market_over = mean(actual) * 100, mean(implied) * 100
    n = len(usable)
    btts_games = [game for game in games if game["home_goals"] is not None and game["away_goals"] is not None]
    actual_btts = mean(int(game["home_goals"] > 0 and game["away_goals"] > 0) for game in btts_games) * 100
    return (actual_over, market_over, actual_over - market_over,
            100 - actual_over, 100 - market_over, market_over - actual_over,
            actual_btts, league_btts * 100, actual_btts - league_btts * 100, len(btts_games),
            n, reliability(n), sources.most_common(1)[0][0])


def categories(rows: list[dict], team: str, all_teams: list[str], all_referees: list[str]) -> list[tuple[str, str, list[dict]]]:
    games = team_rows(rows, team)
    output = [
        ("OVERALL", "All games", games),
        ("OVERALL", "Home games", [g for g in games if g["home"] == team]),
        ("OVERALL", "Away games", [g for g in games if g["away"] == team]),
    ]
    weekdays = ((0, "Monday"), (1, "Tuesday"), (2, "Wednesday"), (3, "Thursday"),
                (4, "Friday"), (5, "Saturday"), (6, "Sunday"))
    output += [("DAY OF WEEK", label, [g for g in games if g["weekday"] == day]) for day, label in weekdays]
    month_names = ((8, "August"), (9, "September"), (10, "October"), (11, "November"),
                   (12, "December"), (1, "January"), (2, "February"), (3, "March"),
                   (4, "April"), (5, "May"))
    output += [("MONTH", label, [g for g in games if g["month"] == month]) for month, label in month_names]
    output += [
        ("PREVIOUS RESULT", "After a win", [g for g in games if g.get(f"previous__{team}") == "W"]),
        ("PREVIOUS RESULT", "After a draw", [g for g in games if g.get(f"previous__{team}") == "D"]),
        ("PREVIOUS RESULT", "After a loss", [g for g in games if g.get(f"previous__{team}") == "L"]),
        ("REST", "Short rest (≤3 days)", [g for g in games if g.get(f"rest__{team}") is not None and g[f"rest__{team}"] <= 3]),
        ("REST", "Normal rest (4–8 days)", [g for g in games if g.get(f"rest__{team}") is not None and 4 <= g[f"rest__{team}"] <= 8]),
        ("REST", "Long rest (≥9 days)", [g for g in games if g.get(f"rest__{team}") is not None and g[f"rest__{team}"] >= 9]),
    ]
    bands = ("Short favourite (<1.80 fair)", "Favourite (1.80–2.49 fair)",
             "Competitive (2.50–3.49 fair)", "Underdog (3.50+ fair)")
    output += [("TEAM CLOSING PRICE BAND", band, [g for g in games if odds_band(g, team) == band]) for band in bands]
    if ephem is not None:
        output += [
            ("MOON PHASE", "New Moon (+/-1 day)", [g for g in games if g.get("moon_phase") == "new"]),
            ("MOON PHASE", "Full Moon (+/-1 day)", [g for g in games if g.get("moon_phase") == "full"]),
        ]
    output += [("REFEREE", referee, [g for g in games if g["referee"] == referee])
               for referee in all_referees]
    output += [("HEAD TO HEAD", f"vs {opponent}", [g for g in games if opponent in (g["home"], g["away"])])
               for opponent in all_teams if opponent != team]
    return output


def setup_sheet(ws, title: str, headers: list[str]) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(1, 1, title)
    cell.fill, cell.font, cell.alignment = NAVY, Font(color="FFFFFF", bold=True, size=12), Alignment(horizontal="center")
    for column, header in enumerate(headers, 1):
        c = ws.cell(2, column, header)
        c.fill, c.font, c.alignment, c.border = NAVY, WHITE_FONT, Alignment(horizontal="center", wrap_text=True), BORDER
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(headers))}2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 34
    for column in range(3, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 17


def write_matrix_sheet(
    ws, rows: list[dict], team: str, all_teams: list[str], all_referees: list[str],
    market: str, seasons: tuple[str, ...], league_name: str = "EFL Championship",
) -> None:
    if market == "1x2":
        headers = ["Section", "Category", "Actual win %", "Market win %", "Win edge pp",
                   "Actual draw %", "Market draw %", "Draw edge pp", "N", "Reliability", "Main odds source"]
        calculator = lambda games: stats_1x2(games, team)
    else:
        headers = ["Section", "Category", "Actual over %", "Market over %", "Over edge pp",
                   "Actual under %", "Market under %", "Under edge pp", "N", "Reliability", "Main odds source"]
        headers = ["Section", "Category", "Actual over %", "Market over %", "Over edge pp",
                   "Actual under %", "Market under %", "Under edge pp",
                   "Actual BTTS %", "League BTTS %", "BTTS difference pp", "BTTS N",
                   "Totals N", "Reliability", "Main odds source"]
        league_btts = mean(int(row["home_goals"] > 0 and row["away_goals"] > 0) for row in rows)
        calculator = lambda games: stats_goals(games, league_btts)
    setup_sheet(ws, f"{team} — {league_name} {market.upper()} confluence ({', '.join(seasons)})", headers)
    for row_number, (section, label, games) in enumerate(categories(rows, team, all_teams, all_referees), 3):
        values = calculator(games)
        ws.cell(row_number, 1, section)
        ws.cell(row_number, 2, label)
        if values is None:
            for column in range(3, len(headers) + 1):
                ws.cell(row_number, column, "—")
        else:
            for column, value in enumerate(values, 3):
                ws.cell(row_number, column, round(value, 1) if isinstance(value, float) else value)
            for edge_column in ((5, 8) if market == "1x2" else (5, 8, 11)):
                edge = ws.cell(row_number, edge_column)
                if isinstance(edge.value, (int, float)) and abs(edge.value) >= FLAG_EDGE_PP:
                    edge.fill = GREEN if edge.value > 0 else AMBER
        for column in range(1, len(headers) + 1):
            cell = ws.cell(row_number, column)
            cell.border = BORDER
            if not cell.fill.fill_type:
                cell.fill = ALT if row_number % 2 else PatternFill("solid", fgColor="FFFFFF")
            cell.alignment = Alignment(horizontal="left" if column <= 2 else "center")
        if row_number == 3 or ws.cell(row_number - 1, 1).value != section:
            ws.cell(row_number, 1).fill = BLUE
            ws.cell(row_number, 1).font = WHITE_FONT


def add_readme(
    wb, market: str, rows: list[dict], seasons: tuple[str, ...],
    league_name: str = "EFL Championship",
    holdout_note: str = "The sealed 2025/26 model-evaluation vault is excluded.",
) -> None:
    ws = wb.create_sheet("README", 0)
    lines = [
        (f"{league_name} historical confluence matrix", True),
        (f"Market: {'1X2 team win/draw' if market == '1x2' else 'Over/Under 2.5 goals'}", False),
        (f"Seasons: {', '.join(seasons)}", False),
        (holdout_note, False),
        (f"Source matches loaded: {len(rows)}", False),
        ("Market probabilities are normalized/de-vigged before comparison.", False),
        (f"Rows require at least {MIN_SAMPLE} matches; 5–14 is explicitly a small sample.", False),
        (f"Cells highlight when actual-minus-market differs by at least {FLAG_EDGE_PP:.1f} percentage points.", False),
        ("Signals overlap and are not independent. Multiple green cells do not imply additive edge.", False),
        ("This is descriptive historical evidence, not a standalone betting recommendation.", False),
        ("Kickoff-time splits are excluded because the consolidated source does not retain reliable times.", False),
        (f"BTTS difference is versus the {league_name} base rate, not a de-vigged BTTS market.", False),
    ]
    for row, (value, title) in enumerate(lines, 1):
        ws.cell(row, 1, value)
        if title:
            ws.cell(row, 1).fill, ws.cell(row, 1).font = NAVY, Font(color="FFFFFF", bold=True, size=13)
    ws.column_dimensions["A"].width = 105


def build_workbook(
    rows: list[dict], teams: list[str], seasons: tuple[str, ...], market: str,
    output: Path, league_name: str = "EFL Championship",
    holdout_note: str = "The sealed 2025/26 model-evaluation vault is excluded.",
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_readme(wb, market, rows, seasons, league_name, holdout_note)
    all_teams = sorted(set(row["home"] for row in rows) | set(row["away"] for row in rows))
    all_referees = sorted(set(row["referee"] for row in rows if row["referee"]))
    for team in teams:
        ws = wb.create_sheet(team[:31])
        write_matrix_sheet(ws, rows, team, all_teams, all_referees, market, seasons, league_name)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--seasons", nargs="+", default=list(DEFAULT_SEASONS))
    args = parser.parse_args()
    seasons = tuple(args.seasons)
    rows = load_rows(args.source, seasons)
    historical_teams = set(row["home"] for row in rows) | set(row["away"] for row in rows)
    teams = sorted(historical_teams | set(CURRENT_2026_27_TEAMS))
    print(f"Loaded {len(rows)} matches across {len(teams)} teams: {', '.join(seasons)}")
    one_x_two = args.output_dir / "efl_championship_1x2_confluence_matrix.xlsx"
    goals = args.output_dir / "efl_championship_goals_confluence_matrix.xlsx"
    build_workbook(rows, teams, seasons, "1x2", one_x_two)
    build_workbook(rows, teams, seasons, "goals", goals)
    print(f"Saved {one_x_two}")
    print(f"Saved {goals}")


if __name__ == "__main__":
    main()
