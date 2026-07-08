"""
NRL Totals Matrix -- Penrith Panthers vs New Zealand Warriors
R13 2026 -- Sunday 31 May, CommBank Stadium

Model total: 44.5 (5-10pt NRL bias means fair line ~44-46)
Market total: 48.5 (4pt gap -- model UNDER lean)

Panthers: 14 days rest (bye R12), prev WIN (beat St George 28-6, May 17)
Warriors: 14 days rest (bye R12), prev WIN (beat Brisbane 42-12, May 17)
Warriors travel 2182km from NZ.

May 31 = full moon (99.7% illumination, is_full_moon = True)
"""

import os
from datetime import datetime, timedelta, date
from collections import defaultdict

import ephem
import openpyxl

SOURCE_PATH    = os.path.join("C:\\Users\\ElliotBladen\\Apps", "data", "nrl", "historical", "latest.xlsx")
SEASONS        = (2022, 2023, 2024, 2025, 2026)
MIN_SAMPLE     = 5
EDGE_THRESHOLD = 10.0

GAME_DATE      = date(2026, 5, 31)
GAME_TIME      = datetime(2026, 5, 31, 14, 5)   # Sunday afternoon NRL
HOME_TEAM      = "Penrith Panthers"
AWAY_TEAM      = "New Zealand Warriors"
VENUE          = "CommBank Stadium"
MARKET_TOTAL   = 48.5

PANTHERS_PREV_WIN = True   # beat St George Dragons 28-6 (May 17)
WARRIORS_PREV_WIN = True   # beat Brisbane Broncos 42-12 (May 17)
PANTHERS_REST     = 14     # May 17 -> May 31
WARRIORS_REST     = 14     # May 17 -> May 31


def moon_phase(target: date, window: int = 1):
    d = ephem.Date(target - timedelta(days=30))
    end = ephem.Date(target + timedelta(days=30))
    new_moons, full_moons = set(), set()
    while d < end:
        nm = ephem.Date(ephem.next_new_moon(d)).datetime().date()
        fm = ephem.Date(ephem.next_full_moon(d)).datetime().date()
        for delta in range(-window, window + 1):
            new_moons.add(nm + timedelta(days=delta))
            full_moons.add(fm + timedelta(days=delta))
        d = ephem.next_new_moon(d) + 1
    observer = ephem.Observer()
    observer.date = ephem.Date(target)
    moon = ephem.Moon(observer)
    def next_moon(func, start):
        return ephem.Date(func(ephem.Date(start))).datetime().date()
    return {
        "phase_pct": round(moon.phase, 1),
        "is_new_moon": target in new_moons,
        "is_full_moon": target in full_moons,
        "nearest_new":  next_moon(ephem.next_new_moon,  target - timedelta(days=15)),
        "nearest_full": next_moon(ephem.next_full_moon, target - timedelta(days=15)),
    }


def load_data():
    wb = openpyxl.load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        gd = raw[0]
        if not gd or not hasattr(gd, "year") or gd.year not in SEASONS:
            continue
        hs, as_ = raw[5], raw[6]
        total_close = raw[40]   # Total Score Close
        over_odds   = raw[44]   # Total Score Over Close
        under_odds  = raw[48]   # Total Score Under Close
        if hs is None or as_ is None or total_close is None or over_odds is None:
            continue

        game_date = gd.date() if hasattr(gd, "date") else gd
        ko = raw[1]
        if ko and hasattr(ko, "hour") and not hasattr(ko, "date"):
            dt = datetime.combine(game_date, ko)
        elif ko and hasattr(ko, "date"):
            dt = datetime.combine(game_date, ko.time())
        else:
            dt = datetime(game_date.year, game_date.month, game_date.day, 14, 0)

        actual_total = int(hs) + int(as_)
        market_line  = float(total_close)

        rows.append({
            "dt": dt,
            "game_date": game_date,
            "season": gd.year,
            "home_team": str(raw[2] or "").strip(),
            "away_team": str(raw[3] or "").strip(),
            "venue":     str(raw[4] or "Unknown").strip(),
            "home_score": int(hs),
            "away_score": int(as_),
            "actual_total": actual_total,
            "market_line":  market_line,
            "over_implied": round(1 / float(over_odds), 4),
            "actual_over":  1 if actual_total > market_line else 0,
            "is_final":     bool(raw[7]),
        })

    wb.close()
    rows.sort(key=lambda r: r["dt"])
    return rows


def enrich_rows(rows):
    dates = [r["game_date"] for r in rows]
    new_moon_dates, full_moon_dates = set(), set()
    d = ephem.Date(min(dates) - timedelta(days=30))
    end_e = ephem.Date(max(dates) + timedelta(days=30))
    while d < end_e:
        nm = ephem.Date(ephem.next_new_moon(d)).datetime().date()
        fm = ephem.Date(ephem.next_full_moon(d)).datetime().date()
        for delta in range(-1, 2):
            new_moon_dates.add(nm + timedelta(days=delta))
            full_moon_dates.add(fm + timedelta(days=delta))
        d = ephem.next_new_moon(d) + 1

    for r in rows:
        dt = r["dt"]
        wd = dt.weekday()
        r["is_night"]    = dt.hour >= 18
        r["is_day"]      = dt.hour < 18
        r["is_thu_fri"]  = wd in (3, 4)
        r["is_saturday"] = wd == 5
        r["is_sunday"]   = wd == 6
        r["is_new_moon"]  = r["game_date"] in new_moon_dates
        r["is_full_moon"] = r["game_date"] in full_moon_dates

    team_games: dict[str, list] = defaultdict(list)
    for r in rows:
        team_games[r["home_team"]].append(r)
        team_games[r["away_team"]].append(r)

    for team, games in team_games.items():
        games.sort(key=lambda x: x["dt"])
        for i, g in enumerate(games):
            if i == 0:
                g[f"rest__{team}"] = None
                g[f"prev_win__{team}"] = None
            else:
                prev = games[i - 1]
                g[f"rest__{team}"] = (g["game_date"] - prev["game_date"]).days
                g[f"prev_win__{team}"] = (
                    prev["home_score"] > prev["away_score"]
                ) if prev["home_team"] == team else (
                    prev["away_score"] > prev["home_score"]
                )
    return rows


def totals_stats(games):
    n = len(games)
    if n < MIN_SAMPLE:
        return None
    overs       = sum(g["actual_over"] for g in games)
    actual_pct  = overs / n * 100
    implied_pct = sum(g["over_implied"] for g in games) / n * 100
    diff        = actual_pct - implied_pct
    rel_edge    = abs(diff / implied_pct) * 100 if implied_pct else 0
    avg_total   = sum(g["actual_total"] for g in games) / n
    avg_line    = sum(g["market_line"] for g in games) / n
    return {
        "actual": round(actual_pct, 1),
        "implied": round(implied_pct, 1),
        "diff": round(diff, 1),
        "rel_edge": round(rel_edge, 1),
        "direction": "OVER" if diff > 0 else "UNDER",
        "n": n,
        "flag": rel_edge >= EDGE_THRESHOLD,
        "avg_total": round(avg_total, 1),
        "avg_line":  round(avg_line, 1),
        "avg_diff":  round(avg_total - avg_line, 1),
    }


def print_row(label, s):
    if s is None:
        print(f"  --     {label:<46}  n<{MIN_SAMPLE}")
        return
    flag = ">> " if s["flag"] else "   "
    edge_str = f"{s['rel_edge']:.0f}% {s['direction']}" if s["rel_edge"] >= 1 else "--"
    print(f"  {flag}  {label:<46}  over {s['actual']:5.1f}%  impl {s['implied']:5.1f}%  "
          f"diff {s['diff']:+5.1f}pp  edge {edge_str:<16}  "
          f"avg {s['avg_total']:.0f} vs line {s['avg_line']:.0f} ({s['avg_diff']:+.0f})  n={s['n']}")


def main():
    print("=" * 120)
    print(f"  NRL Totals Matrix -- {HOME_TEAM} vs {AWAY_TEAM}")
    print(f"  {GAME_DATE.strftime('%A %d %B %Y')} | {VENUE} | Market total: {MARKET_TOTAL}")
    print(f"  Rules: 44.5 (-4.0 below market) | NRL bias +5-10pts => fair line ~44-46 | Threshold: {EDGE_THRESHOLD:.0f}%+")
    print("=" * 120)

    moon = moon_phase(GAME_DATE)
    wd = GAME_TIME.weekday()
    print(f"\n  MOON: {moon['phase_pct']}% illumination | nearest full: {moon['nearest_full']} | "
          f"is_full={moon['is_full_moon']} | is_new={moon['is_new_moon']}")
    print(f"  DAY:  {GAME_TIME.strftime('%A')} afternoon (thu_fri={wd in (3,4)}, sat={wd==5}, sun={wd==6}, night={GAME_TIME.hour>=18})")
    print(f"  FORM: Panthers prev WIN ({PANTHERS_REST}d rest, bye R12) | Warriors prev WIN ({WARRIORS_REST}d rest, bye R12)")
    print(f"  NOTE: Warriors travel 2182km from NZ. Panthers home at CommBank (non-standard venue).")

    print("\n  Loading data...")
    rows = load_data()
    rows = enrich_rows(rows)
    print(f"  {len(rows)} games loaded.\n")

    for team in [HOME_TEAM, AWAY_TEAM]:
        role = "HOME" if team == HOME_TEAM else "AWAY"
        prev_win  = PANTHERS_PREV_WIN if team == HOME_TEAM else WARRIORS_PREV_WIN
        rest_days = PANTHERS_REST if team == HOME_TEAM else WARRIORS_REST

        g = [r for r in rows if r["home_team"] == team or r["away_team"] == team]

        print(f"\n{'=' * 120}")
        print(f"  {team.upper()} ({role})  -- >> = {EDGE_THRESHOLD:.0f}%+ edge | OVER/UNDER market total line")
        print(f"{'=' * 120}")

        flags = []

        def check(label, subset):
            s = totals_stats(subset)
            print_row(label, s)
            if s and s["flag"]:
                flags.append((label, s))
            return s

        print("\n  [OVERALL]")
        check("All games",       g)
        check("Home games",      [x for x in g if x["home_team"] == team])
        check("Away games",      [x for x in g if x["away_team"] == team])
        check("Regular season",  [x for x in g if not x["is_final"]])
        check("Finals",          [x for x in g if x["is_final"]])

        print("\n  [TIME OF DAY]")
        check("Night games (>=18:00)", [x for x in g if x["is_night"]])
        check("Day games (<18:00)",    [x for x in g if x["is_day"]])

        print("\n  [DAY OF WEEK]")
        check("Thu/Fri games",   [x for x in g if x["is_thu_fri"]])
        check("Saturday games",  [x for x in g if x["is_saturday"]])
        check("Sunday games",    [x for x in g if x["is_sunday"]])

        print("\n  [FORM]")
        kr = f"prev_win__{team}"
        check("After a win",     [x for x in g if x.get(kr) is True])
        check("After a loss",    [x for x in g if x.get(kr) is False])

        print("\n  [REST]")
        krest = f"rest__{team}"
        check("Short rest (<=6d)",   [x for x in g if x.get(krest) is not None and x[krest] <= 6])
        check("Normal rest (7-9d)",  [x for x in g if x.get(krest) is not None and 7 <= x[krest] <= 9])
        check("Long rest (>=10d)",   [x for x in g if x.get(krest) is not None and x[krest] >= 10])

        print("\n  [MOON]")
        check("New moon (+-1d)",  [x for x in g if x["is_new_moon"]])
        check("Full moon (+-1d)", [x for x in g if x["is_full_moon"]])

        print("\n  [H2H]")
        opp = AWAY_TEAM if team == HOME_TEAM else HOME_TEAM
        h2h = [x for x in g if x["home_team"] == opp or x["away_team"] == opp]
        check(f"vs {opp} (all H2H)",  h2h)
        check(f"vs {opp} as home",    [x for x in h2h if x["home_team"] == team])
        check(f"vs {opp} as away",    [x for x in h2h if x["away_team"] == team])

        print("\n  [VENUE]")
        venue_games = [x for x in g if VENUE.lower() in (x["venue"] or "").lower()]
        check(f"At {VENUE}",     venue_games)

        print("\n  [MONTH]")
        check("May games",       [x for x in g if x["dt"].month == 5])
        check("June games",      [x for x in g if x["dt"].month == 6])

        applicable_map = {
            "Night games (>=18:00)": GAME_TIME.hour >= 18,
            "Day games (<18:00)":    GAME_TIME.hour < 18,
            "Thu/Fri games":         wd in (3, 4),
            "Saturday games":        wd == 5,
            "Sunday games":          wd == 6,
            "After a win":           prev_win is True,
            "After a loss":          prev_win is False,
            "Short rest (<=6d)":     rest_days <= 6,
            "Normal rest (7-9d)":    7 <= rest_days <= 9,
            "Long rest (>=10d)":     rest_days >= 10,
            "Home games":            team == HOME_TEAM,
            "Away games":            team == AWAY_TEAM,
            "Regular season":        True,
            "May games":             True,
            f"vs {opp} (all H2H)":   True,
            f"vs {opp} as home":     team == HOME_TEAM,
            f"vs {opp} as away":     team == AWAY_TEAM,
            f"At {VENUE}":           team == HOME_TEAM,
            "New moon (+-1d)":       moon["is_new_moon"],
            "Full moon (+-1d)":      moon["is_full_moon"],
        }

        applicable = [(lbl, s) for lbl, s in flags if applicable_map.get(lbl, False)]

        print(f"\n  -- FLAGGED ({len(flags)} total) --")
        if not flags:
            print("    None.")
        else:
            for lbl, s in flags:
                mark = "  OK" if applicable_map.get(lbl, False) else "  --"
                print(f"  {mark}  {lbl:<46} {s['rel_edge']:.0f}% {s['direction']}  "
                      f"over {s['actual']:.1f}% vs impl {s['implied']:.1f}%  "
                      f"avg {s['avg_total']:.0f} vs line {s['avg_line']:.0f} ({s['avg_diff']:+.0f})  n={s['n']}")

        over_flags  = [(l, s) for l, s in applicable if s["direction"] == "OVER"]
        under_flags = [(l, s) for l, s in applicable if s["direction"] == "UNDER"]

        print(f"\n  -- APPLICABLE: {len(applicable)} edges ({len(over_flags)} OVER / {len(under_flags)} UNDER) --")
        for lbl, s in applicable:
            print(f"    OK  {lbl:<46} {s['rel_edge']:.0f}% {s['direction']}  "
                  f"avg {s['avg_total']:.0f} vs line {s['avg_line']:.0f}  n={s['n']}")

        print(f"\n  CONFLUENCE: {len(over_flags)} OVER flags / {len(under_flags)} UNDER flags")
        if len(under_flags) >= 3 and len(over_flags) == 0:
            print(f"  ** TRIPLE UNDER CONFLUENCE -- {team} has {len(under_flags)} applicable UNDER edges!")
        elif len(over_flags) >= 3 and len(under_flags) == 0:
            print(f"  ** TRIPLE OVER CONFLUENCE -- {team} has {len(over_flags)} applicable OVER edges!")
        elif len(under_flags) >= 3 or len(over_flags) >= 3:
            print(f"  ** PARTIAL CONFLUENCE -- dominant: {'OVER' if len(over_flags)>len(under_flags) else 'UNDER'} ({max(len(over_flags),len(under_flags))} vs {min(len(over_flags),len(under_flags))})")
        else:
            print(f"  No triple confluence.")
        print()


if __name__ == "__main__":
    main()
