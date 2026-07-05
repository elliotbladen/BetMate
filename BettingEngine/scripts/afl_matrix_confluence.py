#!/usr/bin/env python3
"""
scripts/afl_matrix_confluence.py — AFL T9 Matrix Confluence Analyser

Mirrors scripts/matrix_confluence.py for AFL. Reads the AFL round prep fixture,
loads all 3 AFL matrices (XLSX), derives team context from AFL historical results,
and flags games where 3+ applicable matrix edges of 20%+ point same direction.

RESEARCH/FLAGGING only — not wired into AFL pricing yet.

Usage:
    python scripts/afl_matrix_confluence.py --season 2026 --round 12
    python scripts/afl_matrix_confluence.py --min-edges 2 --min-edge-pct 15
    python scripts/afl_matrix_confluence.py --push
"""

import argparse
import csv
import json
import sys
from calendar import month_name
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.supabase_push import push as _sb_push, load_env as _load_env

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

SCRIPTS_DIR  = Path(__file__).resolve().parent
ENGINE_ROOT  = SCRIPTS_DIR.parent
OUTPUTS_DIR  = ENGINE_ROOT / 'outputs'
HIST_XLSX    = OUTPUTS_DIR / 'afl_weekly_review' / 'historical' / 'latest.xlsx'

MIN_EDGE  = 20.0
MIN_COUNT = 3

# Per-market overrides — AFL totals edges are structurally smaller (venue rotation
# means fewer games per venue row, so even a 5% edge is meaningful).
MIN_EDGE_BY_MARKET  = {'totals': 5.0}   # h2h/handicap stay at MIN_EDGE
MIN_COUNT_BY_MARKET = {'totals': 3}     # h2h/handicap stay at MIN_COUNT

# ---------------------------------------------------------------------------
# Fixture name → matrix sheet name
# AFL fixture CSVs use full "mascot" names; matrices use short forms.
# ---------------------------------------------------------------------------

FIXTURE_TO_MATRIX = {
    "Adelaide Crows":                  "Adelaide",
    "Brisbane Lions":                  "Brisbane",
    "Carlton Blues":                   "Carlton",
    "Collingwood Magpies":             "Collingwood",
    "Essendon Bombers":                "Essendon",
    "Fremantle Dockers":               "Fremantle",
    "Geelong Cats":                    "Geelong",
    "Gold Coast Suns":                 "Gold Coast",
    "Greater Western Sydney Giants":   "GWS Giants",
    "Hawthorn Hawks":                  "Hawthorn",
    "Melbourne Demons":                "Melbourne",
    "North Melbourne Kangaroos":       "North Melbourne",
    "Port Adelaide Power":             "Port Adelaide",
    "Richmond Tigers":                 "Richmond",
    "St Kilda Saints":                 "St Kilda",
    "Sydney Swans":                    "Sydney",
    "West Coast Eagles":               "West Coast",
    "Western Bulldogs":                "Western Bulldogs",
}

# Historical XLSX short names → matrix keys (they already match)
HIST_TO_MATRIX = {
    "GWS Giants": "GWS Giants",
    "Western Bulldogs": "Western Bulldogs",
    "North Melbourne": "North Melbourne",
    "Port Adelaide": "Port Adelaide",
    "West Coast": "West Coast",
    "Gold Coast": "Gold Coast",
    "St Kilda": "St Kilda",
}  # all others are same as-is


def fixture_to_key(name: str) -> str:
    return FIXTURE_TO_MATRIX.get(name, name)


# ---------------------------------------------------------------------------
# Matrix / fixture loaders
# ---------------------------------------------------------------------------

def _parse_edge(val):
    """'6.2% opposing' | '20.6% covers' | '3.2% unders edge' → (6.2, 'word') | None."""
    if not val:
        return None
    s = str(val).strip()
    if s in ('—', '-', ''):
        return None
    import re
    m = re.match(r'^([\d.]+)%\s+(.+)$', s)
    if not m:
        return None
    return float(m.group(1)), m.group(2).strip()


def load_xlsx_matrix(path: Path) -> dict:
    """Returns {sheet_name: {category_string: (edge_pct, direction_string)}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    result = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        data = {}
        for row in rows[2:]:
            cat = row[0]
            raw = row[4] if len(row) > 4 else None
            if not cat:
                continue
            data[str(cat)] = _parse_edge(raw)
        result[sheet] = data
    return result


def load_fixture_csv(path: Path) -> list[dict]:
    games = []
    with open(path, newline='', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            games.append({
                'home_team': row['home_team'].strip(),
                'away_team': row['away_team'].strip(),
                'venue':     row['venue'].strip(),
                'date':      row['date'].strip(),
            })
    return games


def load_afl_history(path: Path) -> list[dict]:
    """
    Returns list of {date, home_team, away_team, home_score, away_score}
    from the aussportsbetting XLSX. Team names are short matrix-key forms.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    games = []
    for row in ws.iter_rows(values_only=True):
        if not isinstance(row[0], datetime):
            continue
        if row[2] is None or row[3] is None:
            continue
        try:
            hs = int(row[5]) if row[5] is not None else None
            aws = int(row[6]) if row[6] is not None else None
        except (TypeError, ValueError):
            hs, aws = None, None
        games.append({
            'date':       row[0].date(),
            'home_team':  str(row[2]).strip(),
            'away_team':  str(row[3]).strip(),
            'home_score': hs,
            'away_score': aws,
        })
    return games


def get_afl_team_context(history: list[dict], matrix_key: str, game_date: date) -> dict:
    """
    Returns {'rest_days': int|None, 'last_result': 'win'|'loss'|'draw'|None}
    by finding the team's most recent game before game_date in history.
    """
    last_game = None
    for g in history:
        if g['date'] >= game_date:
            continue
        if g['home_team'] == matrix_key or g['away_team'] == matrix_key:
            if last_game is None or g['date'] > last_game['date']:
                last_game = g

    if last_game is None:
        return {'rest_days': None, 'last_result': None}

    rest_days = (game_date - last_game['date']).days

    hs  = last_game['home_score']
    aws = last_game['away_score']
    if hs is None or aws is None:
        last_result = None
    elif last_game['home_team'] == matrix_key:
        last_result = 'win' if hs > aws else ('loss' if hs < aws else 'draw')
    else:
        last_result = 'win' if aws > hs else ('loss' if aws < hs else 'draw')

    return {'rest_days': rest_days, 'last_result': last_result}


# ---------------------------------------------------------------------------
# Row selection
# ---------------------------------------------------------------------------

MONTH_LABELS = {
    3: 'March', 4: 'April', 5: 'May', 6: 'June',
    7: 'July',  8: 'August', 9: 'September', 10: 'October',
}


def get_moon_phase_row(game_date: date) -> str | None:
    """Return 'New Moon (±1 day)' / 'Full Moon (±1 day)' or None."""
    try:
        import ephem
        d = ephem.Date(game_date.strftime('%Y/%m/%d 12:00:00'))
        prev_new  = ephem.previous_new_moon(d)
        next_new  = ephem.next_new_moon(d)
        prev_full = ephem.previous_full_moon(d)
        next_full = ephem.next_full_moon(d)
        closest_new  = min(abs(float(d) - float(prev_new)),  abs(float(next_new)  - float(d)))
        closest_full = min(abs(float(d) - float(prev_full)), abs(float(next_full) - float(d)))
        if closest_new  <= 1.0:
            return 'New Moon (±1 day)'
        if closest_full <= 1.0:
            return 'Full Moon (±1 day)'
    except Exception:
        pass
    return None


def applicable_row_names(
    kickoff_hour: int,
    kickoff_weekday: int,
    game_date: date,
    team_role: str,
    home_matrix_key: str,
    away_matrix_key: str,
    venue: str,
    rest_days: int | None,
    last_result: str | None,
    generic_row: str,
) -> list[str]:
    rows = [generic_row]

    # Time of day
    if kickoff_hour >= 18:
        rows.append('Night Games (kick-off ≥ 18:00)')
    else:
        rows.append('Day Games (kick-off < 18:00)')

    # Day of week
    if kickoff_weekday in (3, 4):
        rows.append('Thursday / Friday Games')
    elif kickoff_weekday == 5:
        rows.append('Saturday Games')
    elif kickoff_weekday == 6:
        rows.append('Sunday Games')

    # Rest — AFL uses "Long Rest / Bye" label
    if rest_days is not None:
        if rest_days <= 6:
            rows.append('Short Rest (≤ 6 days)')
        elif rest_days >= 10:
            rows.append('Long Rest / Bye (≥ 10 days)')

    # Form
    if last_result == 'win':
        rows.append('After a Win')
    elif last_result == 'loss':
        rows.append('After a Loss')

    # Month
    month_label = MONTH_LABELS.get(game_date.month)
    if month_label:
        rows.append(month_label)

    # Moon phase
    moon_row = get_moon_phase_row(game_date)
    if moon_row:
        rows.append(moon_row)

    # Opponent
    opponent_key = away_matrix_key if team_role == 'home' else home_matrix_key
    rows.append(f'vs {opponent_key}')

    # Venue
    rows.append(venue)

    return rows


# ---------------------------------------------------------------------------
# Direction normalisation
# ---------------------------------------------------------------------------

def normalise_direction(raw_direction: str, team_role: str, market: str):
    d = raw_direction.lower().strip()

    if market == 'h2h':
        is_team_wins = ('backing' in d)
        is_home = (team_role == 'home')
        return 'HOME_WIN' if (is_team_wins == is_home) else 'AWAY_WIN'

    elif market == 'handicap':
        # "covers" = team covers, "fades" = team doesn't cover (opponent does)
        is_covers = ('covers' in d and 'fades' not in d)
        is_home   = (team_role == 'home')
        if is_home:
            return 'HOME_COVERS' if is_covers else 'AWAY_COVERS'
        else:
            return 'AWAY_COVERS' if is_covers else 'HOME_COVERS'

    elif market == 'totals':
        if 'overs' in d:
            return 'OVERS'
        elif 'unders' in d:
            return 'UNDERS'

    return None


# ---------------------------------------------------------------------------
# Core analysis
# ---------------------------------------------------------------------------

GENERIC_ROWS = {
    'h2h':      ('Win % — Home',        'Win % — Away'),
    'handicap': ('Cover Rate — Home',   'Cover Rate — Away'),
    'totals':   ('Total Points — Home', 'Total Points — Away'),
}


def analyse_game(
    home_team, away_team,
    home_matrix_key, away_matrix_key,
    venue, kickoff_hour, kickoff_weekday, game_date,
    home_ctx, away_ctx,
    h2h_matrix, totals_matrix, handicap_matrix,
    min_edge_by_market: dict | None = None,
    min_edge=MIN_EDGE,
) -> dict:
    results = {}

    for market, matrix in [('h2h', h2h_matrix), ('totals', totals_matrix),
                            ('handicap', handicap_matrix)]:
        threshold = (min_edge_by_market or {}).get(market, min_edge)
        buckets = defaultdict(list)
        home_generic, away_generic = GENERIC_ROWS[market]

        for team_key, team_label, team_role, generic_row, ctx in [
            (home_matrix_key, home_team, 'home', home_generic, home_ctx),
            (away_matrix_key, away_team, 'away', away_generic, away_ctx),
        ]:
            team_data = matrix.get(team_key, {})
            row_names = applicable_row_names(
                kickoff_hour, kickoff_weekday, game_date,
                team_role, home_matrix_key, away_matrix_key, venue,
                ctx['rest_days'], ctx['last_result'], generic_row,
            )
            for row_name in row_names:
                val = team_data.get(row_name)
                if not val:
                    continue
                edge_pct, raw_dir = val
                if edge_pct < threshold:
                    continue
                norm_dir = normalise_direction(raw_dir, team_role, market)
                if not norm_dir:
                    continue
                buckets[norm_dir].append((edge_pct, row_name, team_label))

        results[market] = dict(buckets)

    return results


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

DIR_LABELS = {
    'HOME_WIN':    'BACK HOME',
    'AWAY_WIN':    'BACK AWAY',
    'HOME_COVERS': 'HOME COVERS',
    'AWAY_COVERS': 'AWAY COVERS',
    'OVERS':       'OVERS',
    'UNDERS':      'UNDERS',
}


def print_report(games, results_by_game, min_count=MIN_COUNT, min_edge=MIN_EDGE,
                 min_count_by_market: dict | None = None, market_filter='all'):
    w = 88
    bar  = '=' * w
    thin = '-' * w

    market_label = f' [{market_filter.upper()} only]' if market_filter != 'all' else ''
    print(f'\n{bar}')
    print(f'  T9 MATRIX CONFLUENCE -- AFL  (threshold: {int(min_edge)}%+ edge, {min_count}+ signals same direction){market_label}')
    print(bar)

    confluence_found = 0

    for game, analysis in zip(games, results_by_game):
        flags = [
            (market, direction, edges)
            for market, buckets in analysis.items()
            if market_filter == 'all' or market == market_filter
            for direction, edges in buckets.items()
            if len(edges) >= (min_count_by_market or {}).get(market, min_count)
        ]
        if not flags:
            continue

        confluence_found += 1
        print(f'\n{thin}')
        print(f'  {game["home_team"]}  vs  {game["away_team"]}')
        print(f'  {game["date"]}  |  {game["venue"]}')
        print(f'  Context:')
        print(f'    HOME  rest={game["home_ctx"]["rest_days"]}d  form={game["home_ctx"]["last_result"] or "?"}')
        print(f'    AWAY  rest={game["away_ctx"]["rest_days"]}d  form={game["away_ctx"]["last_result"] or "?"}')
        print()

        for market, direction, edges in sorted(flags, key=lambda x: -len(x[2])):
            label = DIR_LABELS.get(direction, direction)
            print(f'  ** {market.upper():<10}  {len(edges)}-way confluence  -->  {label}')
            for edge_pct, row_name, team_label in sorted(edges, key=lambda x: -x[0]):
                short = team_label.split()[-1]
                print(f'       {edge_pct:5.1f}%  {short:<14}  {row_name}')
            print()

    if confluence_found == 0:
        print(f'\n  No games with {min_count}+ edges >= {int(min_edge)}% same direction this round.\n')
    else:
        print(f'\n{bar}')
        print(f'  {confluence_found} game(s) with {min_count}+ confluence signals found.')
    print(bar + '\n')


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--season',       type=int,   default=2026)
    ap.add_argument('--round',        type=int,   default=None)
    ap.add_argument('--min-edges',    type=int,   default=MIN_COUNT)
    ap.add_argument('--min-edge-pct', type=float, default=MIN_EDGE)
    ap.add_argument('--market', choices=['h2h', 'handicap', 'totals', 'all'],
                    default='all', help='Show only signals for this market')
    ap.add_argument('--push',         action='store_true')
    args = ap.parse_args()

    # ── Find fixture ──────────────────────────────────────────────────────
    if args.round:
        fixture_path = (ENGINE_ROOT / 'outputs' / 'afl_round_prep'
                        / f'r{args.round}_{args.season}' / f'fixture_r{args.round}_{args.season}.csv')
    else:
        # auto-detect: most recent r*_SEASON dir
        prep_root = ENGINE_ROOT / 'outputs' / 'afl_round_prep'
        dirs = sorted(prep_root.glob(f'r*_{args.season}'), key=lambda p: p.name)
        if not dirs:
            print('ERROR: No AFL round prep directories found.')
            sys.exit(1)
        latest = dirs[-1]
        rnum   = int(latest.name.split('_')[0][1:])
        fixture_path = latest / f'fixture_r{rnum}_{args.season}.csv'
        args.round   = rnum

    if not fixture_path.exists():
        print(f'ERROR: Fixture not found: {fixture_path}')
        sys.exit(1)

    games_raw = load_fixture_csv(fixture_path)
    print(f'\nLoading AFL R{args.round} fixture  ({len(games_raw)} games) ...')

    # ── Load matrices ─────────────────────────────────────────────────────
    print('Loading matrices ...')
    h2h_matrix      = load_xlsx_matrix(OUTPUTS_DIR / 'afl_h2h_matrix.xlsx')
    totals_matrix   = load_xlsx_matrix(OUTPUTS_DIR / 'afl_team_totals_matrix.xlsx')
    handicap_matrix = load_xlsx_matrix(OUTPUTS_DIR / 'afl_handicap_matrix.xlsx')

    # ── Load AFL history for team context ─────────────────────────────────
    print('Loading team context from AFL history ...')
    history = load_afl_history(HIST_XLSX)

    # ── Analyse each game ─────────────────────────────────────────────────
    games_enriched  = []
    results_by_game = []

    for g in games_raw:
        home_full = g['home_team']
        away_full = g['away_team']
        venue     = g['venue']
        date_str  = g['date']          # 'YYYY-MM-DD'

        game_date = date.fromisoformat(date_str)
        weekday   = game_date.weekday()  # 0=Mon ... 6=Sun
        month     = game_date.month

        # Estimate kickoff hour: Thu/Fri evenings, everything else afternoon
        if weekday in (3, 4):          # Thu / Fri
            kickoff_hour = 19
        elif weekday == 6:             # Sunday
            kickoff_hour = 13
        else:                          # Saturday — most are day games
            kickoff_hour = 14

        home_key = fixture_to_key(home_full)
        away_key = fixture_to_key(away_full)

        home_ctx = get_afl_team_context(history, home_key, game_date)
        away_ctx = get_afl_team_context(history, away_key, game_date)

        analysis = analyse_game(
            home_full, away_full,
            home_key, away_key,
            venue, kickoff_hour, weekday, game_date,
            home_ctx, away_ctx,
            h2h_matrix, totals_matrix, handicap_matrix,
            min_edge_by_market=MIN_EDGE_BY_MARKET,
            min_edge=args.min_edge_pct,
        )

        games_enriched.append({**g,
                                'home_matrix_key': home_key,
                                'away_matrix_key': away_key,
                                'home_ctx': home_ctx,
                                'away_ctx': away_ctx})
        results_by_game.append(analysis)

    # ── Print ─────────────────────────────────────────────────────────────
    print_report(games_enriched, results_by_game,
                 min_count=args.min_edges, min_edge=args.min_edge_pct,
                 min_count_by_market=MIN_COUNT_BY_MARKET,
                 market_filter=args.market)

    # ── Write local JSON ──────────────────────────────────────────────────
    def _meets_count(market, n):
        return n >= MIN_COUNT_BY_MARKET.get(market, args.min_edges)

    local_games = []
    for game, analysis in zip(games_enriched, results_by_game):
        flags = {}
        for market, buckets in analysis.items():
            for direction, edges in buckets.items():
                if _meets_count(market, len(edges)):
                    flags[f'{market}_{direction}'] = {
                        'count': len(edges),
                        'edges': [{'edge_pct': e, 'row': r, 'team': t} for e, r, t in edges],
                    }
        if flags:
            local_games.append({
                'home': game['home_team'],
                'away': game['away_team'],
                'confluence': flags,
            })
    out = OUTPUTS_DIR / 'afl_t9_confluence_latest.json'
    out.write_text(json.dumps({
        'season': args.season,
        'round':  args.round,
        'generated_at': datetime.now().isoformat(),
        'games':  local_games,
    }, indent=2), encoding='utf-8')
    print(f'  Confluence JSON written -> {out}')

    # ── Supabase push ─────────────────────────────────────────────────────
    if args.push:
        _load_env()
        payload = []
        for game, analysis in zip(games_enriched, results_by_game):
            flags = {}
            for market, buckets in analysis.items():
                for direction, edges in buckets.items():
                    if _meets_count(market, len(edges)):
                        flags[f'{market}_{direction}'] = [
                            {'edge_pct': e, 'row': r, 'team': t}
                            for e, r, t in edges
                        ]
            if flags:
                payload.append({
                    'home': game['home_team'],
                    'away': game['away_team'],
                    'venue': game['venue'],
                    'date':  game['date'],
                    'confluence': flags,
                })
        key = f'afl_t9_confluence_r{args.round}_{args.season}'
        if _sb_push(key, payload):
            print(f'  Pushed to Supabase: {key}')
        else:
            print('  Supabase push skipped (env vars not set)')


if __name__ == '__main__':
    main()
