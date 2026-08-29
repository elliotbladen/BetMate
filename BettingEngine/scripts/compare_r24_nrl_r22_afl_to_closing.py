#!/usr/bin/env python3
"""Compare the retained NRL R24 and AFL R22 model sheets with ASB closes."""

from __future__ import annotations

import csv
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
RUN_DATE = "2026-08-12"

CONFIGS = (
    {
        "sport": "NRL", "round": 24,
        "model_db": Path("/Users/elliotbladen/Betting_model/data/model.db"),
        "workbook": ROOT / "outputs" / "nrl_weekly_review" / "historical" / "latest.xlsx",
        "margin": "final_margin", "total": "final_total",
        "home_odds": "fair_home_odds", "away_odds": "fair_away_odds",
    },
    {
        "sport": "AFL", "round": 22,
        "pricing": ROOT / "results" / "r22_afl_2026.csv",
        "workbook": ROOT / "outputs" / "afl_weekly_review" / "historical" / "latest.xlsx",
        "margin": "rules_margin", "total": "rules_total",
        "home_odds": "rules_home_odds", "away_odds": "rules_away_odds",
    },
)


def number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def team_key(name):
    """Normalise the small naming differences between model sheets and ASB."""
    key = (name or "").lower().replace("-", " ").replace(".", "").strip()
    aliases = {
        "st george illawarra dragons": "st george dragons",
        "cronulla sutherland sharks": "cronulla sharks",
        "manly warringah sea eagles": "manly sea eagles",
        "canterbury bankstown bulldogs": "canterbury bulldogs",
        "north queensland cowboys": "north qld cowboys",
        "western bulldogs": "western bulldogs",
        "north melbourne kangaroos": "north melbourne",
        "brisbane lions": "brisbane",
        "hawthorn hawks": "hawthorn",
        "melbourne demons": "melbourne",
        "fremantle dockers": "fremantle",
        "sydney swans": "sydney",
        "port adelaide power": "port adelaide",
        "geelong cats": "geelong",
        "essendon bombers": "essendon",
        "adelaide crows": "adelaide",
        "richmond tigers": "richmond",
        "greater western sydney giants": "gws giants",
        "gold coast suns": "gold coast",
        "west coast eagles": "west coast",
        "collingwood magpies": "collingwood",
        "st kilda saints": "st kilda",
        "carlton blues": "carlton",
    }
    return aliases.get(key, key)


def load_market(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = [cell.value for cell in ws[2]]
    col = {name: index for index, name in enumerate(headers) if name}
    games = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        date = row[col["Date"]]
        if not hasattr(date, "year") or date.year != 2026:
            continue
        home, away = row[col["Home Team"]], row[col["Away Team"]]
        games[(team_key(home), team_key(away))] = {
            "date": date.date().isoformat(), "home": home, "away": away,
            "actual_margin": number(row[col["Home Score"]]) - number(row[col["Away Score"]]),
            "actual_total": number(row[col["Home Score"]]) + number(row[col["Away Score"]]),
            "close_home_odds": number(row[col["Home Odds Close"]]),
            "close_away_odds": number(row[col["Away Odds Close"]]),
            # ASB's home line is the wager line (e.g. -6.5). Invert to obtain
            # the market's expected home margin, comparable to model margin.
            "close_home_line": number(row[col["Home Line Close"]]),
            "close_total": number(row[col["Total Score Close"]]),
        }
    return games


def load_model_rows(cfg):
    if "pricing" in cfg:
        with cfg["pricing"].open(encoding="utf-8-sig") as file:
            return list(csv.DictReader(file))
    conn = sqlite3.connect(cfg["model_db"])
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT m.match_date, ht.team_name AS home_team, at.team_name AS away_team,
                   mr.final_margin, mr.final_total, mr.fair_home_odds, mr.fair_away_odds
            FROM model_runs mr
            JOIN matches m ON m.match_id = mr.match_id
            JOIN teams ht ON ht.team_id = m.home_team_id
            JOIN teams at ON at.team_id = m.away_team_id
            JOIN (
                SELECT mr2.match_id, MAX(mr2.run_timestamp) AS run_timestamp
                FROM model_runs mr2
                JOIN matches m2 ON m2.match_id = mr2.match_id
                WHERE m2.season = 2026 AND m2.round_number = ?
                  AND mr2.run_timestamp < m2.kickoff_datetime
                GROUP BY mr2.match_id
            ) latest ON latest.match_id = mr.match_id AND latest.run_timestamp = mr.run_timestamp
            WHERE m.season = 2026 AND m.round_number = ?
            ORDER BY m.match_date, m.kickoff_datetime
            """, (cfg["round"], cfg["round"])
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def side(expected, actual):
    if expected is None or actual is None:
        return ""
    return "home" if expected > actual else "away" if expected < actual else "push"


def main():
    output = []
    for cfg in CONFIGS:
        market = load_market(cfg["workbook"])
        for model in load_model_rows(cfg):
                home, away = model["home_team"], model["away_team"]
                game = market.get((team_key(home), team_key(away)))
                if not game:
                    raise ValueError(f"No ASB match for {cfg['sport']} R{cfg['round']}: {home} v {away}")
                model_margin = number(model.get(cfg["margin"]))
                model_total = number(model.get(cfg["total"]))
                model_home_odds = number(model.get(cfg["home_odds"]))
                model_away_odds = number(model.get(cfg["away_odds"]))
                market_margin = -game["close_home_line"] if game["close_home_line"] is not None else None
                model_hcap_side = side(model_margin, market_margin)
                actual_hcap_side = side(game["actual_margin"], market_margin)
                model_total_side = side(model_total, game["close_total"])
                actual_total_side = side(game["actual_total"], game["close_total"])
                model_favourite = "home" if model_home_odds < model_away_odds else "away"
                market_favourite = "home" if game["close_home_odds"] < game["close_away_odds"] else "away"
                actual_winner = "home" if game["actual_margin"] > 0 else "away" if game["actual_margin"] < 0 else "draw"
                output.append({
                    "sport": cfg["sport"], "round": cfg["round"], "date": game["date"],
                    "home_team": home, "away_team": away,
                    "model_margin": round(model_margin, 1), "market_margin": round(market_margin, 1),
                    "margin_edge_model_minus_close": round(model_margin - market_margin, 1),
                    "actual_margin": round(game["actual_margin"], 1),
                    "model_margin_abs_error": round(abs(model_margin - game["actual_margin"]), 1),
                    "market_margin_abs_error": round(abs(market_margin - game["actual_margin"]), 1),
                    "model_handicap_side": model_hcap_side, "actual_handicap_side": actual_hcap_side,
                    "model_handicap_correct": model_hcap_side == actual_hcap_side,
                    "model_total": round(model_total, 1), "close_total": game["close_total"],
                    "total_edge_model_minus_close": round(model_total - game["close_total"], 1),
                    "actual_total": round(game["actual_total"], 1),
                    "model_total_abs_error": round(abs(model_total - game["actual_total"]), 1),
                    "market_total_abs_error": round(abs(game["close_total"] - game["actual_total"]), 1),
                    "model_total_side": model_total_side, "actual_total_side": actual_total_side,
                    "model_total_correct": model_total_side == actual_total_side,
                    "model_home_odds": model_home_odds, "model_away_odds": model_away_odds,
                    "close_home_odds": game["close_home_odds"], "close_away_odds": game["close_away_odds"],
                    "model_favourite": model_favourite, "market_favourite": market_favourite,
                    "actual_winner": actual_winner,
                    "model_h2h_correct": model_favourite == actual_winner,
                    "market_h2h_correct": market_favourite == actual_winner,
                })

    out_dir = ROOT / "outputs" / "results"
    out_dir.mkdir(exist_ok=True)
    csv_path = out_dir / f"nrl_r24_afl_r22_model_vs_closing_{RUN_DATE}.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=output[0].keys())
        writer.writeheader(); writer.writerows(output)

    lines = ["# NRL R24 and AFL R22: model vs closing market", "",
             f"AusSportsBetting workbooks refreshed {RUN_DATE}. Margin is the expected home winning margin; positive edge means the model was higher on the home side than the close.", ""]
    for sport in ("NRL", "AFL"):
        rows = [r for r in output if r["sport"] == sport]
        if not rows:
            continue
        def pct(key): return f"{sum(bool(r[key]) for r in rows)}/{len(rows)} ({sum(bool(r[key]) for r in rows)/len(rows):.0%})"
        mae_model_margin = sum(r["model_margin_abs_error"] for r in rows) / len(rows)
        mae_market_margin = sum(r["market_margin_abs_error"] for r in rows) / len(rows)
        mae_model_total = sum(r["model_total_abs_error"] for r in rows) / len(rows)
        mae_market_total = sum(r["market_total_abs_error"] for r in rows) / len(rows)
        lines += [f"## {sport} R{rows[0]['round']}", "",
                  f"- H2H winner: model {pct('model_h2h_correct')}; closing favourite {pct('market_h2h_correct')}",
                  f"- Handicap side: model {pct('model_handicap_correct')}; MAE model {mae_model_margin:.1f}, close {mae_market_margin:.1f}",
                  f"- Total side: model {pct('model_total_correct')}; MAE model {mae_model_total:.1f}, close {mae_market_total:.1f}", "",
                  "| Match | Margin (model / close / actual) | Total (model / close / actual) | H2H (model / market / winner) |",
                  "|---|---:|---:|---|"]
        for r in rows:
            lines.append(f"| {r['home_team']} v {r['away_team']} | {r['model_margin']:+.1f} / {r['market_margin']:+.1f} / {r['actual_margin']:+.1f} | {r['model_total']:.1f} / {r['close_total']:.1f} / {r['actual_total']:.0f} | {r['model_favourite']} / {r['market_favourite']} / {r['actual_winner']} |")
        lines.append("")
    md_path = out_dir / f"nrl_r24_afl_r22_model_vs_closing_{RUN_DATE}.md"
    md_path.write_text("\n".join(lines), encoding="utf-8")
    print(csv_path)
    print(md_path)


if __name__ == "__main__":
    main()
