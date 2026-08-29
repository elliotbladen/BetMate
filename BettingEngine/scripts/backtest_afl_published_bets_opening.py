#!/usr/bin/env python3
"""Audit published AFL pre-match H2H/handicap bets against consensus openers."""

from __future__ import annotations

import argparse
import re
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
PAGE_DATA = ROOT.parent / "lib" / "researchData.ts"
WORKBOOK = ROOT / "outputs/afl_weekly_review/historical/latest.xlsx"


def team(value: str) -> str:
    value = re.sub(r"[^a-z]+", " ", value.lower()).strip()
    aliases = {
        "port adelaide": ("port adelaide", "power"),
        "north melbourne": ("north melbourne", "kangaroos", "nm"),
        "western bulldogs": ("western bulldogs", "bulldogs"),
        "adelaide": ("adelaide", "crows"), "brisbane": ("brisbane", "lions"),
        "carlton": ("carlton", "blues"), "collingwood": ("collingwood", "magpies"),
        "essendon": ("essendon", "bombers"), "fremantle": ("fremantle", "dockers"),
        "geelong": ("geelong", "cats"), "gold coast": ("gold coast", "suns"),
        "gws": ("gws", "giants"), "hawthorn": ("hawthorn", "hawks"),
        "melbourne": ("melbourne", "demons"), "richmond": ("richmond",),
        "st kilda": ("st kilda", "saints"), "sydney": ("sydney", "swans"),
        "west coast": ("west coast", "eagles"),
    }
    for canonical, names in aliases.items():
        if any(name in value for name in names):
            return canonical
    return value


def published_bets() -> list[dict]:
    source = PAGE_DATA.read_text(encoding="utf-8-sig")
    block = source.split("export const AFL_MODEL_BETS: ModelBet[] = [", 1)[1].split("];", 1)[0]
    rows=[]
    for body in re.findall(r"\{\s*id:.*?\}", block, re.S):
        def text(field):
            m=re.search(rf"{field}:'([^']*)'",body); return m.group(1) if m else ""
        def number(field):
            m=re.search(rf"{field}:(null|-?\d+(?:\.\d+)?)",body)
            return None if not m or m.group(1)=="null" else float(m.group(1))
        rows.append({"id":int(number("id")),"date":text("date"),"match":text("match"),
            "market_text":text("market"),"taken_odds":number("takenPrice"),"result_page":text("result")})
    return rows


def games() -> list[dict]:
    wb=openpyxl.load_workbook(WORKBOOK,read_only=True,data_only=True);ws=wb["Data"]
    headers=list(next(ws.iter_rows(min_row=2,max_row=2,values_only=True)))
    rows=[dict(zip(headers,v)) for v in ws.iter_rows(min_row=3,values_only=True)];wb.close()
    return [r for r in rows if getattr(r.get("Date"),"year",None)==2026]


def main():
    ap=argparse.ArgumentParser();ap.add_argument("--out",type=Path,
        default=ROOT/"outputs/results/afl_published_bets_opening_backtest_2026.csv");args=ap.parse_args()
    historical=games();out=[]
    for bet in published_bets():
        label=bet["market_text"];lower=label.lower()
        if any(x in lower for x in ("live","pyl","multi","cash out","2nd half")):continue
        kind="h2h" if ("win" in lower or lower.startswith("game ")) else ("handicap" if re.search(r"[+-]\d",label) else "")
        if not kind:continue
        names=re.split(r"\s+vs\s+",bet["match"],flags=re.I)
        if len(names)!=2:continue
        wanted={team(x) for x in names};candidates=[g for g in historical if {team(str(g["Home Team"])),team(str(g["Away Team"]))}==wanted]
        if not candidates:continue
        bd=date.fromisoformat(bet["date"]);game=min(candidates,key=lambda g:abs((g["Date"].date()-bd).days))
        if abs((game["Date"].date()-bd).days)>8:continue
        home,away=team(str(game["Home Team"])),team(str(game["Away Team"]))
        if lower.startswith("game "):
            # Old ledger labels omitted the team; infer it from the quoted price.
            known={1:"collingwood",2:"carlton",10:"st kilda"}
            options={home:float(game["Home Odds Open"]),away:float(game["Away Odds Open"])}
            selection=known.get(bet["id"],min(options,key=lambda side:abs(options[side]-bet["taken_odds"])))
        else:
            selection=team(re.sub(r"\s+(?:Win|[+-]\d.*)$","",label,flags=re.I))
        side="home" if selection==home else "away" if selection==away else ""
        if not side:continue
        hs,aws=float(game["Home Score"]),float(game["Away Score"])
        if kind=="h2h":
            opening_line=None;opening_odds=float(game["Home Odds Open"] if side=="home" else game["Away Odds Open"])
            won=hs>aws if side=="home" else aws>hs;push=False
            edge=(bet["taken_odds"]/opening_odds-1)*100
        else:
            opening_line=float(game["Home Line Open"] if side=="home" else game["Away Line Open"])
            opening_odds=float(game["Home Line Odds Open"] if side=="home" else game["Away Line Odds Open"])
            taken_line=float(re.search(r"([+-]\d+(?:\.\d+)?)",label).group(1));edge=taken_line-opening_line
            adjusted=(hs if side=="home" else aws)+opening_line-(aws if side=="home" else hs)
            push=adjusted==0;won=adjusted>0
        profit_taken=0 if bet["result_page"]=="push" else bet["taken_odds"]-1 if bet["result_page"]=="win" else -1
        profit_open=0 if push else opening_odds-1 if won else -1
        out.append({**bet,"market":kind,"selection":selection,"game_date":game["Date"].date(),
            "opening_line":opening_line,"opening_odds":opening_odds,"edge_obtained":edge,
            "opening_result":"push" if push else "win" if won else "loss",
            "profit_$1_taken":profit_taken,"profit_$1_at_open":profit_open})
    df=pd.DataFrame(out);args.out.parent.mkdir(parents=True,exist_ok=True);df.to_csv(args.out,index=False)
    print(f"Matched {len(df)} pre-match tickets")
    for market,g in df.groupby("market"):
        print(market,len(g),round(g['profit_$1_taken'].sum(),2),round(g['profit_$1_taken'].mean()*100,2),
              round(g['profit_$1_at_open'].sum(),2),round(g['profit_$1_at_open'].mean()*100,2))
        print(g.groupby(pd.cut(g.edge_obtained,[-float('inf'),0,2,7,float('inf')],right=False))["profit_$1_taken"].agg(['count','sum','mean']))


if __name__=="__main__":main()
