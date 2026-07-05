"""
scripts/push_matrices_to_supabase.py

Reads NRL matrix files (xlsx + csv), converts to JSON matching the structure
that BetMate's lib/matrixEV.ts expects, and pushes to Supabase betmate_data_store.

Run after prepare_round.py completes (already wired into run_nrl_pricing.ps1).

Usage:
  python scripts/push_matrices_to_supabase.py
"""

from __future__ import annotations

import csv
import json
import logging
import os
import sys
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ── Env / Supabase ────────────────────────────────────────────────────────────

def _load_env() -> None:
    env_file = ROOT.parent / ".env.local"
    if not env_file.exists():
        return
    for line in env_file.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())


def _push(key: str, data: dict) -> bool:
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not service_key:
        log.warning("Supabase env vars not set — skipping push for %s", key)
        return False
    try:
        resp = requests.post(
            f"{url}/rest/v1/betmate_data_store",
            headers={
                "apikey": service_key,
                "Authorization": f"Bearer {service_key}",
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates",
            },
            data=json.dumps([{"key": key, "data": data}]),
            timeout=15,
        )
        resp.raise_for_status()
        log.info("Pushed %s (%d teams)", key, len(data))
        return True
    except Exception as exc:
        log.warning("Push failed for %s: %s", key, exc)
        return False


# ── Edge string parser (mirrors matrixEV.ts logic) ───────────────────────────

def _parse_edge(s: str | None) -> dict | None:
    if not s or s.strip() in ("", "—", "—"):
        return None
    import re
    m = re.match(r"^([\d.]+)%\s+(.+)$", s.strip())
    if not m:
        return None
    pct = float(m.group(1))
    return {"edgePct": pct, "direction": m.group(2).strip()}


# ── xlsx matrix parser ────────────────────────────────────────────────────────

def parse_xlsx_matrix(path: Path) -> dict:
    """Parse an NRL h2h or totals matrix xlsx.
    Returns: {teamName: {category: {edgePct, direction} | None}}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        sheet: dict = {}
        for row in rows[1:]:  # skip header
            if not row or not row[0]:
                continue
            category = str(row[0]).strip()
            if category == "Category":
                continue
            # Edge value is in column index 3 (0-based) = 4th column
            raw = row[3] if len(row) > 3 else None
            sheet[category] = _parse_edge(str(raw) if raw is not None else None)
        result[sheet_name] = sheet
    wb.close()
    return result


# ── handicap CSV parser ───────────────────────────────────────────────────────

def parse_handicap_csv(path: Path) -> dict:
    """Parse nrl_handicap_matrix.csv.
    Returns: {teamName: {category: {edgePct, direction}}}
    """
    result: dict = {}
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            team = row.get("team", "").strip()
            category = row.get("category", "").strip()
            if not team or not category:
                continue
            try:
                edge_pct = float(row.get("edge_pct", ""))
                direction = row.get("direction", "").strip()
            except ValueError:
                continue
            if not direction:
                continue
            if team not in result:
                result[team] = {}
            result[team][category] = {"edgePct": edge_pct, "direction": direction}
    return result


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    _load_env()

    files = {
        "nrl_h2h_matrix":      OUTPUTS / "nrl_h2h_matrix.xlsx",
        "nrl_totals_matrix":   OUTPUTS / "nrl_team_totals_matrix.xlsx",
        "nrl_handicap_matrix": OUTPUTS / "nrl_handicap_matrix.csv",
    }

    for key, path in files.items():
        if not path.exists():
            log.warning("File not found — skipping %s (%s)", key, path)
            continue
        log.info("Parsing %s ...", path.name)
        if path.suffix == ".xlsx":
            data = parse_xlsx_matrix(path)
        else:
            data = parse_handicap_csv(path)
        _push(key, data)

    log.info("Done.")


if __name__ == "__main__":
    main()
