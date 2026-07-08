"""
AFL Totals Matrix — Sydney Swans vs Richmond Tigers
R12 2026 — Saturday 30 May, SCG

For each team, under each condition, asks:
  "When [team] plays, does the actual total go OVER or UNDER the market line
   more often than the market odds implied?"

Edge = |actual_over_pct - market_implied_over_pct| / market_implied_over_pct * 100
Flags 20%+ relative edge.

Condition that applies to this game:
  - Sydney: home, Saturday, ~19:25 AEST night, coming off R11 win (?)
  - Richmond: away, Saturday, ~19:25 AEST night
  - H2H at SCG
  - Market total: 178.5
  - Model: rules 195.6 / ML 178.0 (ML at market; rules inflated by ELO dominance)
  - Sydney major outs: Gulden, Adams, Campbell, King (loaded in T5)
"""

import math
import os
from datetime import datetime, timedelta, date
from collections import defaultdict
from pathlib import Path

import ephem
import openpyxl

SOURCE_PATH = r"C:\Users\ElliotBladen\Apps\BettingEngine\outputs\afl_weekly_review\historical\latest.xlsx"
SEASONS     = (2022, 2023, 2024, 2025, 2026)
MIN_SAMPLE  = 5   # AFL: raise to 5 given noisier data
EDGE_THRESHOLD = 10.0

# ── Game details ─────────────────────────────────────────────────────────────
GAME_DATE     = date(2026, 5, 30)
GAME_TIME     = datetime(2026, 5, 30, 19, 25)   # ~19:25 AEST Saturday night
HOME_TEAM     = "Sydney"
AWAY_TEAM     = "Richmond"
VENUE         = "SCG"
MARKET_TOTAL  = 178.5

# Rest/form — need to check AFL R11 results
# Sydney: last played R11, won? Check below note.
# Richmond: last played R11, won? Check below note.
# R11 games not fully known — using None if uncertain
SYDNEY_PREV_WIN   = False  # R12 2026-05-23: lost to Geelong 80-107
RICHMOND_PREV_WIN = True   # R12 2026-05-22: beat Essendon 74-56
SYDNEY_REST       = 7      # May 23 → May 30
RICHMOND_REST     = 8      # May 22 → May 30


# ── Moon phase ────────────────────────────────────────────────────────────────
def moon_phase(target: date, window: int = 1):
    d = ephem.Date(target - timedelta(days=30))
    end = ephem.Date(target + timedelta(days=30))
    new_moons, full_moons = set(), set()
    while d < end:
        nm_date = ephem.Date(ephem.next_new_moon(d)).datetime().date()
        fm_date = ephem.Date(ephem.next_full_moon(d)).datetime().date()
        for delta in range(-window, window + 1):
            new_moons.add(nm_date + timedelta(days=delta))
            full_moons.add(fm_date + timedelta(days=delta))
        d = ephem.next_new_moon(d) + 1

    observer = ephem.Observer()
    observer.date = ephem.Date(target)
    moon = ephem.Moon(observer)
    phase_pct = moon.phase
    is_new  = target in new_moons
    is_full = target in full_moons

    def next_moon(func, start):
        return ephem.Date(func(ephem.Date(start))).datetime().date()

    nearest_new  = next_moon(ephem.next_new_moon,  target - timedelta(days=15))
    nearest_full = next_moon(ephem.next_full_moon, target - timedelta(days=15))

    return {
        "phase_pct": round(phase_pct, 1),
        "is_new_moon": is_new,
        "is_full_moon": is_full,
        "nearest_new": nearest_new,
        "nearest_full": nearest_full,
    }


# ── Data loading ──────────────────────────────────────────────────────────────
def load_data():
    wb = openpyxl.load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        game_date_raw   = raw[0]
        kickoff_raw     = raw[1]
        home_team_raw   = raw[2]
        away_team_raw   = raw[3]
        venue           = raw[4]
        home_score      = raw[5]
        away_score      = raw[6]
        is_final        = raw[7]
        total_open      = raw[39]
        total_close     = raw[42]
        over_odds_close = raw[46]
        under_odds_close = raw[50]

        if not game_date_raw or not hasattr(game_date_raw, "year"):
            continue
        if game_date_raw.year not in SEASONS:
            continue
        if home_score is None or away_score is None:
            continue
        if total_close is None or over_odds_close is None:
            continue

        game_date = game_date_raw.date() if hasattr(game_date_raw, "date") else game_date_raw
        if kickoff_raw and hasattr(kickoff_raw, "hour") and not hasattr(kickoff_raw, "date"):
            dt = datetime.combine(game_date, kickoff_raw)
        elif kickoff_raw and hasattr(kickoff_raw, "date"):
            dt = datetime.combine(game_date, kickoff_raw.time())
        else:
            dt = datetime(game_date.year, game_date.month, game_date.day, 14, 0)

        actual_total   = int(home_score) + int(away_score)
        market_line    = float(total_close)
        over_implied   = round(1 / float(over_odds_close), 4)
        under_implied  = round(1 / float(under_odds_close), 4) if under_odds_close else None
        actual_over    = 1 if actual_total > market_line else 0

        rows.append({
            "dt": dt,
            "game_date": game_date,
            "season": game_date_raw.year,
            "home_team": str(home_team_raw).strip() if home_team_raw else "",
            "away_team": str(away_team_raw).strip() if away_team_raw else "",
            "venue": str(venue).strip() if venue else "Unknown",
            "home_score": int(home_score),
            "away_score": int(away_score),
            "actual_total": actual_total,
            "market_line": market_line,
            "market_line_open": float(total_open) if total_open else market_line,
            "over_odds": float(over_odds_close),
            "over_implied": over_implied,
            "actual_over": actual_over,
            "is_final": bool(is_final),
        })

    wb.close()
    rows.sort(key=lambda r: r["dt"])
    return rows


def enrich_rows(rows):
    dates = [r["game_date"] for r in rows]
    new_moon_dates, full_moon_dates = set(), set()
    d = ephem.Date(min(dates) - timedelta(days=30))
    end_e = ephem.Date(max(dates) + timedelta(days=30))
    while d < end_e:
        nm_date = ephem.Date(ephem.next_new_moon(d)).datetime().date()
        fm_date = ephem.Date(ephem.next_full_moon(d)).datetime().date()
        for delta in range(-1, 2):
            new_moon_dates.add(nm_date + timedelta(days=delta))
            full_moon_dates.add(fm_date + timedelta(days=delta))
        d = ephem.next_new_moon(d) + 1

    for r in rows:
        dt = r["dt"]
        wd = dt.weekday()
        r["is_night"]    = dt.hour >= 17
        r["is_day"]      = dt.hour < 17
        r["is_thu_fri"]  = wd in (3, 4)
        r["is_saturday"] = wd == 5
        r["is_sunday"]   = wd == 6
        r["is_new_moon"]  = r["game_date"] in new_moon_dates
        r["is_full_moon"] = r["game_date"] in full_moon_dates

    # Form (did team win last game) and rest
    team_games_map: dict[str, list] = defaultdict(list)
    for r in rows:
        team_games_map[r["home_team"]].append(r)
        team_games_map[r["away_team"]].append(r)

    for team, games in team_games_map.items():
        games.sort(key=lambda x: x["dt"])
        for i, g in enumerate(games):
            kr = f"rest__{team}"
            kp = f"prev_win__{team}"
            if i == 0:
                g[kr] = None
                g[kp] = None
            else:
                prev = games[i - 1]
                g[kr] = (g["game_date"] - prev["game_date"]).days
                g[kp] = (prev["home_score"] > prev["away_score"]) if prev["home_team"] == team \
                        else (prev["away_score"] > prev["home_score"])

    return rows


def totals_stats(games, team):
    """Calculate over/under performance for a team under given conditions."""
    n = len(games)
    if n < MIN_SAMPLE:
        return None
    overs = sum(g["actual_over"] for g in games)
    actual_pct  = (overs / n) * 100
    implied_pct = (sum(g["over_implied"] for g in games) / n) * 100
    diff        = actual_pct - implied_pct
    rel_edge    = abs(diff / implied_pct) * 100 if implied_pct else 0
    direction   = "OVER" if diff > 0 else "UNDER"
    avg_actual  = sum(g["actual_total"] for g in games) / n
    avg_line    = sum(g["market_line"] for g in games) / n
    avg_diff    = avg_actual - avg_line
    return {
        "actual": round(actual_pct, 1),
        "implied": round(implied_pct, 1),
        "diff": round(diff, 1),
        "rel_edge": round(rel_edge, 1),
        "direction": direction,
        "n": n,
        "flag": rel_edge >= EDGE_THRESHOLD,
        "avg_actual": round(avg_actual, 1),
        "avg_line": round(avg_line, 1),
        "avg_diff": round(avg_diff, 1),
    }


def print_row(label, s):
    if s is None:
        print(f"  --     {label:<44}  n<{MIN_SAMPLE}")
        return
    flag = ">> " if s["flag"] else "   "
    edge_str = f"{s['rel_edge']:.0f}% {s['direction']}" if s["rel_edge"] >= 1 else "--"
    print(f"  {flag}  {label:<44}  over {s['actual']:5.1f}%  impl {s['implied']:5.1f}%  "
          f"diff {s['diff']:+5.1f}pp  edge {edge_str:<16}  "
          f"avg total {s['avg_actual']:.0f} vs line {s['avg_line']:.0f} ({s['avg_diff']:+.0f})  n={s['n']}")


def main():
    print("=" * 110)
    print(f"  AFL Totals Matrix -- {HOME_TEAM} vs {AWAY_TEAM}")
    print(f"  {GAME_DATE.strftime('%A %d %B %Y')} | {VENUE} | Market total: {MARKET_TOTAL}")
    print(f"  Source seasons: {SEASONS}")
    print("=" * 110)

    moon = moon_phase(GAME_DATE)
    print(f"\n  MOON PHASE (30 May 2026):")
    print(f"    Illumination:  {moon['phase_pct']}%  (nearest full: {moon['nearest_full']})")
    print(f"    Is new moon:   {moon['is_new_moon']}   Is full moon: {moon['is_full_moon']}")

    wd = GAME_TIME.weekday()
    print(f"\n  GAME CONDITIONS:")
    print(f"    Day of week:   {GAME_TIME.strftime('%A')} (saturday={wd==5})")
    print(f"    Kickoff:       {GAME_TIME.strftime('%H:%M')} AEST (night={GAME_TIME.hour >= 17})")
    print(f"    Sydney outs:   Gulden, Adams, Campbell, King (T5 loaded)")

    print("\n  Loading historical data...")
    all_rows = load_data()
    all_rows = enrich_rows(all_rows)
    print(f"  {len(all_rows)} games loaded (with totals lines).\n")

    for team in [HOME_TEAM, AWAY_TEAM]:
        role = "HOME" if team == HOME_TEAM else "AWAY"
        prev_win  = SYDNEY_PREV_WIN if team == HOME_TEAM else RICHMOND_PREV_WIN
        rest_days = SYDNEY_REST if team == HOME_TEAM else RICHMOND_REST

        g = [r for r in all_rows if r["home_team"] == team or r["away_team"] == team]

        print(f"\n{'=' * 110}")
        print(f"  {team.upper()} ({role})  -- >> = {EDGE_THRESHOLD:.0f}%+ relative edge on OVER/UNDER")
        print(f"  Metric: when [team] plays in condition, does actual total beat or miss the market line?")
        print(f"{'=' * 110}")

        flags = []

        def check(label, subset):
            s = totals_stats(subset, team)
            print_row(label, s)
            if s and s["flag"]:
                flags.append((label, s))
            return s

        print("\n  [OVERALL]")
        check("All games",         g)
        check("Home games",        [x for x in g if x["home_team"] == team])
        check("Away games",        [x for x in g if x["away_team"] == team])
        check("Regular season",    [x for x in g if not x["is_final"]])
        check("Finals",            [x for x in g if x["is_final"]])

        print("\n  [TIME OF DAY]")
        check("Night games (>=17:00)", [x for x in g if x["is_night"]])
        check("Day games (<17:00)",    [x for x in g if x["is_day"]])

        print("\n  [DAY OF WEEK]")
        check("Thu/Fri games",     [x for x in g if x["is_thu_fri"]])
        check("Saturday games",    [x for x in g if x["is_saturday"]])
        check("Sunday games",      [x for x in g if x["is_sunday"]])

        print("\n  [FORM]")
        kr = f"prev_win__{team}"
        check("After a win",       [x for x in g if x.get(kr) is True])
        check("After a loss",      [x for x in g if x.get(kr) is False])

        print("\n  [REST]")
        krest = f"rest__{team}"
        check("Short rest (<=6 days)", [x for x in g if x.get(krest) is not None and x[krest] <= 6])
        check("Normal rest (7-9d)",    [x for x in g if x.get(krest) is not None and 7 <= x[krest] <= 9])
        check("Long rest (>=10 days)", [x for x in g if x.get(krest) is not None and x[krest] >= 10])

        print("\n  [MOON PHASE]")
        check("New moon (+-1 day)",  [x for x in g if x["is_new_moon"]])
        check("Full moon (+-1 day)", [x for x in g if x["is_full_moon"]])

        print("\n  [HEAD TO HEAD]")
        opp = AWAY_TEAM if team == HOME_TEAM else HOME_TEAM
        h2h_games = [x for x in g if x["home_team"] == opp or x["away_team"] == opp]
        check(f"vs {opp}",             h2h_games)
        h2h_home = [x for x in h2h_games if x["home_team"] == team]
        h2h_away = [x for x in h2h_games if x["away_team"] == team]
        check(f"vs {opp} (as home)",   h2h_home)
        check(f"vs {opp} (as away)",   h2h_away)

        print("\n  [VENUE]")
        check(f"At {VENUE}",       [x for x in g if VENUE.lower() in (x["venue"] or "").lower()])

        print("\n  [MONTH]")
        check("May games",         [x for x in g if x["dt"].month == 5])
        check("June games",        [x for x in g if x["dt"].month == 6])

        print(f"\n  -- FLAGGED EDGES FOR {team} ({len(flags)} total) --")
        if not flags:
            print("    None at 20%+ threshold.")
        else:
            for label, s in flags:
                print(f"    >> {label:<44} over {s['actual']:.1f}% vs impl {s['implied']:.1f}%  "
                      f"({s['diff']:+.1f}pp, {s['rel_edge']:.0f}% edge, {s['direction']}, "
                      f"avg {s['avg_actual']:.0f} vs line {s['avg_line']:.0f}, n={s['n']})")

        # Applicable conditions for THIS specific game
        applicable_map = {
            "Night games (>=17:00)": GAME_TIME.hour >= 17,
            "Day games (<17:00)":    GAME_TIME.hour < 17,
            "Thu/Fri games":         wd in (3, 4),
            "Saturday games":        wd == 5,
            "Sunday games":          wd == 6,
            "After a win":           prev_win is True,
            "After a loss":          prev_win is False,
            "Short rest (<=6 days)": rest_days <= 6,
            "Normal rest (7-9d)":    7 <= rest_days <= 9,
            "Long rest (>=10 days)": rest_days >= 10,
            "Home games":            team == HOME_TEAM,
            "Away games":            team == AWAY_TEAM,
            "May games":             True,
            "Regular season":        True,
            f"vs {opp}":             True,
            f"vs {opp} (as home)":   team == HOME_TEAM,
            f"vs {opp} (as away)":   team == AWAY_TEAM,
            f"At {VENUE}":           team == HOME_TEAM,
        }
        applicable_map["New moon (+-1 day)"]  = moon["is_new_moon"]
        applicable_map["Full moon (+-1 day)"] = moon["is_full_moon"]

        applicable_flags = [(label, s) for (label, s) in flags if applicable_map.get(label, False)]

        print(f"\n  -- APPLICABLE EDGES (conditions that apply TODAY) --")
        if not applicable_flags:
            print(f"    None.")
        else:
            for label, s in applicable_flags:
                direction_str = "OVER" if s["direction"] == "OVER" else "UNDER"
                print(f"    OK  {label:<44} {s['rel_edge']:.0f}% {direction_str}  "
                      f"avg {s['avg_actual']:.0f} vs line {s['avg_line']:.0f}  n={s['n']}")

        confluence = len(applicable_flags)
        print(f"\n  CONFLUENCE COUNT: {confluence}/3+ needed")
        if confluence >= 3:
            # Check if they all agree on direction
            directions = set(s["direction"] for _, s in applicable_flags)
            if len(directions) == 1:
                print(f"  ** TRIPLE CONFLUENCE ({list(directions)[0]}) -- {team} has {confluence} applicable 20%+ edges all pointing {list(directions)[0]}!")
            else:
                print(f"  ** TRIPLE CONFLUENCE (MIXED) -- {confluence} edges but directions conflict: {directions}")
        print()


if __name__ == "__main__":
    main()
