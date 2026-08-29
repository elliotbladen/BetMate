#!/usr/bin/env python3
"""Audit published NRL pre-match H2H/handicap bets against consensus openers."""

from __future__ import annotations

import argparse
import re
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
PAGE_DATA = ROOT.parent / "lib" / "researchData.ts"
WORKBOOK = ROOT / "outputs/nrl_weekly_review/historical/latest.xlsx"


def team(value: str) -> str:
    value = re.sub(r"[^a-z]+", " ", value.lower()).strip()
    aliases = {
        "warriors": ("warriors", "new zealand", "nz"),
        "cowboys": ("cowboys", "north queensland", "north qld"),
        "dragons": ("dragons", "st george"), "souths": ("souths", "rabbitohs", "south sydney"),
        "bulldogs": ("bulldogs", "canterbury"), "roosters": ("roosters",),
        "sharks": ("sharks", "cronulla"), "titans": ("titans", "gold coast"),
        "manly": ("manly", "sea eagles"), "tigers": ("tigers",), "storm": ("storm", "melbourne"),
        "panthers": ("panthers", "penrith"), "knights": ("knights", "newcastle"),
        "broncos": ("broncos", "brisbane"), "raiders": ("raiders", "canberra"),
        "eels": ("eels", "parramatta"),
    }
    for canonical, names in aliases.items():
        if any(name in value for name in names):
            return canonical
    return value


def published_bets(path: Path) -> list[dict]:
    source = path.read_text(encoding="utf-8-sig")
    block = source.split("export const MODEL_BETS: ModelBet[] = [", 1)[1].split("];", 1)[0]
    rows = []
    for body in re.findall(r"\{\s*id:.*?\}", block, re.S):
        def text(field):
            match = re.search(rf"{field}:'([^']*)'", body)
            return match.group(1) if match else ""
        def number(field):
            match = re.search(rf"{field}:(null|-?\d+(?:\.\d+)?)", body)
            return None if not match or match.group(1) == "null" else float(match.group(1))
        rows.append({
            "id": int(number("id")), "date": text("date"), "match": text("match"),
            "market_text": text("market"), "taken_odds": number("takenPrice"),
            "result_page": text("result"),
        })
    return rows


def workbook_games(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data"]
    headers = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    rows = [dict(zip(headers, values)) for values in ws.iter_rows(min_row=3, values_only=True)]
    wb.close()
    return [r for r in rows if getattr(r.get("Date"), "year", None) == 2026]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, default=ROOT / "outputs/results/nrl_published_bets_opening_backtest_2026.csv")
    args = parser.parse_args()
    games = workbook_games(WORKBOOK)
    output = []
    for bet in published_bets(PAGE_DATA):
        market_text = bet["market_text"]
        lower = market_text.lower()
        if "live" in lower or "pyl" in lower or "multi" in lower or "1 to 12" in lower:
            continue
        kind = "h2h" if "win" in lower else ("handicap" if re.search(r"[+-]\d", market_text) else "")
        if not kind:
            continue
        names = re.split(r"\s+vs\s+", re.sub(r"\s*\([^)]*\)\s*$", "", bet["match"]), flags=re.I)
        if len(names) != 2:
            continue
        wanted = {team(x) for x in names}
        candidates = [g for g in games if {team(str(g["Home Team"])), team(str(g["Away Team"]))} == wanted]
        if not candidates:
            continue
        bet_date = date.fromisoformat(bet["date"])
        game = min(candidates, key=lambda g: abs((g["Date"].date() - bet_date).days))
        # Four early ledger dates were entered up to a week before kickoff;
        # team-pair matching resolves those unambiguously within this window.
        if abs((game["Date"].date() - bet_date).days) > 8:
            continue
        home, away = team(str(game["Home Team"])), team(str(game["Away Team"]))
        selection = team(re.sub(r"\s+(?:Win|[+-]\d.*)$", "", market_text, flags=re.I))
        side = "home" if selection == home else "away" if selection == away else ""
        if not side:
            continue
        hs, aws = float(game["Home Score"]), float(game["Away Score"])
        if kind == "h2h":
            opening_odds = float(game["Home Odds Open"] if side == "home" else game["Away Odds Open"])
            won = hs > aws if side == "home" else aws > hs
            opening_line = None
            obtained_edge = ((bet["taken_odds"] / opening_odds) - 1) * 100
        else:
            opening_line = float(game["Home Line Open"] if side == "home" else game["Away Line Open"])
            opening_odds = float(game["Home Line Odds Open"] if side == "home" else game["Away Line Odds Open"])
            taken_match = re.search(r"([+-]\d+(?:\.\d+)?)", market_text)
            taken_line = float(taken_match.group(1))
            adjusted = (hs if side == "home" else aws) + opening_line - (aws if side == "home" else hs)
            won = adjusted > 0
            obtained_edge = taken_line - opening_line
        push = kind == "handicap" and adjusted == 0
        profit_taken = 0 if push else ((bet["taken_odds"] - 1) if bet["result_page"] == "win" else -1)
        profit_open = 0 if push else ((opening_odds - 1) if won else -1)
        output.append({**bet, "market": kind, "selection": selection, "game_date": game["Date"].date(),
            "opening_line": opening_line, "opening_odds": opening_odds,
            "edge_obtained": obtained_edge, "opening_result": "push" if push else "win" if won else "loss",
            "profit_$1_taken": profit_taken, "profit_$1_at_open": profit_open})
    df = pd.DataFrame(output)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(args.out, index=False)
    print(f"Matched {len(df)} pre-match tickets")
    for market, group in df.groupby("market"):
        for profit_col in ("profit_$1_taken", "profit_$1_at_open"):
            print(market, profit_col, len(group), round(group[profit_col].sum(), 2), round(group[profit_col].mean()*100, 2))
        print(group.groupby(pd.cut(group.edge_obtained, [-float('inf'),0,2,7,float('inf')], right=False))["profit_$1_taken"].agg(['count','sum','mean']))


if __name__ == "__main__":
    main()
