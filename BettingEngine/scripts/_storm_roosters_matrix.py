"""
Matrix analysis for Melbourne Storm vs Sydney Roosters
R13 2026 — Saturday 30 May, AAMI Park, ~19:35 AEST kickoff

Checks all H2H matrix conditions. Flags any condition where either team
has 20%+ relative edge (actual win% vs market implied win%).
"""

import math
import os
from datetime import datetime, timedelta, date
from collections import defaultdict
from pathlib import Path

import ephem
import openpyxl

_ROOT = Path(__file__).resolve().parents[1]
_BETMATE = Path(os.environ.get("BETMATE_ROOT", "")) if os.environ.get("BETMATE_ROOT") else _ROOT.parent

SOURCE_PATH = str(os.environ.get("NRL_HISTORICAL_XLSX") or _BETMATE / "data/nrl/historical/latest.xlsx")
SEASONS     = (2022, 2023, 2024, 2025)
MIN_SAMPLE  = 3
EDGE_THRESHOLD = 20.0

# ── Game details ────────────────────────────────────────────────────────────
GAME_DATE   = date(2026, 5, 30)
GAME_TIME   = datetime(2026, 5, 30, 19, 35)   # ~19:35 AEST Saturday night
HOME_TEAM   = "Melbourne Storm"
AWAY_TEAM   = "Sydney Roosters"
VENUE       = "AAMI Park"

# Rest days
# Storm: played R12 2026-05-22 (Thu) → May 30 = 8 days
# Roosters: last played R11 2026-05-16 (bye R12) → May 30 = 14 days
STORM_REST   = 8    # days since last game
ROOSTERS_REST = 14  # days since last game (bye week)

# Previous result
# Storm R12: lost 20-30 to Bulldogs
# Roosters R11: lost 12-18 to Cowboys (had bye R12)
STORM_PREV_WIN    = False   # R12: Storm 20 lost to Bulldogs 30
ROOSTERS_PREV_WIN = False   # R11: Roosters 12 lost to Cowboys 18


# ── Moon phase ──────────────────────────────────────────────────────────────
def moon_phase(target: date, window: int = 1):
    d = ephem.Date(target - timedelta(days=30))
    end = ephem.Date(target + timedelta(days=30))
    new_moons, full_moons = set(), set()
    while d < end:
        nm_date = ephem.Date(ephem.next_new_moon(d)).datetime().date()
        fm_date = ephem.Date(ephem.next_full_moon(d)).datetime().date()
        for delta in range(-window, window + 1):
            new_moons.add(nm_date + timedelta(days=delta))
            full_moons.add(fm_date + timedelta(days=delta))
        d = ephem.next_new_moon(d) + 1

    observer = ephem.Observer()
    observer.date = ephem.Date(target)
    moon = ephem.Moon(observer)
    phase_pct = moon.phase

    is_new  = target in new_moons
    is_full = target in full_moons

    def next_moon(func, start):
        return ephem.Date(func(ephem.Date(start))).datetime().date()

    nearest_new  = next_moon(ephem.next_new_moon,  target - timedelta(days=15))
    nearest_full = next_moon(ephem.next_full_moon, target - timedelta(days=15))

    return {
        "phase_pct": round(phase_pct, 1),
        "is_new_moon": is_new,
        "is_full_moon": is_full,
        "nearest_new": nearest_new,
        "nearest_full": nearest_full,
    }


# ── Data loading ─────────────────────────────────────────────────────────────
TEAM_NAME_MAP = {
    "Brisbane Broncos":        "Brisbane Broncos",
    "Canberra Raiders":        "Canberra Raiders",
    "Canterbury Bulldogs":     "Canterbury Bulldogs",
    "Cronulla Sharks":         "Cronulla Sharks",
    "Dolphins":                "Dolphins",
    "Gold Coast Titans":       "Gold Coast Titans",
    "Manly Sea Eagles":        "Manly Sea Eagles",
    "Melbourne Storm":         "Melbourne Storm",
    "New Zealand Warriors":    "New Zealand Warriors",
    "Newcastle Knights":       "Newcastle Knights",
    "North QLD Cowboys":       "North QLD Cowboys",
    "Parramatta Eels":         "Parramatta Eels",
    "Penrith Panthers":        "Penrith Panthers",
    "South Sydney Rabbitohs":  "South Sydney Rabbitohs",
    "St George Dragons":       "St George Dragons",
    "Sydney Roosters":         "Sydney Roosters",
    "Wests Tigers":            "Wests Tigers",
}

def load_data():
    wb = openpyxl.load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        game_date_raw = raw[0]
        kickoff_raw   = raw[1]
        home_team_raw = raw[2]
        away_team_raw = raw[3]
        venue         = raw[4]
        home_score    = raw[5]
        away_score    = raw[6]
        home_odds_close = raw[16] if raw[16] is not None else raw[13]
        away_odds_close = raw[20] if raw[20] is not None else raw[17]

        if not game_date_raw or not hasattr(game_date_raw, "year"):
            continue
        if game_date_raw.year not in SEASONS:
            continue
        if home_score is None or away_score is None:
            continue
        if home_odds_close is None or away_odds_close is None:
            continue

        game_date = game_date_raw.date() if hasattr(game_date_raw, "date") else game_date_raw
        if kickoff_raw and hasattr(kickoff_raw, "hour") and not hasattr(kickoff_raw, "date"):
            dt = datetime.combine(game_date, kickoff_raw)
        elif kickoff_raw and hasattr(kickoff_raw, "date"):
            dt = datetime.combine(game_date, kickoff_raw.time())
        else:
            dt = datetime(game_date.year, game_date.month, game_date.day, 19, 0)

        home_team = TEAM_NAME_MAP.get(home_team_raw, home_team_raw)
        away_team = TEAM_NAME_MAP.get(away_team_raw, away_team_raw)

        rows.append({
            "dt": dt,
            "game_date": game_date,
            "season": game_date_raw.year,
            "home_team": home_team,
            "away_team": away_team,
            "venue": venue or "Unknown",
            "home_score": int(home_score),
            "away_score": int(away_score),
            "home_odds": float(home_odds_close),
            "away_odds": float(away_odds_close),
            "implied_home": round(1 / float(home_odds_close), 4),
            "implied_away": round(1 / float(away_odds_close), 4),
        })

    wb.close()
    rows.sort(key=lambda r: r["dt"])
    return rows


def enrich_rows(rows):
    dates = [r["game_date"] for r in rows]
    new_moon_dates, full_moon_dates = set(), set()
    d = ephem.Date(min(dates) - timedelta(days=30))
    end_e = ephem.Date(max(dates) + timedelta(days=30))
    while d < end_e:
        nm_date = ephem.Date(ephem.next_new_moon(d)).datetime().date()
        fm_date = ephem.Date(ephem.next_full_moon(d)).datetime().date()
        for delta in range(-1, 2):
            new_moon_dates.add(nm_date + timedelta(days=delta))
            full_moon_dates.add(fm_date + timedelta(days=delta))
        d = ephem.next_new_moon(d) + 1

    for r in rows:
        dt = r["dt"]
        wd = dt.weekday()
        r["is_night"]    = dt.hour >= 18
        r["is_day"]      = dt.hour < 18
        r["is_thu_fri"]  = wd in (3, 4)
        r["is_saturday"] = wd == 5
        r["is_sunday"]   = wd == 6
        r["is_new_moon"]  = r["game_date"] in new_moon_dates
        r["is_full_moon"] = r["game_date"] in full_moon_dates

    team_games_map: dict[str, list] = defaultdict(list)
    for r in rows:
        team_games_map[r["home_team"]].append(r)
        team_games_map[r["away_team"]].append(r)

    for team, games in team_games_map.items():
        games.sort(key=lambda x: x["dt"])
        for i, g in enumerate(games):
            kr = f"rest__{team}"
            kp = f"prev_win__{team}"
            if i == 0:
                g[kr] = None
                g[kp] = None
            else:
                prev = games[i - 1]
                g[kr] = (g["game_date"] - prev["game_date"]).days
                g[kp] = (prev["home_score"] > prev["away_score"]) if prev["home_team"] == team \
                        else (prev["away_score"] > prev["home_score"])

    return rows


def team_win(game, team):
    return game["home_score"] > game["away_score"] if game["home_team"] == team \
           else game["away_score"] > game["home_score"]


def implied_prob(game, team):
    return game["implied_home"] if game["home_team"] == team else game["implied_away"]


def stats(games, team):
    n = len(games)
    if n < MIN_SAMPLE:
        return None
    wins       = sum(1 for g in games if team_win(g, team))
    actual_pct = (wins / n) * 100
    impl_pct   = (sum(implied_prob(g, team) for g in games) / n) * 100
    diff       = actual_pct - impl_pct
    rel_edge   = abs(diff / impl_pct) * 100 if impl_pct else 0
    direction  = "BACKING" if diff > 0 else "OPPOSING"
    return {
        "actual": round(actual_pct, 1),
        "implied": round(impl_pct, 1),
        "diff": round(diff, 1),
        "rel_edge": round(rel_edge, 1),
        "direction": direction,
        "n": n,
        "flag": rel_edge >= EDGE_THRESHOLD,
    }


def print_row(label, s, team):
    if s is None:
        print(f"  {'--':5}  {label:<42}  n<{MIN_SAMPLE}")
        return
    flag = ">> " if s["flag"] else "   "
    edge_str = f"{s['rel_edge']:.0f}% {s['direction']}" if s["rel_edge"] >= 1 else "--"
    print(f"  {flag}  {label:<42}  {s['actual']:5.1f}%  impl {s['implied']:5.1f}%  "
          f"diff {s['diff']:+5.1f}pp  edge {edge_str:<20}  n={s['n']}")


def main():
    print("=" * 80)
    print(f"  NRL H2H Matrix -- {HOME_TEAM} vs {AWAY_TEAM}")
    print(f"  {GAME_DATE.strftime('%A %d %B %Y')} | {VENUE}")
    print(f"  Source seasons: {SEASONS}")
    print("=" * 80)

    moon = moon_phase(GAME_DATE)
    print(f"\n  MOON PHASE (30 May 2026):")
    print(f"    Illumination:  {moon['phase_pct']}%")
    print(f"    Is new moon:   {moon['is_new_moon']}  (nearest new: {moon['nearest_new']})")
    print(f"    Is full moon:  {moon['is_full_moon']}  (nearest full: {moon['nearest_full']})")

    wd = GAME_TIME.weekday()
    print(f"\n  GAME CONDITIONS:")
    print(f"    Day of week:   {GAME_TIME.strftime('%A')} (is_thu_fri={wd in (3,4)}, is_saturday={wd==5}, is_sunday={wd==6})")
    print(f"    Kickoff time:  {GAME_TIME.strftime('%H:%M')} AEST (is_night={GAME_TIME.hour >= 18})")
    print(f"    Storm rest:    {STORM_REST} days | prev: {'Win' if STORM_PREV_WIN else 'Loss'}")
    print(f"    Roosters rest: {ROOSTERS_REST} days | prev: {'Win' if ROOSTERS_PREV_WIN else 'Loss'}")

    print("\n  Loading historical data...")
    all_rows = load_data()
    all_rows = enrich_rows(all_rows)
    print(f"  {len(all_rows)} games loaded.\n")

    for team in [HOME_TEAM, AWAY_TEAM]:
        role = "HOME" if team == HOME_TEAM else "AWAY"
        prev_win  = STORM_PREV_WIN if team == HOME_TEAM else ROOSTERS_PREV_WIN
        rest_days = STORM_REST if team == HOME_TEAM else ROOSTERS_REST

        g = [r for r in all_rows if r["home_team"] == team or r["away_team"] == team]

        print(f"\n{'=' * 80}")
        print(f"  {team.upper()} ({role})  -- >> = {EDGE_THRESHOLD:.0f}%+ relative edge")
        print(f"{'=' * 80}")

        flags = []

        def check(label, subset):
            s = stats(subset, team)
            print_row(label, s, team)
            if s and s["flag"]:
                flags.append((label, s))
            return s

        print("\n  [OVERALL]")
        check("All games",         g)
        check("As home team",      [x for x in g if x["home_team"] == team])
        check("As away team",      [x for x in g if x["away_team"] == team])

        print("\n  [TIME OF DAY]")
        check("Night games (>=18:00)", [x for x in g if x["is_night"]])
        check("Day games (<18:00)",    [x for x in g if x["is_day"]])

        print("\n  [DAY OF WEEK]")
        check("Thu/Fri games",     [x for x in g if x["is_thu_fri"]])
        check("Saturday games",    [x for x in g if x["is_saturday"]])
        check("Sunday games",      [x for x in g if x["is_sunday"]])

        print("\n  [FORM]")
        kr = f"prev_win__{team}"
        check("After a win",       [x for x in g if x.get(kr) is True])
        check("After a loss",      [x for x in g if x.get(kr) is False])

        print("\n  [REST]")
        krest = f"rest__{team}"
        check("Short rest (<=6 days)", [x for x in g if x.get(krest) is not None and x[krest] <= 6])
        check("Long rest (>=10 days)", [x for x in g if x.get(krest) is not None and x[krest] >= 10])

        print("\n  [MOON PHASE]")
        check("New moon (+-1 day)",  [x for x in g if x["is_new_moon"]])
        check("Full moon (+-1 day)", [x for x in g if x["is_full_moon"]])

        print("\n  [HEAD TO HEAD vs OPPONENT]")
        opp = AWAY_TEAM if team == HOME_TEAM else HOME_TEAM
        h2h_games = [x for x in g if x["home_team"] == opp or x["away_team"] == opp]
        check(f"vs {opp}", h2h_games)

        print("\n  [VENUE]")
        venue_aliases = [VENUE]  # Only confirmed Melbourne Storm home venue
        for alias in venue_aliases:
            venue_games = [x for x in g if alias.lower() in (x["venue"] or "").lower()]
            if venue_games:
                check(f"At {alias}", venue_games)

        print("\n  [MONTH -- MAY]")
        check("May games",         [x for x in g if x["dt"].month == 5])

        print(f"\n  -- FLAGGED EDGES FOR {team} ({len(flags)} total) --")
        if not flags:
            print("    None at 20%+ threshold.")
        else:
            for label, s in flags:
                print(f"    >> {label:<42} actual {s['actual']:.1f}% vs implied {s['implied']:.1f}%  "
                      f"({s['diff']:+.1f}pp, {s['rel_edge']:.0f}% edge, {s['direction']}, n={s['n']})")

        # Applicable conditions for this specific game
        applicable_map = {
            "Night games (>=18:00)": GAME_TIME.hour >= 18,
            "Day games (<18:00)":    GAME_TIME.hour < 18,
            "Thu/Fri games":         wd in (3, 4),
            "Saturday games":        wd == 5,
            "Sunday games":          wd == 6,
            "After a win":           prev_win is True,
            "After a loss":          prev_win is False,
            "Short rest (<=6 days)": rest_days <= 6,
            "Long rest (>=10 days)": rest_days >= 10,
            "As home team":          team == HOME_TEAM,
            "As away team":          team == AWAY_TEAM,
            "May games":             True,
            f"vs {opp}":             True,
        }
        applicable_map["New moon (+-1 day)"]  = moon["is_new_moon"]
        applicable_map["Full moon (+-1 day)"] = moon["is_full_moon"]
        for alias in venue_aliases:
            applicable_map[f"At {alias}"] = True

        applicable_flags = [(label, s) for (label, s) in flags if applicable_map.get(label, False)]

        print(f"\n  -- APPLICABLE EDGES (conditions that apply TODAY) --")
        if not applicable_flags:
            print(f"    None -- no flagged edges apply to this specific game.")
        else:
            for label, s in applicable_flags:
                print(f"    OK  {label:<42} {s['rel_edge']:.0f}% {s['direction']}  n={s['n']}")

        confluence = len(applicable_flags)
        print(f"\n  CONFLUENCE COUNT: {confluence}/{'3+ = SIGNAL' if confluence >= 3 else '3 needed'}")
        if confluence >= 3:
            print(f"  ** TRIPLE CONFLUENCE -- {team} has {confluence} applicable 20%+ edges!")
        print()


if __name__ == "__main__":
    main()
