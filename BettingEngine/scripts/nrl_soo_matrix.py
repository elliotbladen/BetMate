"""
NRL State of Origin (SOO) Matrix Builder
Analyses NRL team performance during the State of Origin window each year (2022-2025).

During the SOO period, NSW and QLD clubs lose Origin-selected players, creating
systematic edges for some teams and systematic fade signals for others.

Source: outputs/nrl_weekly_review/historical/latest.xlsx  (or NRL_HISTORICAL_XLSX env var)
Output: outputs/nrl_soo_matrix.xlsx
        outputs/nrl_soo_matrix.csv

SOO Game Dates (verified from schedule gaps in historical data):
  2022: Game 1 Jun 8  | Game 2 Jun 26 | Game 3 Jul 13
  2023: Game 1 May 31 | Game 2 Jun 21 | Game 3 Jul 12
  2024: Game 1 May 29 | Game 2 Jun 26 | Game 3 Jul 17
  2025: Game 1 May 28 | Game 2 Jun 18 | Game 3 Jul 16

NRL games are tagged into three SOO phases per year:
  Phase 1 — Post Game 1  (day after G1 -> day before G2)
  Phase 2 — Post Game 2  (day after G2 -> day before G3)
  Phase 3 — Post Game 3  (day after G3 -> ~10 days after G3)

Metrics:
  H2H        -- actual win % vs market implied win % (closing odds)
  Handicap   -- cover rate vs 50% baseline
  Totals     -- over rate (% games > market total) + avg actual vs avg market total
"""

import csv
import os
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─────────────────────────────────────────────
#  Paths
# ─────────────────────────────────────────────

_ROOT = Path(__file__).resolve().parents[1]
_BETMATE = (
    Path(os.environ["BETMATE_ROOT"])
    if os.environ.get("BETMATE_ROOT")
    else _ROOT.parent / "BetMate"
)

SOURCE_PATH = str(
    os.environ.get("NRL_HISTORICAL_XLSX")
    or _BETMATE / "data/nrl/historical/latest.xlsx"
)
OUTPUT_XLSX = str(_ROOT / "outputs" / "nrl_soo_matrix.xlsx")
OUTPUT_CSV  = str(_ROOT / "outputs" / "nrl_soo_matrix.csv")

SEASONS       = (2022, 2023, 2024, 2025)
MIN_SAMPLE    = 3
EDGE_FLAG_PCT = 15.0

# ─────────────────────────────────────────────
#  SOO calendar
# ─────────────────────────────────────────────

SOO_GAME_DATES: dict[int, list[date]] = {
    2022: [date(2022, 6, 8),  date(2022, 6, 26), date(2022, 7, 13)],
    2023: [date(2023, 5, 31), date(2023, 6, 21), date(2023, 7, 12)],
    2024: [date(2024, 5, 29), date(2024, 6, 26), date(2024, 7, 17)],
    2025: [date(2025, 5, 28), date(2025, 6, 18), date(2025, 7, 16)],
}


def build_phase_windows(year: int) -> list[tuple[date, date, str]]:
    g1, g2, g3 = SOO_GAME_DATES[year]
    return [
        (g1 + timedelta(days=1), g2 - timedelta(days=1), "Post-G1"),
        (g2 + timedelta(days=1), g3 - timedelta(days=1), "Post-G2"),
        (g3 + timedelta(days=1), g3 + timedelta(days=10), "Post-G3"),
    ]


def soo_phase(game_date: date) -> tuple[int | None, str | None]:
    for year in SEASONS:
        for start, end, label in build_phase_windows(year):
            if start <= game_date <= end:
                return year, label
    return None, None


# ─────────────────────────────────────────────
#  Team name map
# ─────────────────────────────────────────────

TEAM_NAME_MAP = {
    "Brisbane Broncos":        "Brisbane Broncos",
    "Canberra Raiders":        "Canberra Raiders",
    "Canterbury Bulldogs":     "Canterbury Bulldogs",
    "Cronulla Sharks":         "Cronulla Sharks",
    "Dolphins":                "Dolphins",
    "Gold Coast Titans":       "Gold Coast Titans",
    "Manly Sea Eagles":        "Manly Sea Eagles",
    "Melbourne Storm":         "Melbourne Storm",
    "New Zealand Warriors":    "New Zealand Warriors",
    "Newcastle Knights":       "Newcastle Knights",
    "North QLD Cowboys":       "North QLD Cowboys",
    "Parramatta Eels":         "Parramatta Eels",
    "Penrith Panthers":        "Penrith Panthers",
    "South Sydney Rabbitohs":  "South Sydney Rabbitohs",
    "St George Dragons":       "St George Dragons",
    "Sydney Roosters":         "Sydney Roosters",
    "Wests Tigers":            "Wests Tigers",
}

# ─────────────────────────────────────────────
#  Data loading
# ─────────────────────────────────────────────

def load_data() -> list[dict]:
    wb = openpyxl.load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        game_date_raw   = raw[0]
        kickoff_raw     = raw[1]
        home_team_raw   = raw[2]
        away_team_raw   = raw[3]
        venue           = raw[4]
        home_score      = raw[5]
        away_score      = raw[6]
        home_odds_close = raw[16] if raw[16] is not None else raw[13]
        away_odds_close = raw[20] if raw[20] is not None else raw[17]
        hcap_home_close = raw[24] if raw[24] is not None else raw[21]
        total_close     = raw[40] if raw[40] is not None else raw[37]

        if not game_date_raw or not hasattr(game_date_raw, "year"):
            continue
        if game_date_raw.year not in SEASONS:
            continue
        if home_score is None or away_score is None:
            continue
        if home_odds_close is None or away_odds_close is None:
            continue

        game_date = game_date_raw.date() if hasattr(game_date_raw, "date") else game_date_raw
        soo_year, soo_phase_label = soo_phase(game_date)
        if soo_year is None:
            continue

        if kickoff_raw and hasattr(kickoff_raw, "hour") and not hasattr(kickoff_raw, "date"):
            dt = datetime.combine(game_date, kickoff_raw)
        elif kickoff_raw and hasattr(kickoff_raw, "date"):
            dt = datetime.combine(game_date, kickoff_raw.time())
        else:
            dt = datetime(game_date.year, game_date.month, game_date.day, 19, 0)

        home_team = TEAM_NAME_MAP.get(home_team_raw, home_team_raw)
        away_team = TEAM_NAME_MAP.get(away_team_raw, away_team_raw)
        total_score = int(home_score) + int(away_score)

        rows.append({
            "dt":           dt,
            "game_date":    game_date,
            "season":       soo_year,
            "soo_phase":    soo_phase_label,
            "home_team":    home_team,
            "away_team":    away_team,
            "venue":        venue or "Unknown",
            "home_score":   int(home_score),
            "away_score":   int(away_score),
            "total_score":  total_score,
            "home_odds":    float(home_odds_close),
            "away_odds":    float(away_odds_close),
            "implied_home": round(1 / float(home_odds_close), 4),
            "implied_away": round(1 / float(away_odds_close), 4),
            "hcap_home":    float(hcap_home_close) if hcap_home_close is not None else None,
            "market_total": float(total_close)      if total_close      is not None else None,
            "margin_home":  int(home_score) - int(away_score),
        })

    wb.close()
    rows.sort(key=lambda r: r["dt"])
    return rows


# ─────────────────────────────────────────────
#  Stats helpers
# ─────────────────────────────────────────────

def h2h_stats(games: list[dict], team: str) -> tuple | None:
    n = len(games)
    if n < MIN_SAMPLE:
        return None
    wins = sum(
        1 for g in games
        if (g["home_team"] == team and g["home_score"] > g["away_score"])
        or (g["away_team"] == team and g["away_score"] > g["home_score"])
    )
    actual_pct  = round(wins / n * 100, 1)
    implied_pct = round(
        sum(g["implied_home"] if g["home_team"] == team else g["implied_away"] for g in games)
        / n * 100, 1
    )
    return actual_pct, implied_pct, n


def hcap_stats(games: list[dict], team: str) -> tuple | None:
    eligible = [g for g in games if g["hcap_home"] is not None]
    n = len(eligible)
    if n < MIN_SAMPLE:
        return None
    covers = 0
    for g in eligible:
        margin = g["margin_home"] if g["home_team"] == team else -g["margin_home"]
        line   = g["hcap_home"]   if g["home_team"] == team else -g["hcap_home"]
        if margin + line > 0:
            covers += 1
    return round(covers / n * 100, 1), 50.0, n


def totals_stats(games: list[dict]) -> tuple | None:
    """Returns (avg_actual, avg_market, n) for average scoring comparison."""
    eligible = [g for g in games if g["market_total"] is not None]
    n = len(eligible)
    if n < MIN_SAMPLE:
        return None
    avg_actual = round(sum(g["total_score"]  for g in eligible) / n, 1)
    avg_market = round(sum(g["market_total"] for g in eligible) / n, 1)
    return avg_actual, avg_market, n


def over_rate_stats(games: list[dict]) -> tuple | None:
    """Returns (over_rate_pct, 50.0, avg_actual, avg_market, n).
    over_rate = % of games where actual total > market total line."""
    eligible = [g for g in games if g["market_total"] is not None]
    n = len(eligible)
    if n < MIN_SAMPLE:
        return None
    overs      = sum(1 for g in eligible if g["total_score"] > g["market_total"])
    over_rate  = round(overs / n * 100, 1)
    avg_actual = round(sum(g["total_score"]  for g in eligible) / n, 1)
    avg_market = round(sum(g["market_total"] for g in eligible) / n, 1)
    return over_rate, 50.0, avg_actual, avg_market, n


# ─────────────────────────────────────────────
#  Edge label helpers
# ─────────────────────────────────────────────

def edge_label_h2h(actual_pct: float, implied_pct: float):
    diff = actual_pct - implied_pct
    if implied_pct == 0:
        return round(diff, 1), 0.0, "", False
    edge_pct  = abs(diff / implied_pct) * 100
    direction = "over" if diff > 0 else "under"
    flag      = edge_pct >= EDGE_FLAG_PCT
    return round(diff, 1), round(edge_pct, 1), direction, flag


def edge_label_hcap(cover_rate: float):
    diff      = cover_rate - 50.0
    edge_pct  = abs(diff / 50.0) * 100
    direction = "covers" if diff > 0 else "fades"
    flag      = edge_pct >= EDGE_FLAG_PCT
    return round(diff, 1), round(edge_pct, 1), direction, flag


def edge_label_over(over_rate: float):
    """Edge vs the 50% baseline for over/under rate."""
    diff      = over_rate - 50.0
    edge_pct  = abs(diff / 50.0) * 100
    direction = "overs" if diff > 0 else "unders"
    flag      = edge_pct >= EDGE_FLAG_PCT
    return round(diff, 1), round(edge_pct, 1), direction, flag


def edge_label_avg(avg_actual: float, avg_market: float):
    diff = avg_actual - avg_market
    if avg_market == 0:
        return round(diff, 1), 0.0, "", False
    edge_pct  = abs(diff / avg_market) * 100
    direction = "overs" if diff > 0 else "unders"
    flag      = edge_pct >= EDGE_FLAG_PCT
    return round(diff, 1), round(edge_pct, 1), direction, flag


def team_games_for(all_rows: list[dict], team: str) -> list[dict]:
    return [r for r in all_rows if r["home_team"] == team or r["away_team"] == team]


# ─────────────────────────────────────────────
#  Formatting constants
# ─────────────────────────────────────────────

TITLE_FILL       = PatternFill("solid", fgColor="0D2137")
TITLE_FONT       = Font(color="FFFFFF", bold=True, size=12)
HEADER_FILL      = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT      = Font(color="FFFFFF", bold=True, size=10)
SECTION_FILL     = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT     = Font(color="FFFFFF", bold=True, size=9)
TOT_SECTION_FILL = PatternFill("solid", fgColor="1A5276")   # darker blue for totals sections
LABEL_FILL       = PatternFill("solid", fgColor="D6E4F0")
ALT_ROW_FILL     = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL       = PatternFill("solid", fgColor="FFFFFF")
BLANK_FILL       = PatternFill("solid", fgColor="F2F2F2")
FLAG_FILL        = PatternFill("solid", fgColor="6CE58D")
STRONG_FLAG_FILL = PatternFill("solid", fgColor="00FF00")
POS_FILL         = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL         = PatternFill("solid", fgColor="FFC7CE")
THIN             = Side(style="thin", color="CCCCCC")
THIN_BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER           = Alignment(horizontal="center", vertical="center")
CENTER_WRAP      = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _border_cell(cell):
    cell.border = THIN_BORDER


def style_section_row(ws, row: int, label: str, ncols: int, fill=None):
    ws.cell(row=row, column=1).value = label
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill or SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = CENTER
        _border_cell(cell)


# ─────────────────────────────────────────────
#  H2H + Handicap combined row writer
#  Columns: A=Category B=H2H% C=Impl% D=H2H Edge E=Cover% F=Hcap Edge G=N
# ─────────────────────────────────────────────

H2H_HCAP_HEADERS = [
    "Category",
    "H2H Win %", "Market Implied %", "H2H Edge",
    "Hcap Cover %", "Hcap Edge",
    "N",
]


def write_h2h_hcap_row(ws, row: int, label: str,
                        h2h: tuple | None, hcp: tuple | None,
                        alt: bool = False,
                        csv_rows: list | None = None,
                        team: str = "", section: str = "") -> None:
    bg = ALT_ROW_FILL if alt else WHITE_FILL

    lc = ws.cell(row=row, column=1, value=label)
    lc.fill = LABEL_FILL
    lc.font = Font(size=9)
    lc.alignment = Alignment(vertical="center", indent=1)
    _border_cell(lc)

    def _v(col, val, fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(size=9, bold=bold)
        c.alignment = CENTER
        c.fill = fill or bg
        c.number_format = "0.0"
        _border_cell(c)

    if h2h:
        actual, implied, _ = h2h
        diff, edge_pct, direction, flag = edge_label_h2h(actual, implied)
        edge_text = f"{edge_pct}% {direction}" if edge_pct >= 1.0 else "—"
        ff = STRONG_FLAG_FILL if (flag and edge_pct >= 30) else FLAG_FILL if flag else None
        _v(2, actual);  _v(3, implied);  _v(4, edge_text, fill=ff, bold=flag)
    else:
        for c in (2, 3, 4): _v(c, "—", fill=BLANK_FILL)

    if hcp:
        cover, _, _n = hcp
        diff, edge_pct, direction, flag = edge_label_hcap(cover)
        edge_text = f"{cover}% ({'+' if diff >= 0 else ''}{diff}pp)"
        ff = STRONG_FLAG_FILL if (flag and edge_pct >= 30) else FLAG_FILL if flag else None
        _v(5, cover);  _v(6, edge_text, fill=ff, bold=flag)
    else:
        for c in (5, 6): _v(c, "—", fill=BLANK_FILL)

    n_val = (h2h[2] if h2h else None) or (hcp[2] if hcp else None)
    _v(7, n_val or "—")

    if csv_rows is not None:
        entry = {
            "team": team, "section": section, "category": label, "type": "h2h_hcap",
            "h2h_actual": h2h[0] if h2h else "",
            "h2h_implied": h2h[1] if h2h else "",
            "h2h_edge_pct": edge_label_h2h(h2h[0], h2h[1])[1] if h2h else "",
            "h2h_direction": edge_label_h2h(h2h[0], h2h[1])[2] if h2h else "",
            "hcap_cover": hcp[0] if hcp else "",
            "hcap_edge_pct": edge_label_hcap(hcp[0])[1] if hcp else "",
            "hcap_direction": edge_label_hcap(hcp[0])[2] if hcp else "",
            "over_rate": "", "avg_actual": "", "avg_market": "",
            "totals_edge_pct": "", "totals_direction": "",
            "n": n_val or "",
        }
        csv_rows.append(entry)


# ─────────────────────────────────────────────
#  Totals row writer
#  Columns: A=Category B=Over Rate% C=Market(50%) D=Over Edge E=Avg Actual F=Avg Market G=Pts Diff H=N
# ─────────────────────────────────────────────

TOTALS_HEADERS = [
    "Category",
    "Over Rate %", "Market (50%)", "Over/Under Edge",
    "Avg Actual Total", "Avg Market Total", "Pts Diff",
    "N",
]


def write_totals_row(ws, row: int, label: str,
                     stats: tuple | None,
                     alt: bool = False,
                     col_offset: int = 0,
                     csv_rows: list | None = None,
                     team: str = "", section: str = "") -> None:
    """stats = (over_rate, 50.0, avg_actual, avg_market, n) from over_rate_stats()"""
    bg = ALT_ROW_FILL if alt else WHITE_FILL
    c0 = col_offset  # column offset so totals table can live anywhere

    lc = ws.cell(row=row, column=1 + c0, value=label)
    lc.fill = LABEL_FILL
    lc.font = Font(size=9)
    lc.alignment = Alignment(vertical="center", indent=1)
    _border_cell(lc)

    def _v(col, val, fill=None, bold=False):
        c = ws.cell(row=row, column=col + c0, value=val)
        c.font = Font(size=9, bold=bold)
        c.alignment = CENTER
        c.fill = fill or bg
        c.number_format = "0.0"
        _border_cell(c)

    if stats:
        over_rate, _, avg_actual, avg_market, n = stats
        _, edge_or, dir_or, flag_or = edge_label_over(over_rate)
        or_text = f"{edge_or}% {dir_or}" if edge_or >= 1.0 else "—"
        pts_diff = round(avg_actual - avg_market, 1)
        ff = STRONG_FLAG_FILL if (flag_or and edge_or >= 30) else FLAG_FILL if flag_or else None
        _v(2, over_rate)
        _v(3, 50.0)
        _v(4, or_text, fill=ff, bold=flag_or)
        _v(5, avg_actual)
        _v(6, avg_market)
        _v(7, f"{'+' if pts_diff >= 0 else ''}{pts_diff}")
        _v(8, n)
    else:
        for c in range(2, 9): _v(c, "—", fill=BLANK_FILL)

    if csv_rows is not None:
        entry = {
            "team": team, "section": section, "category": label, "type": "totals",
            "h2h_actual": "", "h2h_implied": "", "h2h_edge_pct": "", "h2h_direction": "",
            "hcap_cover": "", "hcap_edge_pct": "", "hcap_direction": "",
            "over_rate": stats[0] if stats else "",
            "avg_actual": stats[2] if stats else "",
            "avg_market": stats[3] if stats else "",
            "totals_edge_pct": edge_label_over(stats[0])[1] if stats else "",
            "totals_direction": edge_label_over(stats[0])[2] if stats else "",
            "n": stats[4] if stats else "",
        }
        csv_rows.append(entry)


# ─────────────────────────────────────────────
#  Per-team sheet builder
# ─────────────────────────────────────────────

def build_team_sheet(wb, team: str, all_rows: list[dict],
                     all_teams: list[str], csv_rows: list) -> None:
    g = team_games_for(all_rows, team)
    ws = wb.create_sheet(title=team[:31])

    # Column widths — 7 cols for H2H/hcap block, then gap, then 8 cols for totals
    # Layout: A-G = H2H+hcap | I-P = Totals (col H is spacer)
    widths = {
        "A": 30, "B": 12, "C": 16, "D": 20, "E": 12, "F": 22, "G": 8,
        "H": 3,
        "I": 30, "J": 12, "K": 14, "L": 22, "M": 16, "N": 16, "O": 10, "P": 8,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # ── Title spanning both blocks ──────────────────────────────
    ws.merge_cells("A1:P1")
    tc = ws.cell(row=1, column=1, value=f"{team} — NRL State of Origin Matrix (2022–2025)")
    tc.fill = TITLE_FILL
    tc.font = TITLE_FONT
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # ── H2H + Handicap block headers (cols A-G) ────────────────
    ws.merge_cells("A2:G2")
    lh = ws.cell(row=2, column=1, value="H2H & HANDICAP")
    lh.fill = HEADER_FILL
    lh.font = HEADER_FONT
    lh.alignment = CENTER

    for c, h in enumerate(H2H_HCAP_HEADERS, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        _border_cell(cell)

    # ── Totals block headers (cols I-P, offset=8) ──────────────
    ws.merge_cells("I2:P2")
    lh2 = ws.cell(row=2, column=9, value="TOTALS — OVER/UNDER")
    lh2.fill = PatternFill("solid", fgColor="145A32")
    lh2.font = HEADER_FONT
    lh2.alignment = CENTER

    for c, h in enumerate(TOTALS_HEADERS, start=9):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="145A32")
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        _border_cell(cell)

    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 30

    cur_row = 4
    current_section = [""]

    def section(label):
        nonlocal cur_row
        current_section[0] = label
        # H2H/hcap section header in cols 1-7
        style_section_row(ws, cur_row, label, ncols=7)
        # Totals section header in cols 9-16
        style_section_row(ws, cur_row, label, ncols=8,
                          fill=PatternFill("solid", fgColor="1A6B3C"))
        ws.cell(row=cur_row, column=9).value = label
        for c in range(9, 17):
            ws.cell(row=cur_row, column=c).fill = PatternFill("solid", fgColor="1A6B3C")
            ws.cell(row=cur_row, column=c).font = SECTION_FONT
            ws.cell(row=cur_row, column=c).alignment = CENTER
            _border_cell(ws.cell(row=cur_row, column=c))
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

    def row(label, games, alt=False):
        nonlocal cur_row
        write_h2h_hcap_row(
            ws, cur_row, label,
            h2h_stats(games, team), hcap_stats(games, team),
            alt=alt, csv_rows=csv_rows, team=team, section=current_section[0],
        )
        write_totals_row(
            ws, cur_row, label,
            over_rate_stats(games),
            alt=alt, col_offset=8,
            csv_rows=csv_rows, team=team, section=current_section[0],
        )
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

    # ── OVERALL SOO ─────────────────────────────────────────────
    section("OVERALL SOO PERIOD")
    row("All SOO Games",  g)
    row("Home (SOO)",     [x for x in g if x["home_team"] == team], alt=True)
    row("Away (SOO)",     [x for x in g if x["away_team"] == team])

    # ── BY SOO PHASE ─────────────────────────────────────────────
    section("BY SOO PHASE")
    row("Post-Game 1",  [x for x in g if x["soo_phase"] == "Post-G1"])
    row("Post-Game 2",  [x for x in g if x["soo_phase"] == "Post-G2"], alt=True)
    row("Post-Game 3",  [x for x in g if x["soo_phase"] == "Post-G3"])

    # ── BY YEAR ───────────────────────────────────────────────────
    section("BY YEAR")
    for i, yr in enumerate(SEASONS):
        row(str(yr), [x for x in g if x["season"] == yr], alt=(i % 2 == 1))

    # ── ROLE ──────────────────────────────────────────────────────
    section("ROLE")
    row("As Favourite (H2H)", [x for x in g if
        (x["home_team"] == team and x["implied_home"] > 0.5) or
        (x["away_team"] == team and x["implied_away"] > 0.5)])
    row("As Underdog (H2H)", [x for x in g if
        (x["home_team"] == team and x["implied_home"] <= 0.5) or
        (x["away_team"] == team and x["implied_away"] <= 0.5)], alt=True)

    # ── H2H vs OPPONENT ──────────────────────────────────────────
    section("H2H vs OPPONENT (SOO)")
    opponents = sorted(t for t in all_teams if t != team)
    for i, opp in enumerate(opponents):
        opp_games = [x for x in g if x["home_team"] == opp or x["away_team"] == opp]
        row(f"vs {opp}", opp_games, alt=(i % 2 == 1))

    ws.freeze_panes = "B4"


# ─────────────────────────────────────────────
#  Summary sheet
# ─────────────────────────────────────────────

SUMMARY_HEADERS = [
    "Team",
    "SOO Games",
    # H2H block
    "H2H Win %", "Market Impl %", "H2H Edge (pp)", "H2H Direction",
    # Handicap block
    "Hcap Cover %", "Hcap Edge",
    # Totals block
    "Over Rate %", "Avg Actual", "Avg Market", "Over/Under Edge",
    # Phase breakdown
    "Post-G1 Win %", "Post-G2 Win %", "Post-G3 Win %",
    # Over rate by phase
    "Post-G1 Over%", "Post-G2 Over%", "Post-G3 Over%",
]


def build_summary_sheet(wb, all_rows: list[dict], all_teams: list[str]) -> None:
    ws = wb.create_sheet(title="SUMMARY — All Teams", index=0)

    ncols = len(SUMMARY_HEADERS)
    col_widths = [28, 9, 11, 13, 13, 14, 12, 20, 11, 11, 11, 18, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths[:ncols], start=1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + i // 26) + chr(64 + i % 26 or 26)
        ws.column_dimensions[col_letter].width = w

    # Title
    ws.merge_cells(f"A1:R1")
    tc = ws.cell(row=1, column=1, value="NRL State of Origin Matrix — Team Rankings (2022–2025)")
    tc.fill = TITLE_FILL; tc.font = TITLE_FONT; tc.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # Sub-header group labels
    ws.merge_cells("C2:F2")
    ws.cell(row=2, column=3, value="HEAD-TO-HEAD").fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(row=2, column=3).font = Font(color="FFFFFF", bold=True, size=9)
    ws.cell(row=2, column=3).alignment = CENTER

    ws.merge_cells("G2:H2")
    ws.cell(row=2, column=7, value="HANDICAP").fill = PatternFill("solid", fgColor="2E75B6")
    ws.cell(row=2, column=7).font = Font(color="FFFFFF", bold=True, size=9)
    ws.cell(row=2, column=7).alignment = CENTER

    ws.merge_cells("I2:L2")
    ws.cell(row=2, column=9, value="TOTALS (OVER/UNDER)").fill = PatternFill("solid", fgColor="145A32")
    ws.cell(row=2, column=9).font = Font(color="FFFFFF", bold=True, size=9)
    ws.cell(row=2, column=9).alignment = CENTER

    ws.merge_cells("M2:O2")
    ws.cell(row=2, column=13, value="H2H WIN % BY PHASE").fill = PatternFill("solid", fgColor="4A235A")
    ws.cell(row=2, column=13).font = Font(color="FFFFFF", bold=True, size=9)
    ws.cell(row=2, column=13).alignment = CENTER

    ws.merge_cells("P2:R2")
    ws.cell(row=2, column=16, value="OVER RATE % BY PHASE").fill = PatternFill("solid", fgColor="1A6B3C")
    ws.cell(row=2, column=16).font = Font(color="FFFFFF", bold=True, size=9)
    ws.cell(row=2, column=16).alignment = CENTER

    ws.row_dimensions[2].height = 16

    # Column headers
    for c, h in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP; _border_cell(cell)
    ws.row_dimensions[3].height = 36

    # Compute per-team data
    team_rows = []
    for team in all_teams:
        g = team_games_for(all_rows, team)
        h2h   = h2h_stats(g, team)
        hcp   = hcap_stats(g, team)
        ors   = over_rate_stats(g)
        p1_h2h  = h2h_stats([x for x in g if x["soo_phase"] == "Post-G1"], team)
        p2_h2h  = h2h_stats([x for x in g if x["soo_phase"] == "Post-G2"], team)
        p3_h2h  = h2h_stats([x for x in g if x["soo_phase"] == "Post-G3"], team)
        p1_ors  = over_rate_stats([x for x in g if x["soo_phase"] == "Post-G1"])
        p2_ors  = over_rate_stats([x for x in g if x["soo_phase"] == "Post-G2"])
        p3_ors  = over_rate_stats([x for x in g if x["soo_phase"] == "Post-G3"])
        h2h_edge = (h2h[0] - h2h[1]) if h2h else -999
        team_rows.append((team, g, h2h, hcp, ors,
                          p1_h2h, p2_h2h, p3_h2h,
                          p1_ors, p2_ors, p3_ors, h2h_edge))

    team_rows.sort(key=lambda x: x[11], reverse=True)

    for ri, (team, g, h2h, hcp, ors,
             p1_h2h, p2_h2h, p3_h2h,
             p1_ors, p2_ors, p3_ors, h2h_edge) in enumerate(team_rows, start=4):

        alt = ri % 2 == 1
        bg  = ALT_ROW_FILL if alt else WHITE_FILL

        def _c(col, val, fill=None, bold=False):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill = fill or bg
            cell.font = Font(size=9, bold=bold)
            cell.alignment = CENTER
            cell.number_format = "0.0"
            _border_cell(cell)

        # Team name
        tc = ws.cell(row=ri, column=1, value=team)
        tc.fill = LABEL_FILL; tc.font = Font(size=9, bold=True)
        tc.alignment = Alignment(vertical="center", indent=1); _border_cell(tc)

        _c(2, len(g))

        # H2H
        if h2h:
            actual, implied, _ = h2h
            diff, edge_pct, direction, flag = edge_label_h2h(actual, implied)
            ef = POS_FILL if diff > 0 else NEG_FILL
            _c(3, actual); _c(4, implied)
            _c(5, round(diff, 1), fill=ef, bold=abs(diff) >= 5)
            _c(6, f"{edge_pct}% {direction}", fill=ef if flag else None, bold=flag)
        else:
            for c in range(3, 7): _c(c, "—", fill=BLANK_FILL)

        # Handicap
        if hcp:
            cover, _, _n = hcp
            diff, edge_pct, direction, flag = edge_label_hcap(cover)
            _c(7, cover)
            _c(8, f"{cover}% ({'+' if diff >= 0 else ''}{diff}pp)",
               fill=STRONG_FLAG_FILL if flag and edge_pct >= 30 else FLAG_FILL if flag else None,
               bold=flag)
        else:
            for c in (7, 8): _c(c, "—", fill=BLANK_FILL)

        # Totals
        if ors:
            over_rate, _, avg_actual, avg_market, _n2 = ors
            diff_or, edge_or, dir_or, flag_or = edge_label_over(over_rate)
            of = POS_FILL if diff_or > 0 else NEG_FILL
            _c(9,  over_rate, fill=of)
            _c(10, avg_actual)
            _c(11, avg_market)
            _c(12, f"{edge_or}% {dir_or}",
               fill=STRONG_FLAG_FILL if flag_or and edge_or >= 30 else FLAG_FILL if flag_or else None,
               bold=flag_or)
        else:
            for c in range(9, 13): _c(c, "—", fill=BLANK_FILL)

        # Phase H2H
        for col, stat in [(13, p1_h2h), (14, p2_h2h), (15, p3_h2h)]:
            if stat:
                diff_p = stat[0] - stat[1]
                _c(col, stat[0], fill=POS_FILL if diff_p > 0 else NEG_FILL)
            else:
                _c(col, "—", fill=BLANK_FILL)

        # Phase over rate
        for col, stat in [(16, p1_ors), (17, p2_ors), (18, p3_ors)]:
            if stat:
                diff_o = stat[0] - 50.0
                _c(col, stat[0], fill=POS_FILL if diff_o > 0 else NEG_FILL)
            else:
                _c(col, "—", fill=BLANK_FILL)

        ws.row_dimensions[ri].height = 16

    # Legend
    leg_row = 4 + len(team_rows)
    ws.merge_cells(f"A{leg_row}:R{leg_row}")
    leg = ws.cell(row=leg_row, column=1,
                  value="Sorted by H2H edge (best SOO performers to worst)  |  "
                        "Green = above market  |  Red = below market  |  "
                        "Bright green = flagged edge >=15% relative  |  "
                        "Over Rate % = % of games where actual total > market closing total line")
    leg.fill = PatternFill("solid", fgColor="E8F4FD")
    leg.font = Font(size=8, italic=True, color="444444")
    leg.alignment = CENTER
    ws.row_dimensions[leg_row].height = 14

    ws.freeze_panes = "B4"


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────

def main():
    print(f"Loading data from: {SOURCE_PATH}")
    all_rows = load_data()
    print(f"  Found {len(all_rows)} SOO-period games across {SEASONS}")

    from collections import Counter
    phase_counts: Counter = Counter()
    for r in all_rows:
        phase_counts[(r["season"], r["soo_phase"])] += 1
    print("\nGames per year/phase:")
    for yr in SEASONS:
        for phase in ("Post-G1", "Post-G2", "Post-G3"):
            n = phase_counts.get((yr, phase), 0)
            print(f"  {yr} {phase}: {n} games")

    all_teams = sorted(
        set(r["home_team"] for r in all_rows) | set(r["away_team"] for r in all_rows)
    )
    print(f"\n{len(all_teams)} teams | {len(all_rows)} total SOO games\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    csv_rows: list[dict] = []

    print("Building summary sheet...")
    build_summary_sheet(wb, all_rows, all_teams)

    print("Building per-team sheets...")
    for team in all_teams:
        n = len(team_games_for(all_rows, team))
        print(f"  {team} ({n} SOO games)...")
        build_team_sheet(wb, team, all_rows, all_teams, csv_rows)

    wb.save(OUTPUT_XLSX)
    print(f"\nSaved: {OUTPUT_XLSX}")

    csv_fields = [
        "team", "section", "category", "type",
        "h2h_actual", "h2h_implied", "h2h_edge_pct", "h2h_direction",
        "hcap_cover", "hcap_edge_pct", "hcap_direction",
        "over_rate", "avg_actual", "avg_market", "totals_edge_pct", "totals_direction",
        "n",
    ]
    with open(OUTPUT_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fields)
        writer.writeheader()
        writer.writerows(csv_rows)
    print(f"Saved: {OUTPUT_CSV}")
    print(f"\n{len(csv_rows)} rows written.")


if __name__ == "__main__":
    main()
