"""Deep totals matrix analysis for NRL R17 — reads all team sheets."""
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
wb = openpyxl.load_workbook(ROOT / 'outputs' / 'nrl_team_totals_matrix.xlsx')

# Parse every sheet (one per team)
rows_by_team = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows_by_team[sheet_name] = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 1).value
        edge = ws.cell(r, 5).value
        n_raw = ws.cell(r, 6).value
        if not val or not edge or str(edge) in ('—', '—'):
            continue
        try:
            n = int(n_raw)
        except Exception:
            n = 0
        if n >= 3:
            rows_by_team[sheet_name][str(val).strip()] = {'edge': str(edge), 'n': n}

# Sheet name aliases (matrix uses short names)
SHEET_ALIAS = {
    'Parramatta Eels':                'Parramatta Eels',
    'South Sydney Rabbitohs':         'South Sydney Rabbitohs',
    'Gold Coast Titans':              'Gold Coast Titans',
    'Canterbury-Bankstown Bulldogs':  'Canterbury Bulldogs',
    'Brisbane Broncos':               'Brisbane Broncos',
    'Sydney Roosters':                'Sydney Roosters',
    'Dolphins':                       'Dolphins',
    'New Zealand Warriors':           'New Zealand Warriors',
    'North Queensland Cowboys':       'North QLD Cowboys',
    'Penrith Panthers':               'Penrith Panthers',
    'Manly-Warringah Sea Eagles':     'Manly Sea Eagles',
    'Melbourne Storm':                'Melbourne Storm',
    'Canberra Raiders':               'Canberra Raiders',
    'St. George Illawarra Dragons':   'St George Dragons',
    'Newcastle Knights':              'Newcastle Knights',
    'Wests Tigers':                   'Wests Tigers',
    'Cronulla-Sutherland Sharks':     'Cronulla Sharks',
}

OPP_KEY = {
    'Parramatta Eels':                'vs Parramatta Eels',
    'South Sydney Rabbitohs':         'vs South Sydney Rabbitohs',
    'Gold Coast Titans':              'vs Gold Coast Titans',
    'Canterbury-Bankstown Bulldogs':  'vs Canterbury Bulldogs',
    'Brisbane Broncos':               'vs Brisbane Broncos',
    'Sydney Roosters':                'vs Sydney Roosters',
    'Dolphins':                       'vs Dolphins',
    'New Zealand Warriors':           'vs New Zealand Warriors',
    'North Queensland Cowboys':       'vs North QLD Cowboys',
    'Penrith Panthers':               'vs Penrith Panthers',
    'Manly-Warringah Sea Eagles':     'vs Manly Sea Eagles',
    'Melbourne Storm':                'vs Melbourne Storm',
    'Canberra Raiders':               'vs Canberra Raiders',
    'St. George Illawarra Dragons':   'vs St George Dragons',
    'Newcastle Knights':              'vs Newcastle Knights',
    'Wests Tigers':                   'vs Wests Tigers',
    'Cronulla-Sutherland Sharks':     'vs Cronulla Sharks',
}

VENUE_KEY = {
    'CommBank Stadium':       'CommBank Stadium',
    'Cbus Super Stadium':     'Cbus Super Stadium',
    'Suncorp Stadium':        'Suncorp Stadium',
    'QCB Stadium':            'QCB Stadium',
    '4 Pines Park':           '4 Pines Park',
    'GIO Stadium':            'GIO Stadium',
    'McDonald Jones Stadium': 'McDonald Jones Stadium',
}

# Moon: Full moon Jun 30. R17 Jun 25-28 = 2-5 days before full. ±1 day = only Jun 29-Jul 1.
# No game falls within ±1 day of full moon.
# New moon Jun 15 — too far (10-13 days ago). No moon signal applies.

GAMES = [
    {
        'label':        'Parramatta vs South Sydney',
        'home':         'Parramatta Eels',
        'away':         'South Sydney Rabbitohs',
        'venue':        'CommBank Stadium',
        'day':          'Thursday / Friday Games',
        'time':         'Night Games (kick-off',
        'home_rest':    'Long Rest',
        'away_rest':    'Long Rest',
        'home_form':    'After a Win',
        'away_form':    'After a Win',
        'month':        'June',
        'model_total':  53.7,
        'ref_bucket':   'whistle_heavy',
        'ref_delta':    -1.06,
    },
    {
        'label':        'Gold Coast vs Canterbury',
        'home':         'Gold Coast Titans',
        'away':         'Canterbury-Bankstown Bulldogs',
        'venue':        'Cbus Super Stadium',
        'day':          'Thursday / Friday Games',
        'time':         'Night Games (kick-off',
        'home_rest':    'Short Rest',
        'away_rest':    'Short Rest',
        'home_form':    'After a Win',
        'away_form':    'After a Win',
        'month':        'June',
        'model_total':  35.7,
        'ref_bucket':   'neutral',
        'ref_delta':    +0.43,
    },
    {
        'label':        'Brisbane vs Roosters',
        'home':         'Brisbane Broncos',
        'away':         'Sydney Roosters',
        'venue':        'Suncorp Stadium',
        'day':          'Thursday / Friday Games',
        'time':         'Night Games (kick-off',
        'home_rest':    'Long Rest',
        'away_rest':    'Short Rest',
        'home_form':    'After a Loss',
        'away_form':    'After a Win',
        'month':        'June',
        'model_total':  45.9,
        'ref_bucket':   'neutral',
        'ref_delta':    -0.29,
    },
    {
        'label':        'Dolphins vs Warriors',
        'home':         'Dolphins',
        'away':         'New Zealand Warriors',
        'venue':        'Suncorp Stadium',
        'day':          'Saturday Games',
        'time':         'Day Games (kick-off',
        'home_rest':    None,
        'away_rest':    'Short Rest',
        'home_form':    'After a Win',
        'away_form':    'After a Win',
        'month':        'June',
        'model_total':  49.4,
        'ref_bucket':   'neutral',
        'ref_delta':    -0.07,
    },
    {
        'label':        'Cowboys vs Panthers',
        'home':         'North Queensland Cowboys',
        'away':         'Penrith Panthers',
        'venue':        'QCB Stadium',
        'day':          'Saturday Games',
        'time':         'Day Games (kick-off',
        'home_rest':    'Short Rest',
        'away_rest':    None,
        'home_form':    'After a Loss',
        'away_form':    'After a Loss',
        'month':        'June',
        'model_total':  51.9,
        'ref_bucket':   'flow_heavy',
        'ref_delta':    +3.13,
    },
    {
        'label':        'Manly vs Storm',
        'home':         'Manly-Warringah Sea Eagles',
        'away':         'Melbourne Storm',
        'venue':        '4 Pines Park',
        'day':          'Saturday Games',
        'time':         'Night Games (kick-off',
        'home_rest':    None,
        'away_rest':    'Short Rest',
        'home_form':    'After a Loss',
        'away_form':    'After a Win',
        'month':        'June',
        'model_total':  45.9,
        'ref_bucket':   'whistle_heavy',
        'ref_delta':    -1.18,
    },
    {
        'label':        'Raiders vs Dragons',
        'home':         'Canberra Raiders',
        'away':         'St. George Illawarra Dragons',
        'venue':        'GIO Stadium',
        'day':          'Sunday Games',
        'time':         'Day Games (kick-off',
        'home_rest':    None,
        'away_rest':    None,
        'home_form':    'After a Loss',
        'away_form':    'After a Loss',
        'month':        'June',
        'model_total':  46.5,
        'ref_bucket':   'neutral',
        'ref_delta':    0.00,
    },
    {
        'label':        'Knights vs Tigers',
        'home':         'Newcastle Knights',
        'away':         'Wests Tigers',
        'venue':        'McDonald Jones Stadium',
        'day':          'Sunday Games',
        'time':         'Day Games (kick-off',
        'home_rest':    'Long Rest',
        'away_rest':    'Long Rest',
        'home_form':    'After a Win',
        'away_form':    'After a Loss',
        'month':        'June',
        'model_total':  60.3,
        'ref_bucket':   'flow_heavy',
        'ref_delta':    +0.99,
    },
]


def get_team_data(team_name):
    alias = SHEET_ALIAS.get(team_name, team_name)
    return rows_by_team.get(alias, {})


def find_row(tm, keyword):
    """Find first row whose label contains keyword (case-insensitive)."""
    kw = keyword.lower()
    for k, v in tm.items():
        if kw in k.lower():
            return k, v
    return None, None


def parse_edge(edge_str):
    if not edge_str or edge_str == '—':
        return 0.0, None
    s = edge_str.lower()
    try:
        pct = float(s.split('%')[0].strip())
    except Exception:
        return 0.0, None
    direction = 'overs' if 'overs' in s else 'unders'
    return pct, direction


print()
print('=' * 100)
print('  NRL R17 — TOTALS MATRIX FULL CONTEXT ANALYSIS')
print('  Moon: Full Moon Jun 30. No R17 game is within ±1 day — moon signal N/A this round.')
print('  June context: most teams show an UNDERS lean in June.')
print('=' * 100)

summary = []

for g in GAMES:
    home = g['home']
    away = g['away']
    edges = []

    for team, role in [(home, 'HOME'), (away, 'AWAY')]:
        tm = get_team_data(team)
        short = SHEET_ALIAS.get(team, team)

        # vs Opponent
        opp = away if role == 'HOME' else home
        opp_k = OPP_KEY.get(opp, '')
        lbl, row = find_row(tm, opp_k)
        if row:
            pct, d = parse_edge(row['edge'])
            if pct > 0:
                edges.append((role, short, opp_k, pct, d, row['n']))

        # Day of week
        lbl, row = find_row(tm, g['day'].split(' Games')[0])
        if row:
            pct, d = parse_edge(row['edge'])
            if pct > 0:
                edges.append((role, short, lbl, pct, d, row['n']))

        # Time of day
        lbl, row = find_row(tm, g['time'])
        if row:
            pct, d = parse_edge(row['edge'])
            if pct > 0:
                edges.append((role, short, lbl, pct, d, row['n']))

        # Rest
        rest_key = g.get(f'{role.lower()}_rest')
        if rest_key:
            lbl, row = find_row(tm, rest_key)
            if row:
                pct, d = parse_edge(row['edge'])
                if pct > 0:
                    edges.append((role, short, lbl, pct, d, row['n']))

        # Form
        form_key = g[f'{role.lower()}_form']
        lbl, row = find_row(tm, form_key)
        if row:
            pct, d = parse_edge(row['edge'])
            if pct > 0:
                edges.append((role, short, lbl, pct, d, row['n']))

        # Month
        lbl, row = find_row(tm, g['month'])
        if row:
            pct, d = parse_edge(row['edge'])
            if pct > 0:
                edges.append((role, short, lbl, pct, d, row['n']))

        # Venue (home team only)
        if role == 'HOME':
            venue_k = VENUE_KEY.get(g['venue'], g['venue'])
            lbl, row = find_row(tm, venue_k)
            if row:
                pct, d = parse_edge(row['edge'])
                if pct > 0:
                    edges.append((role, short, lbl, pct, d, row['n']))

        # Moon — N/A this round (no game ±1 day of full moon Jun 30)

    overs_w = sum(p for _, _, _, p, d, _ in edges if d == 'overs')
    unders_w = sum(p for _, _, _, p, d, _ in edges if d == 'unders')
    overs_n = sum(1 for _, _, _, p, d, _ in edges if d == 'overs')
    unders_n = sum(1 for _, _, _, p, d, _ in edges if d == 'unders')
    signal = 'OVERS' if overs_w >= unders_w else 'UNDERS'
    net = abs(overs_w - unders_w)

    summary.append((g['label'], g['model_total'], g['ref_delta'], g['ref_bucket'], signal, overs_w, unders_w, overs_n, unders_n, net))

    print()
    print(f'  {"─"*96}')
    print(f'  {g["label"]}  |  Model total: {g["model_total"]}  |  Ref ({g["ref_bucket"]}): {g["ref_delta"]:+.2f}')
    if not edges:
        print(f'    No contextual edges found (teams may have thin sample sizes in these conditions)')
    else:
        print(f'  Matrix signal: {signal}  |  Overs: {overs_w:.1f}% ({overs_n})  |  Unders: {unders_w:.1f}% ({unders_n})  |  Net: {net:.1f}%')
        print()
        for _, short, ctx, pct, d, n in sorted(edges, key=lambda x: -x[3]):
            arrow = 'OVER' if d == 'overs' else 'UNDR'
            sym = '▲' if d == 'overs' else '▼'
            print(f'    {sym} [{arrow}]  {pct:5.1f}%   {short}: {ctx}   (n={n})')

print()
print('=' * 100)
print('  SUMMARY — Ranked by net matrix weight')
print('=' * 100)
print(f'  {"Game":<30} {"Model":>7} {"RefΔ":>6}  {"Ref type":<14}  {"Signal":>7}  {"Net":>5}%  {"Overs":>7}  {"Unders":>7}')
print(f'  {"─"*95}')
summary.sort(key=lambda x: -x[9])
for label, tot, rd, rb, sig, op, up, on, un, net in summary:
    s_sym = '▲' if sig == 'OVERS' else '▼'
    print(f'  {label:<30} {tot:>7.1f} {rd:>+6.2f}  {rb:<14}  {s_sym} {sig:<7}  {net:>5.1f}%  {op:>5.1f}%({on})  {up:>5.1f}%({un})')
