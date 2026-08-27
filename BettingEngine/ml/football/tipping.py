"""EPL tipping decisions built from model prices and shrunk matrix evidence.

The pricing model remains the probability owner.  Historical matrix rows are
overlapping diagnostics, so they are reliability-shrunk, lightly weighted and
capped.  Contest strategy is a final decision layer and never rewrites prices.
"""
from __future__ import annotations

import csv
import math
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

OUTCOMES = ("H", "D", "A")


@dataclass(frozen=True)
class PoolRules:
    correct_points: float = 1.0
    draw_points: float | None = None
    leverage: float = 0.0
    matrix_weight: float = 0.20
    matrix_cap_pp: float = 3.0


def _normalise(values: dict[str, float]) -> dict[str, float]:
    clean = {key: max(0.000001, float(values[key])) for key in OUTCOMES}
    total = sum(clean.values())
    return {key: value / total for key, value in clean.items()}


def _number(value: Any) -> float | None:
    try:
        parsed = float(value)
        return parsed if math.isfinite(parsed) else None
    except (TypeError, ValueError):
        return None


def _matching_matrix_rows(sheet: Any, played: date, venue: str) -> list[dict[str, Any]]:
    wanted = {
        ("OVERALL", "All games"): 0.20,
        ("OVERALL", f"{venue} games"): 0.45,
        ("DAY OF WEEK", played.strftime("%A")): 0.10,
        ("MONTH", played.strftime("%B")): 0.10,
    }
    rows: list[dict[str, Any]] = []
    for raw in sheet.iter_rows(min_row=3, values_only=True):
        key = (str(raw[0] or ""), str(raw[1] or ""))
        if key not in wanted:
            continue
        win_edge, draw_edge, sample = _number(raw[4]), _number(raw[7]), _number(raw[8])
        if win_edge is None or draw_edge is None or sample is None:
            continue
        shrink = sample / (sample + 30.0)
        rows.append({"section": key[0], "category": key[1], "win_edge_pp": win_edge,
                     "draw_edge_pp": draw_edge, "n": int(sample),
                     "weight": wanted[key] * shrink})
    return rows


def matrix_adjustment(workbook: Path | None, home: str, away: str,
                      played: date, cap_pp: float = 3.0) -> tuple[dict[str, float], list[dict]]:
    zero = {key: 0.0 for key in OUTCOMES}
    if workbook is None or not workbook.exists():
        return zero, []
    import openpyxl
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    try:
        if home not in book.sheetnames or away not in book.sheetnames:
            return zero, []
        home_rows = _matching_matrix_rows(book[home], played, "Home")
        away_rows = _matching_matrix_rows(book[away], played, "Away")
    finally:
        book.close()

    def weighted(rows: list[dict], field: str) -> float:
        denominator = sum(row["weight"] for row in rows)
        return sum(row[field] * row["weight"] for row in rows) / denominator if denominator else 0.0

    # Each team's win residual informs its own side. Draw evidence is averaged;
    # it is the same match outcome and must not be counted twice.
    raw = {
        "H": weighted(home_rows, "win_edge_pp"),
        "D": (weighted(home_rows, "draw_edge_pp") + weighted(away_rows, "draw_edge_pp")) / 2,
        "A": weighted(away_rows, "win_edge_pp"),
    }
    clipped = {key: max(-cap_pp, min(cap_pp, value)) / 100.0 for key, value in raw.items()}
    evidence = [{"team": home, **row} for row in home_rows] + [{"team": away, **row} for row in away_rows]
    return clipped, evidence


def tip_match(row: dict[str, Any], matrix_workbook: Path | None = None,
              rules: PoolRules = PoolRules(), crowd: dict[str, float] | None = None) -> dict[str, Any]:
    played = datetime.strptime(str(row["date"]), "%Y-%m-%d").date()
    base = _normalise({"H": row["normal_p_home"], "D": row["normal_p_draw"], "A": row["normal_p_away"]})
    shadow = _normalise({"H": row["shadow_p_home"], "D": row["shadow_p_draw"], "A": row["shadow_p_away"]})
    adjustment, evidence = matrix_adjustment(matrix_workbook, str(row["home"]), str(row["away"]), played,
                                             rules.matrix_cap_pp)
    final = _normalise({key: base[key] + rules.matrix_weight * adjustment[key] for key in OUTCOMES})
    accuracy_pick = max(OUTCOMES, key=lambda key: final[key])
    points = {"H": rules.correct_points, "D": rules.draw_points or rules.correct_points,
              "A": rules.correct_points}
    crowd_probs = _normalise(crowd) if crowd else None
    utility = {}
    for key in OUTCOMES:
        leverage = (max(0.02, 1.0 - crowd_probs[key]) ** rules.leverage) if crowd_probs else 1.0
        utility[key] = final[key] * points[key] * leverage
    strategy_pick = max(OUTCOMES, key=lambda key: utility[key])
    normal_pick, shadow_pick = max(OUTCOMES, key=base.get), max(OUTCOMES, key=shadow.get)
    confidence = "strong" if final[strategy_pick] >= 0.60 else "medium" if final[strategy_pick] >= 0.45 else "low"
    return {
        "date": played.isoformat(), "home": row["home"], "away": row["away"],
        "probabilities": {key: round(final[key], 6) for key in OUTCOMES},
        "base_probabilities": {key: round(base[key], 6) for key in OUTCOMES},
        "matrix_adjustment_pp": {key: round(adjustment[key] * 100, 3) for key in OUTCOMES},
        "matrix_evidence": evidence, "accuracy_pick": accuracy_pick,
        "strategy_pick": strategy_pick, "strategy_utility": {key: round(utility[key], 6) for key in OUTCOMES},
        "confidence": confidence, "normal_shadow_agree": normal_pick == shadow_pick,
        "normal_pick": normal_pick, "shadow_pick": shadow_pick,
        "warning": None if normal_pick == shadow_pick else "Production and shadow models disagree; do not force a contrarian pick.",
    }


def tip_round(pricing_csv: Path, matrix_workbook: Path | None = None,
              rules: PoolRules = PoolRules()) -> list[dict[str, Any]]:
    with pricing_csv.open(newline="", encoding="utf-8-sig") as handle:
        return [tip_match(row, matrix_workbook, rules) for row in csv.DictReader(handle)]
