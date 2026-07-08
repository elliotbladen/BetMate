import openpyxl

afl_path = r"C:\Users\ElliotBladen\Apps\BettingEngine\outputs\afl_weekly_review\historical\latest.xlsx"
wb = openpyxl.load_workbook(afl_path, read_only=True, data_only=True)
ws = wb.active
# Show first few rows
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True)):
    print(f"Row {i+1}:", row[:10])
# Show a 2026 row
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and row[0] and "2026" in str(row[0]):
        print("2026 row:", row[:10])
        count += 1
        if count >= 2:
            break
wb.close()
