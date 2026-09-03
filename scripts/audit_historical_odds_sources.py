"""Summarise the four historical odds sources used by BetMate."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]


def workbook_summary(path: Path) -> dict:
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    next(rows)  # provider/source banner row
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    date_index = next((i for i, name in enumerate(headers) if name.lower() == "date"), 0)
    records = [row for row in rows if any(value is not None for value in row)]
    dates = [pd.to_datetime(row[date_index], errors="coerce", dayfirst=True)
             for row in records if row[date_index] is not None]
    dates = [value for value in dates if not pd.isna(value)]
    odds_columns = [name for name in headers if any(token in name.lower() for token in
                    ("odds", "line", "total", "opening", "closing", "maximum", "minimum"))]
    odds_indexes = {name: index for index, name in enumerate(headers) if name in odds_columns}
    return {
        "path": str(path.relative_to(ROOT)), "sheet": sheet.title,
        "rows": len(records), "date_min": str(min(dates)) if dates else None,
        "date_max": str(max(dates)) if dates else None, "odds_columns": odds_columns,
        "odds_non_null": {name: sum(row[index] is not None for row in records)
                          for name, index in odds_indexes.items()},
        "sheets": book.sheetnames,
    }


def csv_summary(path: Path) -> dict:
    frame = pd.read_csv(path, low_memory=False)
    odds_prefixes = ("B365", "Max", "PS", "PC", "AH", "PAH", "PCA")
    odds_columns = [column for column in frame.columns if column.startswith(odds_prefixes)]
    return {
        "path": str(path.relative_to(ROOT)), "rows": len(frame),
        "seasons": int(frame["Season"].nunique()),
        "date_min": str(frame["Date"].min()), "date_max": str(frame["Date"].max()),
        "odds_columns": odds_columns,
        "odds_non_null": {column: int(frame[column].notna().sum()) for column in odds_columns},
    }


def main() -> None:
    report = {
        "afl": workbook_summary(ROOT / "data/afl/historical/raw/afl_20260901.xlsx"),
        "nrl": workbook_summary(ROOT / "data/nrl/historical/raw/nrl_20260901.xlsx"),
        "epl": csv_summary(ROOT / "BettingEngine/ml/football/data/epl/matches/epl_matches.csv"),
        "championship": csv_summary(ROOT / "BettingEngine/ml/football/data/championship/matches/championship_matches.csv"),
    }
    output = ROOT / "reports/data/historical_odds_sources_2026-09-01.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    print(json.dumps(report, indent=2, default=str))


if __name__ == "__main__":
    main()
