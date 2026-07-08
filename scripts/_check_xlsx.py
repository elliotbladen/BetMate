import openpyxl, os

# Check NRL xlsx structure
nrl_path = r"C:\Users\ElliotBladen\Apps\data\nrl\historical\latest.xlsx"
wb = openpyxl.load_workbook(nrl_path, read_only=True, data_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("NRL headers:", headers)
# Show a few 2026 rows
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and str(row[0] if row[0] else "").startswith("2026"):
        print(row)
        count += 1
        if count >= 3:
            break
wb.close()
