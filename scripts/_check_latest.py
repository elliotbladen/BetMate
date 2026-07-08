import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r"data\nrl\historical\latest.xlsx", read_only=True, data_only=True)
ws = wb.active
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0] and isinstance(r[0], datetime)]
print("NRL latest:", max(r[0] for r in rows).strftime("%Y-%m-%d"), "| total rows:", len(rows))
for r in sorted(rows, key=lambda x: x[0])[-5:]:
    print(" ", r[0].strftime("%Y-%m-%d"), r[2], "vs", r[3], "|", r[5], "-", r[6])
wb.close()

wb2 = openpyxl.load_workbook(r"BettingEngine\outputs\afl_weekly_review\historical\latest.xlsx", read_only=True, data_only=True)
ws2 = wb2.active
rows2 = [r for r in ws2.iter_rows(min_row=3, values_only=True) if r and r[0] and isinstance(r[0], datetime)]
print("AFL latest:", max(r[0] for r in rows2).strftime("%Y-%m-%d"), "| total rows:", len(rows2))
for r in sorted(rows2, key=lambda x: x[0])[-5:]:
    print(" ", r[0].strftime("%Y-%m-%d"), r[2], "vs", r[3], "|", r[5], "-", r[6])
wb2.close()
